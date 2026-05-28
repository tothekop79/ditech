"""
Event Analytics Dashboard — Engine v4
═══════════════════════════════════════════════════════════
Template ที่ใช้ร่วมกับ event_config.py

Sheets:
  1. Overall Summary      — KPIs + hourly + zone + demographics
  2. Per-day dashboards — one sheet per event day
  3. Dwell Time Analysis  — per-zone dwell with distributions
  4. Activity Analytics   — activity traffic + zone matrix
  5. Raw Data Day 1/2/3   — filterable source data

วิธีใช้:
  1. แก้ event_config.py ให้ตรงกับงานของคุณ
  2. รัน: python3 dashboard_engine.py
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
import warnings
import os
warnings.filterwarnings('ignore')
try:
    from insight_engine import generate_all_insights as _gen_insights
    _HAS_INSIGHTS = True
except ImportError:
    _HAS_INSIGHTS = False


# ── Load configuration ─────────────────────────────────────────
# Priority: _config sheet in Rawdata Excel → fallback to event_config.py
import os as _os, glob as _glob
def _find_rawdata():
    """
    Find the Rawdata Excel file that contains a _config sheet.
    Prefers files named Rawdata*.xlsx over others.
    Returns None if no suitable file found (engine falls back to event_config.py).
    """
    _d = _os.path.dirname(_os.path.abspath(__file__))
    import pandas as _pd
    # Check Rawdata*.xlsx first (most likely the right file)
    for f in sorted(_glob.glob(_os.path.join(_d, 'Rawdata*.xlsx'))):
        try:
            if '_config' in _pd.ExcelFile(f).sheet_names:
                return f
        except Exception: pass
    # Then check any other xlsx
    for f in sorted(_glob.glob(_os.path.join(_d, '*.xlsx'))):
        try:
            if '_config' in _pd.ExcelFile(f).sheet_names:
                return f
        except Exception: pass
    return None
_RAWDATA_FILE = _find_rawdata()
_USE_EXCEL_CONFIG = _RAWDATA_FILE is not None
if not _USE_EXCEL_CONFIG:
    from event_config import (
        EVENT_NAME, ORGANIZER, VENUE, SYSTEM_CREDIT, CONFIDENTIAL,
        INPUT_FILES, OUTPUT_XLSX, OUTPUT_HTML,
        EVENT_DATES, EVENT_LABELS, DAY_COLORS,
        ENTRANCE_GATES, PASSERBY_GATES, ZONE_ORDER, ZONE_ABBREV,
        ACTIVITIES, DWELL_MIN_SEC, DWELL_MAX_SEC,
        ENGAGEMENT_THRESHOLD_SEC, DISPLAY_HOURS_START, DISPLAY_HOURS_END,
        PAGE_GROUPS, COL_NAMES,
    )
COL_NAMES  = ['No','unid','VideoId','BodyID','PersonnelNo','CustomerType',
              'AgeGroup','Gender','Event','CameraID','Time','Location']
PAGE_GROUPS = [[1,7,5,6],[8],[3],[2],[4]]
VENUE_TYPE     = 'Booth'   # 'Booth' or 'Event' — affects page titles & labels
SPONSOR_ZONES  = ['Main Stage']  # list of zones to render sponsor value for
SHOW_PASSERBY  = True      # True = show PSB data; False = hide even if has_psb=True

# ── Profile settings ──────────────────────────────────────────
EVENT_PROFILE = 'full'   # simple / standard / full — overridden in main()

# ── Staff exclusion ───────────────────────────────────────────
# True = exclude CustomerType='Staff' rows from unique/dwell metrics
# Overridden in main() from cfg['EXCLUDE_STAFF']
EXCLUDE_STAFF = True

def _non_staff(frame):
    """Filter out staff rows. Used by uv_count() and dwell computations.

    Behavior:
      - If EXCLUDE_STAFF=False: return frame unchanged
      - If 'CustomerType' column missing: return frame unchanged
        (graceful degradation for old rawdata without the column)
      - Otherwise: filter rows where CustomerType != 'Staff'
    """
    if not EXCLUDE_STAFF:
        return frame
    if 'CustomerType' not in frame.columns:
        return frame
    return frame[frame['CustomerType'] != 'Staff']

def uv_count(frame):
    """Unique BodyID count, staff-aware. Replaces frame['BodyID'].nunique()
    everywhere we want 'unique customers' rather than 'unique people'."""
    return _non_staff(frame)['BodyID'].nunique()

# ── Dwell time benchmark by zone (behavior-based) ─────────────
# Per-zone target dwell (seconds). Populated from _config Section D.
# SHOW_DWELL_BENCHMARK toggles the dashboard table. Both set in main().
SHOW_DWELL_BENCHMARK = False
ZONE_BENCHMARK = {}   # zone_name -> benchmark seconds
ZONE_DESC = {}        # zone_name -> human description
ZONE_BENCHMARK_MODE = {}  # zone_name -> 'higher_better' | 'lower_better'




# ══════════════════════════════════════════════════════════════
# AUTO-PAGINATION ENGINE v2 (Sprint 4)
# Adapts PAGE_GROUPS based on actual data volume to fit A4 portrait
# ══════════════════════════════════════════════════════════════
A4_CONTENT_MM = 245  # 297 - 20 margin - 22 header - 10 footer

def estimate_section_height_mm(section_id: int,
                                n_days: int,
                                n_gates: int,
                                n_zones: int,
                                n_activities: int,
                                profile: str,
                                has_psb: bool) -> int:
    """Estimate rendered height in mm for a given section."""
    if section_id == 0:  # Executive Summary
        # big KPIs + programme score + 3 highlights + 5 insights
        return 55 + 35 + 45 + 60  # ≈ 195mm — fits A4
    elif section_id == 1:  # KPIs + traffic insights
        return 40 + 25  # KPI row + insight panel

    elif section_id == 7:  # Heatmap
        return 50 + n_days * 6

    elif section_id == 5:  # Demographics
        if profile == 'simple':
            return 30 + 110  # KPIs + per-day side-by-side columns
        else:
            return 30 + 80   # KPIs + overall gender+age tables

    elif section_id == 6:  # Gate Breakdown
        base = 30
        return base + n_gates * 5 + (5 if has_psb else 0)

    elif section_id == 8:  # Hourly per day
        return 15 + n_days * 90

    elif section_id == 2:  # Zone Summary + extras
        funnel_h  = 45 if has_psb else 0
        ranking_h = 25 + min(n_zones, 5) * 12
        sponsor_h = 65
        zone_tbl_h = 25 + n_zones * 5
        zone_hm_h  = 55 + n_days * 5
        return funnel_h + ranking_h + sponsor_h + zone_tbl_h + zone_hm_h

    elif section_id == 3:  # Activity — many sub-blocks
        corr_h       = 15 + n_activities * 16
        impact_h     = 40
        bda_tbl_h    = 25 + n_activities * 5.5
        detail_tbl_h = 25 + n_activities * 5.5
        matrix_h     = 45 + min(n_activities, 20) * 4
        return int(corr_h + impact_h + bda_tbl_h + detail_tbl_h + matrix_h)

    elif section_id == 4:  # Dwell
        return 50 + n_zones * 5 + 40

    return 50  # fallback


def auto_paginate(base_groups: list,
                  n_days: int, n_gates: int, n_zones: int,
                  n_activities: int, profile: str,
                  has_psb: bool, verbose: bool = True) -> list:
    """
    Greedy bin-packs sections into A4-sized pages.

    - If a section alone exceeds A4, it gets its own page + CSS handles overflow
    - Small sections merge together
    - Always preserves section order
    """
    all_sections = [sec for grp in base_groups for sec in grp]

    heights = {
        s: estimate_section_height_mm(s, n_days, n_gates, n_zones,
                                       n_activities, profile, has_psb)
        for s in all_sections
    }

    pages = []
    current = []
    current_h = 0

    for sec in all_sections:
        h = heights[sec]

        if h > A4_CONTENT_MM:
            # Single section overflows — give it its own page
            if current:
                pages.append(current)
                current = []
                current_h = 0
            pages.append([sec])
            if verbose:
                print(f"  ⚠  Section {sec} estimated {h}mm > A4 ({A4_CONTENT_MM}mm) — "
                      f"will flow across multiple printed pages")
            continue

        if current_h + h > A4_CONTENT_MM and current:
            pages.append(current)
            current = [sec]
            current_h = h
        else:
            current.append(sec)
            current_h += h

    if current:
        pages.append(current)

    return pages


PROFILE_CONFIG = {
    'simple': {
        'page_groups':  [[0],[1,7,6],[5,8]],
        'has_psb':      False,
        'has_zones':    False,
        'has_dwell':    False,
        'has_activity': False,
        'desc': 'Simple — ทางเข้าอย่างเดียว (ENT only)',
    },
    'standard': {
        'page_groups':  [[0],[1,7],[5,6],[8],[2],[4]],
        'has_psb':      True,
        'has_zones':    True,
        'has_dwell':    True,
        'has_activity': False,
        'desc': 'Standard — มีทางเข้า + Zone (ไม่มีกิจกรรม)',
    },
    'full': {
        'page_groups':  [[0],[1,7],[5,6],[8],[3],[2],[4]],
        'has_psb':      True,
        'has_zones':    True,
        'has_dwell':    True,
        'has_activity': True,
        'desc': 'Full — ครบทุกอย่าง (Zone + Activity)',
    },
}
# ZONE_TYPES, DAY_CSS_XL, OUTPUT, and all config vars are set in main()
ZONE_TYPES = {}; OUTPUT = ''; DAY_CSS_XL = []


# ══════════════════════════════════════════════════════════════════
# CONFIGURATION
# ══════════════════════════════════════════════════════════════════
# Script directory — input file must be in the same folder as this .py file
_DIR = os.path.dirname(os.path.abspath(__file__))

# ACTIVITIES, EVENT_DATES, EVENT_LABELS, ZONE_ORDER — from event_config.py

# ══════════════════════════════════════════════════════════════════
# COLOR / STYLE CONSTANTS
# ══════════════════════════════════════════════════════════════════
NAV = '003865'   # Deep navy
BLU = '005B9A'   # Mid blue
LBL = 'E5EEF6'   # Light blue
WHT = 'FFFFFF'
BLK = '1A1A2E'
GRY = 'F5F7FA'   # Alt row
BDR = 'C8D8EC'   # Cell border
GRN = '1A7A45'   # Green accent
AMB = 'B06000'   # Amber
RED = 'B02020'   # Red
MUT = '6B8299'   # Muted
YLW = 'FFF3CD'   # Activity highlight
TAB = 'D6E4F0'   # Total row
TUL = '009E73'   # Teal (dwell chart)

# Chart series colors
CH_V  = '2D7DD2'  # Visitors
CH_U  = '1A9850'  # Unique Visitors
CH_P  = 'F4A261'  # Passersby
CH_D1 = '2D7DD2'  # Dwell series 1
CH_D2 = '009E73'  # Dwell series 2

def F(h): return PatternFill('solid', start_color=h, fgColor=h)
def fn(sz=10, bold=False, color=BLK, italic=False):
    return Font(name='Calibri', size=sz, bold=bold, color=color, italic=italic)
def al(h='center', v='center', wrap=False, indent=0):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap, indent=indent)
def bd(color=BDR, style='thin'):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)
def rh(ws, r, h): ws.row_dimensions[r].height = h

# ══════════════════════════════════════════════════════════════════
# WORKSHEET SETUP  — A4 Portrait
# ══════════════════════════════════════════════════════════════════
# 12 columns: col1=margin, col2=label(wide), col3-11=data, col12=margin
A4_COLS = [1.2, 22.0, 11.0, 11.0, 11.0, 11.0, 11.0, 11.0, 11.0, 11.0, 11.0, 1.2]

def setup_ws(wb, name, tab_color=NAV):
    ws = wb.create_sheet(name)
    ws.sheet_properties.tabColor = tab_color
    ws.sheet_view.showGridLines = False
    ws.sheet_view.showRowColHeaders = False
    ws.page_setup.paperSize   = 9          # A4
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.print_options.horizontalCentered = True
    ws.page_margins.left   = 0.5
    ws.page_margins.right  = 0.5
    ws.page_margins.top    = 0.55
    ws.page_margins.bottom = 0.55
    ws.page_margins.header = 0.3
    ws.page_margins.footer = 0.3
    for i, w in enumerate(A4_COLS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for r in range(1, 180):
        for c in range(1, 14):
            ws.cell(r, c).fill = F(WHT)
    return ws

# ══════════════════════════════════════════════════════════════════
# REUSABLE COMPONENTS
# ══════════════════════════════════════════════════════════════════
def page_header(ws, r, title, subtitle, date_str):
    rh(ws, r,   38); rh(ws, r+1, 16); rh(ws, r+2, 6)
    for c in range(1, 13):
        ws.cell(r,   c).fill = F(NAV)
        ws.cell(r+1, c).fill = F(BLU)
        ws.cell(r+2, c).fill = F(WHT)

    # PTTEP logo text
    lc = ws.cell(r, 2, ORGANIZER)
    lc.font = fn(13, bold=True, color=WHT); lc.fill = F(NAV)
    lc.alignment = al('left', indent=1)

    # Title
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=9)
    tc = ws.cell(r, 3, title)
    tc.font = fn(14, bold=True, color=WHT); tc.fill = F(NAV)
    tc.alignment = al('center')

    # Date badge
    ws.merge_cells(start_row=r, start_column=10, end_row=r, end_column=11)
    dc = ws.cell(r, 10, f'{date_str}  ')
    dc.font = fn(10, bold=True, color='B8D4F0'); dc.fill = F(NAV)
    dc.alignment = al('right')

    # Subtitle
    ws.merge_cells(start_row=r+1, start_column=2, end_row=r+1, end_column=11)
    sc = ws.cell(r+1, 2, subtitle)
    sc.font = fn(9, color=WHT, italic=True); sc.fill = F(BLU)
    sc.alignment = al('center')


def sec_hdr(ws, r, title, cs=2, ce=11, bg=LBL, fg=NAV):
    rh(ws, r, 20)
    ws.merge_cells(start_row=r, start_column=cs, end_row=r, end_column=ce)
    for c in range(cs, ce+1): ws.cell(r, c).fill = F(bg)
    c = ws.cell(r, cs, f'  {title}')
    c.font = fn(9, bold=True, color=fg); c.fill = F(bg)
    c.alignment = al('left')
    c.border = Border(bottom=Side(style='medium', color=NAV),
                      top=Side(style='thin', color=BDR))


def kpi(ws, r, col, label, value, sub='', accent=BLU):
    rh(ws, r,   16); rh(ws, r+1, 28); rh(ws, r+2, 14)
    lt = Side(style='thick', color=accent)
    th = Side(style='thin',  color=BDR)
    for row_, top_, bot_ in [(r, th, None), (r+1, None, None), (r+2, None, th)]:
        ws.merge_cells(start_row=row_, start_column=col, end_row=row_, end_column=col+1)
        cell = ws.cell(row_, col)
        cell.fill = F(WHT)
        cell.border = Border(left=lt, right=th,
                             top=(top_ or th), bottom=(bot_ or th))
    ws.cell(r,   col, f' {label}').font = fn(8, color=MUT)
    ws.cell(r,   col).alignment = al('left')
    val_str = f'{value:,}' if isinstance(value, (int, float)) else str(value)
    vc = ws.cell(r+1, col, val_str)
    vc.font = fn(20, bold=True, color=accent); vc.alignment = al('center')
    sc = ws.cell(r+2, col, sub)
    sc.font = fn(8, color=MUT); sc.fill = F(GRY); sc.alignment = al('center')


def tbl_hdr(ws, r, cols, cs=2):
    rh(ws, r, 24)
    for i, lbl in enumerate(cols):
        c = ws.cell(r, cs+i, lbl)
        c.font = fn(9, bold=True, color=WHT); c.fill = F(NAV)
        c.alignment = al('center', wrap=True)
        c.border = Border(right=Side(style='thin', color='336699'),
                          bottom=Side(style='medium', color=BLU),
                          top=Side(style='thin', color='336699'))


def tbl_row(ws, r, vals, cs=2, alt=False, total=False, hi=False):
    rh(ws, r, 17)
    if total: bg, fg, bold = TAB, NAV, True
    elif hi:  bg, fg, bold = YLW, AMB, True
    else:     bg, fg, bold = (GRY if alt else WHT), BLK, False
    for i, v in enumerate(vals):
        c = ws.cell(r, cs+i, v)
        c.font = fn(9, bold=bold, color=fg); c.fill = F(bg)
        c.alignment = al('left' if i==0 else 'center',
                         indent=(1 if i==0 else 0))
        c.border = Border(right=Side(style='thin', color=BDR),
                          bottom=Side(style='thin', color=BDR))


def gap(ws, r, h=6):
    rh(ws, r, h)
    for c in range(1, 13): ws.cell(r, c).fill = F(WHT)


def footer(ws, r):
    rh(ws, r, 14)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=11)
    c = ws.cell(r, 2,
        f'{SYSTEM_CREDIT}  ·  {EVENT_NAME}, {VENUE}'  + ('  ·  Confidential' if CONFIDENTIAL else ''))
    c.font = fn(7, italic=True, color=MUT); c.fill = F(GRY)
    c.alignment = al('center')
    c.border = Border(top=Side(style='thin', color=BDR))


# ══════════════════════════════════════════════════════════════════
# DATA LOADING
# ══════════════════════════════════════════════════════════════════

# ══════════════════════════════════════════════════════════════════
# LOAD CONFIG FROM EXCEL (_config sheet)
# ══════════════════════════════════════════════════════════════════
def load_config_from_excel(filepath):
    """
    Read configuration from _config sheet in the Rawdata Excel file.
    Returns a dict of all config values — same keys as event_config.py.
    """
    import pandas as pd, os

    xl = pd.ExcelFile(filepath)

    def read_section(sheet_df, section_marker):
        """Find rows belonging to a section (between section markers)."""
        # Strip leading/trailing spaces from cell values before matching
        col0 = sheet_df.iloc[:,0].astype(str).str.strip()
        idxs = sheet_df.index[col0.str.startswith(section_marker)].tolist()
        if not idxs:
            return pd.DataFrame()
        start = idxs[0] + 2           # skip section header + column header
        # Find next section or end
        all_sections = sheet_df.index[
            col0.str.match(r'^[A-G] —')
        ].tolist()
        next_sections = [i for i in all_sections if i > idxs[0]]
        end = next_sections[0] if next_sections else len(sheet_df)
        return sheet_df.iloc[start:end].dropna(how='all')

    raw = pd.read_excel(filepath, sheet_name='_config', header=None)

    cfg = {}

    # ── A: Event info (key/value pairs) ──────────────────────────
    a_sec = read_section(raw, 'A —')
    for _, row in a_sec.iterrows():
        k = str(row.iloc[0]).strip()
        v = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else ''
        if k and k != 'nan' and k != 'Key':
            cfg[k] = v

    cfg['confidential'] = cfg.get('confidential','True').lower() == 'true'

    # ── B: Event dates ────────────────────────────────────────────
    b_sec = read_section(raw, 'B —')
    b_sec = b_sec[b_sec.iloc[:,0].astype(str).str.match(r'\d{4}-\d{2}-\d{2}')]
    cfg['EVENT_DATES']  = [str(r.iloc[0]).strip()[:10] for _,r in b_sec.iterrows()]
    cfg['EVENT_LABELS'] = [str(r.iloc[1]).strip() for _,r in b_sec.iterrows()]
    cfg['DAY_COLORS']   = [str(r.iloc[2]).strip() for _,r in b_sec.iterrows()]

    # ── C: Gates ─────────────────────────────────────────────────
    c_sec = read_section(raw, 'C —')
    c_sec = c_sec[c_sec.iloc[:,0].isin(['Entrance','Passerby'])]
    cfg['ENTRANCE_GATES'] = c_sec[c_sec.iloc[:,0]=='Entrance'].iloc[:,1].str.strip().tolist()
    cfg['PASSERBY_GATES'] = c_sec[c_sec.iloc[:,0]=='Passerby'].iloc[:,1].str.strip().tolist()

    # ── D: Zones ─────────────────────────────────────────────────
    d_sec = read_section(raw, 'D —')
    d_sec = d_sec[pd.to_numeric(d_sec.iloc[:,0], errors='coerce').notna()]
    cfg['ZONE_ORDER']  = d_sec.iloc[:,1].str.strip().tolist()
    cfg['ZONE_ABBREV'] = {r.iloc[1].strip(): r.iloc[2].strip()
                          for _,r in d_sec.iterrows()
                          if pd.notna(r.iloc[2]) and str(r.iloc[2]).strip()}
    # Column 3 = dwell_benchmark_sec, col 4 = description, col 5 = benchmark_mode
    _zbench = {}
    _zdesc = {}
    _zmode = {}
    for _, r in d_sec.iterrows():
        zname = str(r.iloc[1]).strip()
        if not zname:
            continue
        if len(r) > 3 and pd.notna(r.iloc[3]) and str(r.iloc[3]).strip():
            try:
                _zbench[zname] = int(float(r.iloc[3]))
            except (ValueError, TypeError):
                pass
        if len(r) > 4 and pd.notna(r.iloc[4]) and str(r.iloc[4]).strip():
            _zdesc[zname] = str(r.iloc[4]).strip()
        if len(r) > 5 and pd.notna(r.iloc[5]) and str(r.iloc[5]).strip():
            _m = str(r.iloc[5]).strip().lower()
            _zmode[zname] = 'lower_better' if _m == 'lower_better' else 'higher_better'
    cfg['ZONE_BENCHMARK'] = _zbench
    cfg['ZONE_DESC'] = _zdesc
    cfg['ZONE_BENCHMARK_MODE'] = _zmode

    # ── E: Activities ────────────────────────────────────────────
    e_sec = read_section(raw, 'E —')
    e_sec = e_sec[e_sec.iloc[:,0].astype(str).str.match(r'\d{4}-\d{2}-\d{2}')]
    activities = {}
    for _, row in e_sec.iterrows():
        ds   = str(row.iloc[0]).strip()[:10]
        ts   = str(row.iloc[1]).strip()[:5]
        te   = str(row.iloc[2]).strip()[:5]
        name = str(row.iloc[3]).strip()
        activities.setdefault(ds, []).append((ts, te, name))
    cfg['ACTIVITIES'] = activities

    # ── F: Parameters ────────────────────────────────────────────
    f_sec = read_section(raw, 'F —')
    for _, row in f_sec.iterrows():
        k = str(row.iloc[0]).strip()
        v = row.iloc[1]
        if k and k != 'nan' and k != 'parameter' and pd.notna(v):
            cfg[k] = int(v) if isinstance(v, float) and v == int(v) else v

    # Derived
    # exclude_staff: enables staff filter for unique/dwell metrics (default True)
    _exc = str(cfg.get('exclude_staff', 'true')).strip().lower()
    cfg['EXCLUDE_STAFF']             = _exc in ('true', '1', 'yes', 'on')
    _sdb = str(cfg.get('show_dwell_benchmark', 'false')).strip().lower()
    cfg['SHOW_DWELL_BENCHMARK']      = _sdb in ('true', '1', 'yes', 'on')
    cfg['DWELL_MIN_SEC']             = int(cfg.get('dwell_min_sec', 10))
    cfg['DWELL_MAX_SEC']             = int(cfg.get('dwell_max_sec', 3600))
    cfg['ENGAGEMENT_THRESHOLD_SEC']  = int(cfg.get('engagement_threshold_sec', 60))
    cfg['DISPLAY_HOURS_START']       = int(cfg.get('display_hours_start', 9))
    cfg['DISPLAY_HOURS_END']         = int(cfg.get('display_hours_end', 19))

    return cfg


def load_rawdata_from_excel(filepath, col_names, zone_types, event_dates):
    """
    Load raw data from data_YYYY-MM-DD sheets.
    Falls back to 'data' sheet if per-day sheets not found.
    """
    import pandas as pd
    xl     = pd.ExcelFile(filepath)
    sheets = xl.sheet_names
    dfs    = []

    day_sheets = [f'data_{ds}' for ds in event_dates]
    found_day_sheets = [s for s in day_sheets if s in sheets]

    if found_day_sheets:
        # ── Per-day sheets (new format) ───────────────────────────
        for sheet in found_day_sheets:
            try:
                d = xl.parse(sheet)
                # Skip placeholder-only sheets (header + 1 note row, no real data)
                if len(d) <= 1 or (len(d) == 1 and '←' in str(d.iloc[0,0])):
                    print(f'  ⚠  {sheet}: no data (placeholder only)')
                    continue
                d.columns = col_names
                dfs.append(d)
                print(f'  Loaded: {sheet}  ({len(d):,} rows)')
            except Exception as e:
                print(f'  Warning — {sheet}: {e}')
    elif 'data' in sheets:
        # ── Legacy single sheet ───────────────────────────────────
        d = xl.parse('data')
        d.columns = col_names
        dfs.append(d)
        print(f'  Loaded: data sheet  ({len(d):,} rows) [legacy format]')
    else:
        raise ValueError(f'No data sheets found in {filepath}')

    if not dfs:
        # Try loading from separate Rawdata files (legacy single-sheet format)
        import glob as _g
        _dir = os.path.dirname(filepath)
        legacy = [f for f in sorted(_g.glob(os.path.join(_dir, 'Rawdata*.xlsx')))
                  if '_config' not in pd.ExcelFile(f).sheet_names]
        if legacy:
            for f in legacy:
                d = pd.read_excel(f, sheet_name='data')
                d.columns = col_names
                dfs.append(d)
                print(f'  Loaded (legacy): {os.path.basename(f)}  ({len(d):,} rows)')
        if not dfs:
            raise ValueError('No data found. Add raw data to data_YYYY-MM-DD sheets or place Rawdata*.xlsx in the same folder.')

    df = pd.concat(dfs, ignore_index=True)
    df['Time'] = pd.to_datetime(df['Time'])
    df['Date'] = df['Time'].dt.strftime('%Y-%m-%d')
    df['Hour'] = df['Time'].dt.hour
    df['Type'] = df['Location'].map(zone_types).fillna('Zone')
    return df

def load_data():
    dfs = []
    for f in INPUT_FILES:
        try:
            d = pd.read_excel(f, sheet_name='data')
            d.columns = COL_NAMES
            dfs.append(d)
            print(f'  Loaded: {f.split("/")[-1]}  ({len(d):,} rows)')
        except Exception as e:
            print(f'  Warning — could not load {f}: {e}')
    df = pd.concat(dfs, ignore_index=True)
    df['Time']  = pd.to_datetime(df['Time'])
    df['Date']  = df['Time'].dt.strftime('%Y-%m-%d')
    df['Hour']  = df['Time'].dt.hour
    df['Type']  = df['Location'].map(ZONE_TYPES).fillna('Zone')
    return df

# ══════════════════════════════════════════════════════════════════
# DWELL TIME COMPUTATION
# ══════════════════════════════════════════════════════════════════
def compute_zone_dwell(df):
    """
    Match in→out pairs at same Location per BodyID.
    Filter: 10 sec ≤ dwell ≤ 3600 sec (3 hr).
    Returns DataFrame with columns:
      BodyID, Location, dwell_sec, dwell_min, hour
    """
    df = _non_staff(df)   # exclude staff from dwell analysis
    zone_io = (df[df['Type']=='Zone'][['BodyID','Location','Event','Time','Hour']]
               .pipe(lambda x: x[x['Event'].isin(['in','out'])])
               .sort_values(['BodyID','Location','Time']))

    sessions = []
    for (bid, loc), grp in zone_io.groupby(['BodyID','Location'], sort=False):
        ins      = grp[grp['Event']=='in']['Time'].tolist()
        in_hours = grp[grp['Event']=='in']['Hour'].tolist()
        outs     = grp[grp['Event']=='out']['Time'].tolist()
        oi = 0
        for idx, t_in in enumerate(ins):
            while oi < len(outs) and outs[oi] <= t_in:
                oi += 1
            if oi < len(outs):
                sec = (outs[oi] - t_in).total_seconds()
                if DWELL_MIN_SEC <= sec <= DWELL_MAX_SEC:  # 10 sec – 60 min
                    sessions.append({
                        'BodyID': bid, 'Location': loc,
                        'dwell_sec': sec, 'dwell_min': sec / 60,
                        'hour': in_hours[idx] if idx < len(in_hours) else grp['Hour'].iloc[0],
                    })
                oi += 1
    return pd.DataFrame(sessions) if sessions else pd.DataFrame(
        columns=['BodyID','Location','dwell_sec','dwell_min','hour'])


def compute_booth_dwell(df):
    """
    Booth-level: first-seen → last-seen across all ENT+Zone in/out events.
    Filter: dwell > 0 (seen at ≥2 time points).
    Returns DataFrame: BodyID, dwell_min, first_seen, last_seen, hour
    """
    df = _non_staff(df)   # exclude staff from booth dwell
    booth = (df[df['Type'].isin(['Entrance','Zone'])]
             [['BodyID','Event','Time','Hour']]
             .pipe(lambda x: x[x['Event'].isin(['in','out'])]))
    if len(booth) == 0:
        return pd.DataFrame(columns=['BodyID','dwell_min','first_seen','last_seen','hour'])
    spans = booth.groupby('BodyID')['Time'].agg(['min','max']).reset_index()
    spans.columns = ['BodyID','first_seen','last_seen']
    spans['dwell_min'] = (spans['last_seen'] - spans['first_seen']).dt.total_seconds() / 60
    spans['hour'] = spans['first_seen'].dt.hour
    return spans[spans['dwell_min'] > 0].copy()


def dwell_summary(dwell_df):
    """Compute summary stats for a dwell Series or sub-DataFrame."""
    s = dwell_df['dwell_min'] if isinstance(dwell_df, pd.DataFrame) else dwell_df
    if len(s) == 0:
        return dict(n=0, avg=0, med=0, p75=0, p90=0)
    return dict(
        n   = len(s),
        avg = round(s.mean(),  1),
        med = round(s.median(),1),
        p75 = round(s.quantile(0.75), 1),
        p90 = round(s.quantile(0.90), 1),
    )


def dwell_buckets(dwell_min_series):
    """Return counts per time bucket."""
    s = dwell_min_series
    return {
        '< 1 min':    int((s < 1).sum()),
        '1 – 5 min':  int(((s >= 1)  & (s < 5)).sum()),
        '5 – 15 min': int(((s >= 5)  & (s < 15)).sum()),
        '15 – 30 min':int(((s >= 15) & (s < 30)).sum()),
        '> 30 min':   int((s >= 30).sum()),
    }


# ══════════════════════════════════════════════════════════════════
# CHART HELPERS
# ══════════════════════════════════════════════════════════════════
def write_hidden_data(ws, r0, c0, headers, rows_data):
    """Write data to invisible cells for chart source."""
    for ci, h in enumerate(headers):
        ws.cell(r0, c0+ci, h).font = fn(7, color=WHT)
        ws.cell(r0, c0+ci).fill = F(WHT)
    for ri, row in enumerate(rows_data):
        for ci, v in enumerate(row):
            ws.cell(r0+1+ri, c0+ci, v).font = fn(7, color=WHT)
            ws.cell(r0+1+ri, c0+ci).fill = F(WHT)


def make_bar_chart(ws, anchor, r0, c0, n_rows, n_series,
                   series_colors, width=17, height=9,
                   bar_type='col', legend_pos='b'):
    chart = BarChart()
    chart.type      = bar_type
    chart.grouping  = 'clustered'
    chart.gapWidth  = 70
    chart.title     = None
    chart.style     = 2
    chart.legend.position = legend_pos

    dr = Reference(ws, min_col=c0+1, max_col=c0+n_series,
                   min_row=r0, max_row=r0+n_rows)
    cr = Reference(ws, min_col=c0, min_row=r0+1, max_row=r0+n_rows)
    chart.add_data(dr, titles_from_data=True)
    chart.set_categories(cr)

    for i, color in enumerate(series_colors):
        if i < len(chart.series):
            chart.series[i].graphicalProperties.solidFill = color
            chart.series[i].graphicalProperties.line.solidFill = color

    chart.plot_area.graphicalProps = None
    chart.width  = width
    chart.height = height
    ws.add_chart(chart, anchor)


# ══════════════════════════════════════════════════════════════════
# HOURLY TRAFFIC TABLE + CHART
# ══════════════════════════════════════════════════════════════════
def hourly_section(ws, r, ent, psb, acts_list, dc=15,
                   all_days_data=None, date_str=None):
    """
    Hourly traffic table + chart.
    - Single day mode  : pass ent/psb filtered to one date + acts_list
    - multi-day summary mode : pass all_days_data = {date_str: (ent_d, psb_d)}
    dc = start column for hidden chart data.
    Returns: row after section.
    """
    # ── 09:00–19:00 covers all 3 days (Day1 closes at 19:xx) ──────
    HOURS = list(range(9, 20))   # 09:00 … 19:00

    # ── Activity notes helper ──────────────────────────────────────
    def notes(h):
        sm, em = h*60, h*60+60
        hits = [a[2] for a in acts_list
                if (int(a[0][:2])*60+int(a[0][3:])) < em
                and (int(a[1][:2])*60+int(a[1][3:])) > sm]
        return '  ●  ' + ' / '.join(hits) if hits else ''

    # ══════════════════════════════════════════════════════════════
    # MODE A: Single-day table (used in Day sheets)
    # ══════════════════════════════════════════════════════════════
    if all_days_data is None:
        sec_hdr(ws, r,
                '│ HOURLY TRAFFIC — Visitors  /  Unique Visitors  /  Passersby')

        tbl_hdr(ws, r+1,
                ['Hour', 'Visitors', 'Unique Visitors', 'Passersby', 'Booth Activities'])

        rows = []
        for h in HOURS:
            eh = ent[ent['Hour']==h]; ph = psb[psb['Hour']==h]
            v  = len(eh); uv = uv_count(eh); p = len(ph) + v
            rows.append((f'{h:02d}:00', v, uv, p, notes(h)))

        # Only show rows that have any data OR are within operating hours
        has_late = any(r[1] > 0 for r in rows if int(r[0][:2]) >= 19)
        show_hours = HOURS if has_late else list(range(9, 19))

        peak_v = max(row[1] for row in rows) if rows else 0
        for i, h in enumerate(show_hours):
            hl, v, uv, p, note = rows[h - 9]
            tbl_row(ws, r+2+i,
                    [hl, v or '—', uv or '—', p or '—', note],
                    alt=(i%2==0), hi=bool(note))
            if v == peak_v and peak_v > 0:
                c = ws.cell(r+2+i, 3)
                c.font = fn(9, bold=True, color=WHT); c.fill = F(BLU)

        tr = r + 2 + len(show_hours)
        tbl_row(ws, tr,
                ['TOTAL', sum(rows[h-9][1] for h in show_hours), '',
                 sum(rows[h-9][3] for h in show_hours), ''],
                total=True)
        gap(ws, tr+1)

        # Chart
        cr = tr + 2
        sec_hdr(ws, cr, '│ HOURLY TRAFFIC CHART')
        for ri in range(cr+1, cr+16): rh(ws, ri, 14)
        chart_data = [(f'{h:02d}:00', rows[h-9][1], rows[h-9][2], rows[h-9][3])
                      for h in show_hours]
        write_hidden_data(ws, cr+1, dc,
                          ['Hour', 'Visitors', 'Unique', 'Passersby'],
                          chart_data)
        make_bar_chart(ws, f'B{cr+1}', cr+1, dc, len(show_hours), 3,
                       [CH_V, CH_U, CH_P])
        gap(ws, cr+16)
        return cr + 17

    # ══════════════════════════════════════════════════════════════
    # MODE B: multi-day stacked tables — one per day, same layout as Mode A
    # ══════════════════════════════════════════════════════════════
    _DC = [c.lstrip('#') for c in DAY_COLORS[:len(EVENT_DATES)]]  # strip # for openpyxl
    DAY_LABELS = [f'Day {i+1}  \u00b7  {l}' for i,l in enumerate(EVENT_LABELS)]
    sec_hdr(ws, r,
            '│ HOURLY TRAFFIC — Visitors  /  Unique Visitors  /  Passersby  (แยกตามวัน)')

    cur = r + 1   # cursor row, advances after each day block

    # Collect all chart data for combined 3-series chart at end
    all_chart_data = {}   # date_str → [(hour_label, v, uv, p), …]

    for di, ds in enumerate(EVENT_DATES):
        dcolor = _DC[di]
        dlabel = DAY_LABELS[di]

        # ── Day sub-header ─────────────────────────────────────────
        rh(ws, cur, 18)
        ws.merge_cells(start_row=cur, start_column=2,
                       end_row=cur,   end_column=6)
        dh = ws.cell(cur, 2, f'  {dlabel}')
        dh.font = fn(10, bold=True, color=WHT)
        dh.fill = F(dcolor); dh.alignment = al('left')
        dh.border = Border(bottom=Side(style='thin', color=BDR))
        cur += 1

        # ── Get data ───────────────────────────────────────────────
        if ds in all_days_data:
            e_d, p_d = all_days_data[ds]
            has_data = len(e_d) > 0 or len(p_d) > 0
        else:
            e_d = pd.DataFrame(columns=['Hour', 'BodyID'])
            p_d = pd.DataFrame(columns=['Hour'])
            has_data = False

        if not has_data:
            rh(ws, cur, 16)
            ws.merge_cells(start_row=cur, start_column=2,
                           end_row=cur,   end_column=6)
            nc = ws.cell(cur, 2, '  ⏳  No data for this day — pending upload')
            nc.font = fn(9, italic=True, color=MUT)
            nc.fill = F(GRY); nc.alignment = al('left')
            cur += 1
            gap(ws, cur, 4); cur += 1
            all_chart_data[ds] = []
            continue

        # ── Column header ──────────────────────────────────────────
        tbl_hdr(ws, cur,
                ['Hour', 'Visitors', 'Unique Visitors', 'Passersby', 'Booth Activities'],
                cs=2)
        cur += 1

        # ── Compute rows ───────────────────────────────────────────
        acts_for_day = ACTIVITIES.get(ds, [])
        def day_notes(h):
            sm, em = h*60, h*60+60
            hits = [a[2] for a in acts_for_day
                    if (int(a[0][:2])*60+int(a[0][3:])) < em
                    and (int(a[1][:2])*60+int(a[1][3:])) > sm]
            return '  ●  ' + ' / '.join(hits) if hits else ''

        rows = []
        for h in HOURS:
            eh = e_d[e_d['Hour']==h]; ph = p_d[p_d['Hour']==h]
            v  = len(eh); uv = uv_count(eh) if len(eh) > 0 else 0
            p  = len(ph) + v  # Passersby = PSB gate + Visitors
            rows.append((f'{h:02d}:00', v, uv, p, day_notes(h)))

        has_late   = any(row[1] > 0 for row in rows if int(row[0][:2]) >= 19)
        show_hours = HOURS if has_late else list(range(9, 19))
        peak_v     = max(row[1] for row in rows) if rows else 0

        for i, h in enumerate(show_hours):
            hl, v, uv, p, note = rows[h - 9]
            tbl_row(ws, cur,
                    [hl, v or '—', uv or '—', p or '—', note],
                    alt=(i % 2 == 0), hi=bool(note))
            if v == peak_v and peak_v > 0:
                c = ws.cell(cur, 3)
                c.font = fn(9, bold=True, color=WHT); c.fill = F(dcolor)
            cur += 1

        # ── Total row ──────────────────────────────────────────────
        tbl_row(ws, cur,
                ['TOTAL',
                 sum(rows[h-9][1] for h in show_hours),
                 '',
                 sum(rows[h-9][3] for h in show_hours),
                 ''],
                total=True)
        cur += 1

        # Store chart data
        all_chart_data[ds] = [
            (f'{h:02d}:00', rows[h-9][1], rows[h-9][2], rows[h-9][3])
            for h in show_hours
        ]

        gap(ws, cur, 6); cur += 1

    # ── Combined multi-day chart (Visitors only, one series per day) ──
    sec_hdr(ws, cur, f'│ HOURLY VISITOR TRAFFIC CHART  ·  All {len(EVENT_DATES)} Days')
    cur += 1
    for ri in range(cur, cur+16): rh(ws, ri, 14)

    # Use Day 1 hours as base axis (longest)
    base_hours = list(range(9, 19))
    for ds in EVENT_DATES:
        cdata = all_chart_data.get(ds, [])
        if cdata and int(cdata[-1][0][:2]) >= 19:
            base_hours = list(range(9, 20)); break

    chart_rows = []
    for h in base_hours:
        row_vals = [f'{h:02d}:00']
        for ds in EVENT_DATES:
            cdata = all_chart_data.get(ds, [])
            match = next((c[1] for c in cdata if c[0] == f'{h:02d}:00'), 0)
            row_vals.append(match)
        chart_rows.append(tuple(row_vals))

    write_hidden_data(ws, cur, dc,
                      ['Hour'] + [f'Day {i+1} ({l[:6]})' for i,l in enumerate(EVENT_LABELS)],
                      chart_rows)
    make_bar_chart(ws, f'B{cur}', cur, dc, len(base_hours), 3,
                   [BLU, GRN, AMB])
    cur += 16
    gap(ws, cur)
    return cur + 1

def zone_section(ws, r, zon, dc=15):
    sec_hdr(ws, r, '│ ZONE TRAFFIC BREAKDOWN')

    if len(zon) == 0:  # simple profile — no zones
        gap(ws, r+1)
        ws.cell(r+1, 2, 'No zone data for this event type')
        return r + 2

    # Staff-aware: visits count uses raw zon, unique uses _non_staff filter
    _zon_ns = _non_staff(zon)
    _vis = zon.groupby('Location').size().rename('Visits')
    _uniq = _zon_ns.groupby('Location')['BodyID'].nunique().rename('Unique')
    stats = (_vis.to_frame().join(_uniq, how='left').fillna({'Unique': 0})
             .reset_index().sort_values('Visits', ascending=False))
    stats['Unique'] = stats['Unique'].astype(int)
    # Use only Gender column to avoid TypeError on datetime columns
    zm = zon.groupby('Location')['Gender'].apply(lambda x:(x=='Male').sum())
    zf = zon.groupby('Location')['Gender'].apply(lambda x:(x=='Female').sum())
    tv = int(stats['Visits'].sum())

    tbl_hdr(ws, r+1, ['Zone / Area','Total Visits','Unique Visitors',
                       'Male','Female','% of Total'])
    for i, (_, zr) in enumerate(stats.iterrows()):
        loc = zr['Location']
        tbl_row(ws, r+2+i,
                [loc, int(zr['Visits']), int(zr['Unique']),
                 int(zm.get(loc,0)), int(zf.get(loc,0)),
                 f"{int(zr['Visits'])/tv*100:.1f}%" if tv>0 else '—'],
                alt=(i%2==0))
    zt = r+2+len(stats)
    tbl_row(ws, zt, ['TOTAL', tv, int(uv_count(zon)),
                     int(zm.sum()), int(zf.sum()), '100%'], total=True)
    gap(ws, zt+1)

    # Zone chart (horizontal)
    zc = zt+2
    sec_hdr(ws, zc, '│ ZONE VISIT CHART  ·  Total Visits / Unique Visitors')
    for ri in range(zc+1, zc+14): rh(ws, ri, 14)

    zone_rows = [(r['Location'].replace('TV Touchscreen ','TV-')
                               .replace(' Total',''), int(r['Visits']), int(r['Unique']))
                 for _, r in stats.iterrows()]
    write_hidden_data(ws, zc+1, dc, ['Zone','Visits','Unique'], zone_rows)
    make_bar_chart(ws, f'B{zc+1}', zc+1, dc, len(stats), 2,
                   [CH_V, CH_U], bar_type='bar', height=10)

    gap(ws, zc+14)
    return zc+15


# ══════════════════════════════════════════════════════════════════
# DEMOGRAPHICS
# ══════════════════════════════════════════════════════════════════
def demo_section(ws, r, df_day):
    sec_hdr(ws, r, '│ VISITOR DEMOGRAPHICS  ·  Gender & Age Group')

    male   = int((df_day['Gender']=='Male').sum())
    female = int((df_day['Gender']=='Female').sum())
    tg     = male + female

    # Gender table (cols 2-4)
    tbl_hdr(ws, r+1, ['Gender','Count','Share %'], cs=2)
    for i,(lbl,n) in enumerate([('Male',male),('Female',female)]):
        tbl_row(ws, r+2+i, [lbl, n, f'{n/tg*100:.1f}%' if tg>0 else '—'],
                cs=2, alt=(i%2==0))
    tbl_row(ws, r+4, ['TOTAL', tg, '100%'], cs=2, total=True)

    # Age table (cols 6-9)
    age_order  = ['Juvenile (0-18 yrs)','Young Adults (19-35 yrs)',
                  'Middle-Aged (36-55 yrs)','Seniors (55+ yrs)']
    age_labels = ['Juvenile (0–18)','Young Adults (19–35)',
                  'Middle-Aged (36–55)','Seniors (55+)']
    ac = df_day['AgeGroup'].value_counts()
    ta = len(df_day)
    bc = [BLU, GRN, AMB, RED]

    tbl_hdr(ws, r+1, ['Age Group','Count','Share %','▐ Distribution'], cs=6)
    for i,(full,lbl) in enumerate(zip(age_order, age_labels)):
        n   = int(ac.get(full, 0))
        pct = n/ta*100 if ta>0 else 0
        bar = '█'*int(pct/5) + '░'*max(0, 20-int(pct/5))
        tbl_row(ws, r+2+i, [lbl, n, f'{pct:.1f}%', bar[:16]], cs=6, alt=(i%2==0))
        ws.cell(r+2+i, 9).font = fn(8, color=bc[i])
        ws.cell(r+2+i, 9).alignment = al('left')
    tbl_row(ws, r+6, ['TOTAL', ta, '100%', ''], cs=6, total=True)

    gap(ws, r+7)
    return r+8


# ══════════════════════════════════════════════════════════════════
# DWELL TIME SHEET
# ══════════════════════════════════════════════════════════════════
def build_dwell_sheet(wb, df):
    ws = setup_ws(wb, 'Dwell Time Analysis', tab_color=TUL)

    page_header(ws, 1,
        f'{EVENT_NAME}  —  Dwell Time Analysis',
        'Time spent in Booth & each Zone  ·  Matched In / Out Events per Visitor ID',
        f'{EVENT_LABELS[0][:6]} {EVENT_DATES[0][-4:]}')

    print('  Computing dwell times…')
    zone_dwell  = compute_zone_dwell(df)
    booth_dwell = compute_booth_dwell(df)

    # ── BOOTH SUMMARY KPIs ─────────────────────────────────────────
    gap(ws, 4, 8)
    rh(ws, 5, 16); rh(ws, 6, 28); rh(ws, 7, 14)
    gap(ws, 8, 8)

    bd_stat = dwell_summary(booth_dwell)
    zd_stat = dwell_summary(zone_dwell)

    kpi(ws, 5,  2, 'BOOTH DWELL (Avg)',    f"{bd_stat['avg']} min",  f"{bd_stat['n']:,} sessions tracked",  TUL)
    kpi(ws, 5,  4, 'BOOTH DWELL (Median)', f"{bd_stat['med']} min",  'p50 — half stay longer',              GRN)
    kpi(ws, 5,  6, 'ZONE DWELL (Avg)',     f"{zd_stat['avg']} min",  f"{zd_stat['n']:,} zone sessions",     BLU)
    kpi(ws, 5,  8, 'ZONE DWELL (Median)',  f"{zd_stat['med']} min",  'p50 — half stay longer',              NAV)
    kpi(ws, 5, 10, 'ZONE DWELL (p75)',     f"{zd_stat['p75']} min",  '75th percentile',                    AMB)

    # ── BOOTH DWELL DISTRIBUTION ───────────────────────────────────
    r = 9
    sec_hdr(ws, r, '1 │ BOOTH DWELL TIME DISTRIBUTION  '
            '(first seen → last seen across all gates & zones per visitor)')

    bd_bkts = dwell_buckets(booth_dwell['dwell_min'])
    tbl_hdr(ws, r+1, ['Dwell Range','Visitors','Share %','▐ Proportion'])
    total_bd = sum(bd_bkts.values())
    for i,(lbl,n) in enumerate(bd_bkts.items()):
        pct = n/total_bd*100 if total_bd>0 else 0
        bar = '█'*int(pct/5)+'░'*max(0,20-int(pct/5))
        tbl_row(ws, r+2+i, [lbl, n, f'{pct:.1f}%', bar[:16]], alt=(i%2==0))
        ws.cell(r+2+i, 5).font = fn(8, color=TUL)
        ws.cell(r+2+i, 5).alignment = al('left')
    tbl_row(ws, r+2+len(bd_bkts), ['TOTAL', total_bd, '100%', ''], total=True)

    # Booth dwell distribution chart
    dc = 14
    bkr = r+2+len(bd_bkts)+2
    sec_hdr(ws, bkr, '2 │ BOOTH DWELL DISTRIBUTION CHART')
    for ri in range(bkr+1, bkr+11): rh(ws, ri, 14)

    bkt_rows = [(lbl, n) for lbl, n in bd_bkts.items()]
    write_hidden_data(ws, bkr+1, dc, ['Range','Visitors'], bkt_rows)
    make_bar_chart(ws, f'B{bkr+1}', bkr+1, dc, len(bkt_rows), 1,
                   [TUL], height=8)
    gap(ws, bkr+11)

    # ── PER-ZONE DWELL TABLE ───────────────────────────────────────
    ztr = bkr+12
    sec_hdr(ws, ztr, '3 │ PER-ZONE DWELL TIME SUMMARY  '
            '(in → out at same zone per visitor, 10 s – 3 hr filter)')

    tbl_hdr(ws, ztr+1, ['Zone / Area','Unique Visitors','Avg (min)','Median (min)',
                          'p75 (min)','p90 (min)'])
    zone_order_use = [z for z in ZONE_ORDER if z in zone_dwell['Location'].unique()] \
                   + [z for z in zone_dwell['Location'].unique() if z not in ZONE_ORDER]

    for i, loc in enumerate(zone_order_use):
        g = zone_dwell[zone_dwell['Location']==loc]
        s = dwell_summary(g)
        tbl_row(ws, ztr+2+i,
                [loc, s['n'], s['avg'], s['med'], s['p75'], s['p90']],
                alt=(i%2==0))
        # Colour-code avg: green=short, amber=medium, red=long
        avg_cell = ws.cell(ztr+2+i, 4)
        if s['avg'] < 5:
            avg_cell.font = fn(9, color=GRN, bold=True)
        elif s['avg'] < 15:
            avg_cell.font = fn(9, color=AMB, bold=True)
        else:
            avg_cell.font = fn(9, color=RED, bold=True)

    # Grand total row
    s_all = dwell_summary(zone_dwell)
    tbl_row(ws, ztr+2+len(zone_order_use),
            ['ALL ZONES', s_all['n'], s_all['avg'], s_all['med'],
             s_all['p75'], s_all['p90']], total=True)

    gap(ws, ztr+3+len(zone_order_use))

    # ── ZONE DWELL CHART (avg) ─────────────────────────────────────
    zcr = ztr+4+len(zone_order_use)
    sec_hdr(ws, zcr, '4 │ ZONE DWELL TIME CHART  ·  Avg  &  Median (minutes)')
    for ri in range(zcr+1, zcr+14): rh(ws, ri, 14)

    zone_chart_rows = []
    for loc in zone_order_use:
        g = zone_dwell[zone_dwell['Location']==loc]
        s = dwell_summary(g)
        short = loc.replace('TV Touchscreen ','TV-').replace(' Total','')
        zone_chart_rows.append((short, s['avg'], s['med']))

    write_hidden_data(ws, zcr+1, dc, ['Zone','Avg min','Median min'], zone_chart_rows)
    make_bar_chart(ws, f'B{zcr+1}', zcr+1, dc, len(zone_chart_rows), 2,
                   [CH_D1, CH_D2], bar_type='bar', height=10)
    gap(ws, zcr+14)

    # ── HOURLY DWELL TREND ─────────────────────────────────────────
    hdr = zcr+15
    sec_hdr(ws, hdr, '5 │ HOURLY AVERAGE DWELL TIME  ·  All Zones Combined')

    hours = list(range(9, 19))
    tbl_hdr(ws, hdr+1, ['Hour','Sessions','Avg Dwell (min)','Median Dwell (min)'])
    for i, h in enumerate(hours):
        gh = zone_dwell[zone_dwell['hour']==h]
        s  = dwell_summary(gh)
        tbl_row(ws, hdr+2+i,
                [f'{h:02d}:00', s['n'], s['avg'], s['med']],
                alt=(i%2==0))

    tbl_row(ws, hdr+2+len(hours),
            ['ALL HOURS', s_all['n'], s_all['avg'], s_all['med']], total=True)

    gap(ws, hdr+3+len(hours))

    # Hourly dwell chart
    hcr = hdr+4+len(hours)
    sec_hdr(ws, hcr, '6 │ HOURLY DWELL TREND CHART  ·  Avg Dwell per Entry Hour')
    for ri in range(hcr+1, hcr+12): rh(ws, ri, 14)

    hourly_rows = []
    for h in hours:
        gh = zone_dwell[zone_dwell['hour']==h]
        s  = dwell_summary(gh)
        hourly_rows.append((f'{h:02d}:00', s['avg'], s['med']))

    write_hidden_data(ws, hcr+1, dc, ['Hour','Avg','Median'], hourly_rows)
    make_bar_chart(ws, f'B{hcr+1}', hcr+1, dc, len(hourly_rows), 2,
                   [CH_D1, CH_D2], height=8)
    gap(ws, hcr+12)

    # ── METHODOLOGY NOTE ──────────────────────────────────────────
    mn = hcr+13
    rh(ws, mn, 14); rh(ws, mn+1, 14); rh(ws, mn+2, 14); rh(ws, mn+3, 14)
    ws.merge_cells(start_row=mn, start_column=2, end_row=mn, end_column=11)
    nc = ws.cell(mn, 2, '  📌  Methodology Notes')
    nc.font = fn(9, bold=True, color=NAV); nc.fill = F(LBL); nc.alignment = al('left')
    notes = [
        '  • Booth Dwell Time: First sensor detection → last sensor detection across all ENT + Zone sensors per Visitor ID',
        '  • Zone Dwell Time:  Matched in-event → next out-event at the same zone per Visitor ID (10 sec – 60 min valid range)',
        '  • Sessions with dwell < 10 seconds are excluded (likely sensor noise). Only visitors with paired in+out are counted.',
    ]
    for j, note in enumerate(notes):
        ws.merge_cells(start_row=mn+1+j, start_column=2, end_row=mn+1+j, end_column=11)
        c = ws.cell(mn+1+j, 2, note)
        c.font = fn(8, color=BLK, italic=True); c.fill = F(GRY); c.alignment = al('left')

    gap(ws, mn+4)
    footer(ws, mn+5)


# ══════════════════════════════════════════════════════════════════
# ACTIVITY ANALYTICS SHEET
# ══════════════════════════════════════════════════════════════════
def build_activity_sheet(wb, df):
    ws = setup_ws(wb, 'Activity Analytics', tab_color=AMB)

    page_header(ws, 1,
        f'{EVENT_NAME}  —  Booth Activity Analytics',
        'Traffic Analysis During Scheduled Activities  ·  All Days',
        f'{EVENT_LABELS[0][:6]} – {EVENT_LABELS[-1][:6]} {EVENT_DATES[-1][-4:]}')

    ent = df[(df['Type']=='Entrance') & (df['Event']=='in')]
    psb = df[df['Type']=='Passerby']
    zon = df[(df['Type']=='Zone') & (df['Event']=='in')]

    # Build per-activity traffic
    all_act_rows = []
    for date_str in EVENT_DATES:
        for t_s, t_e, name in ACTIVITIES.get(date_str, []):
            sh, sm = int(t_s[:2]), int(t_s[3:])
            eh, em = int(t_e[:2]), int(t_e[3:])
            s_min, e_min = sh*60+sm, eh*60+em
            def in_w(sub):
                mins = sub['Time'].dt.hour*60 + sub['Time'].dt.minute
                return sub[(sub['Date']==date_str) & (mins >= s_min) & (mins <= e_min)]
            v_w = in_w(ent); uv_w = uv_count(v_w)
            p_w = in_w(psb); z_w = in_w(zon)
            all_act_rows.append({
                'date': date_str, 'time': f'{t_s}–{t_e}',
                'name': name, 'visitors': len(v_w),
                'unique': int(uv_w), 'passersby': len(p_w),
                'zone_visits': len(z_w),
            })

    act_df = pd.DataFrame(all_act_rows)

    # ── 🎪 Activity Traffic Summary ────────────────────────────────
    r = 4
    gap(ws, r)
    sec_hdr(ws, r+1, '🎪  Activity Traffic Summary')

    if act_df.empty or len(all_act_rows) == 0:
        ws.cell(r+2, 2, 'ไม่มีกำหนดการกิจกรรมสำหรับ event นี้ — เพิ่มใน Section E ของ _config sheet')
        footer(ws, r+4)
        return

    tbl_hdr(ws, r+2, ['Date','Time','Activity','Visitors','Unique','Passersby','Zone Visits'])
    for i, row in enumerate(all_act_rows):
        tbl_row(ws, r+3+i,
                [row['date'], row['time'], row['name'],
                 row['visitors'], row['unique'],
                 row['passersby'], row['zone_visits']],
                alt=(i%2==0), hi=True)
    tr = r+3+len(all_act_rows)
    tbl_row(ws, tr, ['', 'TOTAL', '',
                      act_df['visitors'].sum(), '',
                      act_df['passersby'].sum(),
                      act_df['zone_visits'].sum()], total=True)
    gap(ws, tr+1)

    # ── 🗺️ Activity × Zone Traffic Matrix ─────────────────────────
    mtr = tr+2
    sec_hdr(ws, mtr, '🗺️  Activity × Zone Traffic Matrix')

    zones_in_data = [z for z in ZONE_ORDER if z in zon['Location'].unique()]
    short_zones   = [z.replace('TV Touchscreen ','TV-').replace(' Total','')[:13]
                     for z in zones_in_data]

    # Header: date/time | activity name | zone cols…
    rh(ws, mtr+1, 28)
    for ci, lbl in enumerate(['Activity'] + short_zones):
        c = ws.cell(mtr+1, 2+ci, lbl)
        c.font = fn(8, bold=True, color=WHT); c.fill = F(NAV)
        c.alignment = al('center', wrap=True)
        c.border = Border(right=Side(style='thin',   color='336699'),
                          bottom=Side(style='medium', color=BLU))

    # Data rows
    max_val = 1
    matrix_data = []
    for row in all_act_rows:
        sh, sm = int(row['time'][:2]), int(row['time'][3:5])
        eh, em = int(row['time'][6:8]), int(row['time'][9:11])
        s_min, e_min = sh*60+sm, eh*60+em
        zone_counts = []
        for z in zones_in_data:
            sub = zon[(zon['Date']==row['date']) & (zon['Location']==z)]
            mins = sub['Time'].dt.hour*60 + sub['Time'].dt.minute
            n = int(len(sub[(mins >= s_min) & (mins <= e_min)]))
            zone_counts.append(n)
            max_val = max(max_val, n)
        matrix_data.append(zone_counts)

    for i, (row, zcounts) in enumerate(zip(all_act_rows, matrix_data)):
        rr = mtr+2+i; rh(ws, rr, 17); alt = (i%2==0)
        # Activity name cell
        nc = ws.cell(rr, 2, f'{row["date"][-5:]}  {row["name"]}')
        nc.font = fn(8, bold=True, color=AMB); nc.fill = F(YLW if alt else WHT)
        nc.alignment = al('left', indent=1)
        nc.border = Border(right=Side(style='thin', color=BDR),
                           bottom=Side(style='thin', color=BDR))
        # Zone count cells — heatmap coloring
        for j, n in enumerate(zcounts):
            mc = ws.cell(rr, 3+j, n if n>0 else '')
            mc.alignment = al('center')
            mc.border = Border(right=Side(style='thin', color=BDR),
                               bottom=Side(style='thin', color=BDR))
            intensity = n / max_val if max_val > 0 else 0
            if   intensity >= 0.70: mc.fill = F('1A5C38'); mc.font = fn(9, bold=True, color=WHT)
            elif intensity >= 0.45: mc.fill = F('2E8B57'); mc.font = fn(9, bold=True, color=WHT)
            elif intensity >= 0.20: mc.fill = F('A8D8A8'); mc.font = fn(9, color=BLK)
            elif intensity >  0.01: mc.fill = F('DFF0D8'); mc.font = fn(9, color=MUT)
            else:                   mc.fill = F(GRY);      mc.font = fn(9, color=BDR)

    gap(ws, mtr+2+len(all_act_rows))

    # ── 📊 Activity Visits Comparison (chart) ─────────────────────
    acr = mtr+3+len(all_act_rows)
    sec_hdr(ws, acr, '📊  Activity Visits Comparison  ·  Visitors / Passersby / Zone Visits')
    for ri in range(acr+1, acr+14): rh(ws, ri, 14)

    act_chart_rows = [(r['name'][:24], r['visitors'], r['passersby'], r['zone_visits'])
                      for r in all_act_rows]
    write_hidden_data(ws, acr+1, 14,
                      ['Activity','Visitors','Passersby','Zone Visits'],
                      act_chart_rows)
    make_bar_chart(ws, f'B{acr+1}', acr+1, 14, len(act_chart_rows), 3,
                   [CH_V, CH_P, CH_U], bar_type='bar', height=12)
    gap(ws, acr+14)

    footer(ws, acr+15)



# ══════════════════════════════════════════════════════════════════
# ZONE DASHBOARD SHEET
# ══════════════════════════════════════════════════════════════════
def build_zone_sheet(wb, df):
    ws = setup_ws(wb, 'Zone Dashboard', tab_color=BLU)

    page_header(ws, 1,
        f'{EVENT_NAME}  —  Zone Traffic Dashboard',
        f'All Zones  ·  Traffic / Heatmap / Activity Matrix  ·  {SYSTEM_CREDIT}',
        f'{EVENT_LABELS[0][:6]} – {EVENT_LABELS[-1][:6]} {EVENT_DATES[-1][-4:]}')

    ent_all = df[(df['Type']=='Entrance') & (df['Event']=='in')]
    zon_all = df[(df['Type']=='Zone')     & (df['Event']=='in')]

    # Per-day slices  {date_str: (ent_d, zon_d)}
    day_slices = {}
    for ds in EVENT_DATES:
        day_slices[ds] = (
            df[(df['Date']==ds) & (df['Type']=='Entrance') & (df['Event']=='in')],
            df[(df['Date']==ds) & (df['Type']=='Zone')     & (df['Event']=='in')],
        )

    # Zone ordering: by total visits desc, consistent across days
    stats_all = (zon_all.groupby('Location')
                 .agg(Visits=('No','count'),
                      Unique=('BodyID', lambda s: (_non_staff(zon_all.loc[s.index])['BodyID'].nunique()
                                                  if 'CustomerType' in zon_all.columns else s.nunique())))
                 .reset_index().sort_values('Visits', ascending=False))
    zones_in_data = stats_all['Location'].tolist()
    tv_all = int(stats_all['Visits'].sum())

    # ── KPI cards ─────────────────────────────────────────────────
    gap(ws, 4, 8); rh(ws, 5, 16); rh(ws, 6, 28); rh(ws, 7, 14); gap(ws, 8, 8)

    top_z   = zones_in_data[0] if zones_in_data else '—'
    top_v   = int(stats_all.iloc[0]['Visits']) if len(stats_all) > 0 else 0
    top_z_s = top_z.replace('TV Touchscreen ','TV-').replace(' Total','')

    kpi(ws, 5,  2, 'TOTAL ZONE VISITS',    tv_all,
        'All zones combined',      BLU)
    kpi(ws, 5,  4, 'UNIQUE ZONE VISITORS', int(uv_count(zon_all)),
        'Distinct visitors',       GRN)
    kpi(ws, 5,  6, 'ACTIVE ZONES',         len(zones_in_data),
        'Zones with traffic',      NAV)
    kpi(ws, 5,  8, 'TOP ZONE',             top_z_s,
        f'{top_v:,} visits',       AMB)
    kpi(ws, 5, 10, 'BOOTH VISITORS',       len(ent_all),
        'Via entrance gates',      MUT)

    # ══════════════════════════════════════════════════════════════
    # 📊  ZONE TRAFFIC SUMMARY — per zone, per day as separate rows
    # ══════════════════════════════════════════════════════════════
    r = 9
    sec_hdr(ws, r, '📊  Zone Traffic Summary — All Zones  (แยกตามวัน)')

    # Helper: top age group label
    AGE_ORDER = ['Young Adults (19-35 yrs)', 'Middle-Aged (36-55 yrs)',
                 'Juvenile (0-18 yrs)', 'Seniors (55+ yrs)']
    AGE_SHORT = {'Young Adults (19-35 yrs)': 'Young Adults (19-35)',
                 'Middle-Aged (36-55 yrs)':  'Middle-Aged (36-55)',
                 'Juvenile (0-18 yrs)':       'Juvenile (0-18)',
                 'Seniors (55+ yrs)':         'Seniors (55+)'}
    DAY_COLORS = [BLU, GRN, AMB]
    DAY_LABELS = [f'Day {i+1}\n{l[:6]}' for i,l in enumerate(EVENT_LABELS)]

    # Table header
    tbl_hdr(ws, r+1, ['Zone / Area', 'Day', 'Total Visits', 'Unique Visitors',
                       'Engmt %', 'Male', 'Female', 'Top Age Group'])

    row_cursor = r + 2
    for zi, loc in enumerate(zones_in_data):
        alt_zone = (zi % 2 == 0)
        # Zone name cell (spans all 3 day rows)
        for di, (ds, dlbl, dcolor) in enumerate(zip(EVENT_DATES, DAY_LABELS, DAY_CSS_XL)):
            rr = row_cursor + di
            rh(ws, rr, 17)
            bg = GRY if alt_zone else WHT

            # Zone name — only first row of this zone block
            if di == 0:
                nc = ws.cell(rr, 2, loc)
                nc.font = fn(9, bold=True, color=BLK)
            else:
                nc = ws.cell(rr, 2, '')
                nc.font = fn(9, color=BLK)
            nc.fill = F(bg)
            nc.alignment = al('left', indent=1)
            nc.border = Border(right=Side(style='thin', color=BDR),
                               bottom=Side(style='thin', color=BDR))

            # Day badge cell
            _, z_d = day_slices[ds]
            z_loc = z_d[z_d['Location']==loc]
            v_d  = len(z_loc)
            uv_d = uv_count(z_loc)
            m_d  = int((z_loc['Gender']=='Male').sum())
            f_d  = int((z_loc['Gender']=='Female').sum())
            eng  = f'{uv_d/v_d*100:.1f}%' if v_d > 0 else '—'
            age_vc = z_loc['AgeGroup'].value_counts()
            top_age = AGE_SHORT.get(age_vc.index[0], age_vc.index[0]) if len(age_vc) > 0 else '—'

            dc = ws.cell(rr, 3, dlbl.replace('\n', ' '))
            dc.font = fn(8, bold=True, color=WHT)
            dc.fill = F(dcolor)
            dc.alignment = al('center')
            dc.border = Border(right=Side(style='thin', color=BDR),
                               bottom=Side(style='thin', color=BDR))

            vals = [v_d or '—', uv_d or '—', eng, m_d or '—', f_d or '—', top_age]
            for ci, val in enumerate(vals):
                cc = ws.cell(rr, 4+ci, val)
                cc.font = fn(9, color=(MUT if val == '—' else BLK))
                cc.fill = F(bg)
                cc.alignment = al('center')
                cc.border = Border(right=Side(style='thin', color=BDR),
                                   bottom=Side(style='thin', color=BDR))
                # Engagement % color
                if ci == 2 and val != '—':
                    pct_val = float(val.replace('%',''))
                    if   pct_val >= 65: cc.font = fn(9, bold=True, color=GRN)
                    elif pct_val >= 55: cc.font = fn(9, bold=True, color=AMB)
                    else:               cc.font = fn(9, bold=True, color=RED)

        # Separator line after each zone block
        for c_idx in range(2, 10):
            ws.cell(row_cursor+2, c_idx).border = Border(
                right=Side(style='thin', color=BDR),
                bottom=Side(style='medium', color=BDR))

        row_cursor += 3

    # Grand total row
    tbl_row(ws, row_cursor,
            ['TOTAL (All Days)', '', tv_all, int(uv_count(zon_all)),
             f"{int(uv_count(zon_all))/tv_all*100:.1f}%" if tv_all>0 else '—',
             int((zon_all['Gender']=='Male').sum()),
             int((zon_all['Gender']=='Female').sum()), ''],
            total=True)
    gap(ws, row_cursor+1)

    # ══════════════════════════════════════════════════════════════
    # 🕐  HOURLY ZONE TRAFFIC HEATMAP — per day (3 blocks)
    # ══════════════════════════════════════════════════════════════
    HOURS = list(range(9, 20))   # 09–19 covers Day1 closing at 19:xx

    def write_heatmap_block(ws, start_r, zon_d, day_label, day_color, zones):
        """Write one day's heatmap block. Returns next row."""
        # Mini day-header
        rh(ws, start_r, 18)
        ws.merge_cells(start_row=start_r, start_column=2,
                       end_row=start_r,   end_column=2+len(HOURS)+1)
        dh = ws.cell(start_r, 2, f'  {day_label}')
        dh.font = fn(9, bold=True, color=WHT)
        dh.fill = F(day_color)
        dh.alignment = al('left')
        dh.border = Border(bottom=Side(style='thin', color=BDR))

        # Column headers: Zone | 09:00 … | Total
        rh(ws, start_r+1, 22)
        for ci, lbl in enumerate(['Zone / Area'] + [f'{h:02d}:00' for h in HOURS] + ['Total']):
            c = ws.cell(start_r+1, 2+ci, lbl)
            c.font = fn(9, bold=True, color=WHT)
            c.fill = F(NAV)
            c.alignment = al('center', wrap=True)
            c.border = Border(right=Side(style='thin', color='336699'),
                              bottom=Side(style='medium', color=day_color))

        if len(zon_d) == 0:
            rh(ws, start_r+2, 16)
            ws.merge_cells(start_row=start_r+2, start_column=2,
                           end_row=start_r+2,   end_column=2+len(HOURS)+1)
            nc = ws.cell(start_r+2, 2, '  ⏳  No data for this day — pending upload')
            nc.font = fn(9, italic=True, color=MUT)
            nc.fill = F(GRY); nc.alignment = al('left')
            return start_r + 3

        zon_hr = zon_d.groupby(['Location','Hour']).size().unstack(fill_value=0)
        max_v  = zon_hr.values.max() if len(zon_hr) > 0 and zon_hr.values.max() > 0 else 1

        for i, loc in enumerate(zones):
            rr  = start_r + 2 + i
            alt = (i % 2 == 0)
            rh(ws, rr, 17)

            nc = ws.cell(rr, 2, loc)
            nc.font = fn(9, color=BLK); nc.fill = F(GRY if alt else WHT)
            nc.alignment = al('left', indent=1)
            nc.border = Border(right=Side(style='thin', color=BDR),
                               bottom=Side(style='thin', color=BDR))

            row_tot = 0
            for j, h in enumerate(HOURS):
                n = int(zon_hr.loc[loc, h]) \
                    if loc in zon_hr.index and h in zon_hr.columns else 0
                row_tot += n
                mc = ws.cell(rr, 3+j, n if n > 0 else '')
                mc.alignment = al('center')
                mc.border = Border(right=Side(style='thin', color=BDR),
                                   bottom=Side(style='thin', color=BDR))
                intensity = n / max_v
                if   intensity >= 0.80: mc.fill = F('084594'); mc.font = fn(9, bold=True,  color=WHT)
                elif intensity >= 0.60: mc.fill = F('2171B5'); mc.font = fn(9, bold=True,  color=WHT)
                elif intensity >= 0.40: mc.fill = F('6BAED6'); mc.font = fn(9,             color=BLK)
                elif intensity >= 0.20: mc.fill = F('C6DBEF'); mc.font = fn(9,             color=BLK)
                elif intensity >  0.00: mc.fill = F('EFF6FB'); mc.font = fn(9,             color=MUT)
                else:                   mc.fill = F(GRY if alt else WHT); mc.font = fn(9,  color=BDR)

            tc = ws.cell(rr, 3+len(HOURS), row_tot if row_tot > 0 else '')
            tc.font = fn(9, bold=True, color=NAV); tc.fill = F(TAB)
            tc.alignment = al('center')
            tc.border = Border(right=Side(style='thin', color=BDR),
                               bottom=Side(style='thin', color=BDR))

        # Total row for this day
        gtr = start_r + 2 + len(zones)
        rh(ws, gtr, 17)
        tc0 = ws.cell(gtr, 2, 'TOTAL')
        tc0.font = fn(9, bold=True, color=NAV); tc0.fill = F(TAB)
        tc0.alignment = al('left', indent=1)
        tc0.border = Border(right=Side(style='thin', color=BDR),
                            bottom=Side(style='thin', color=BDR))
        day_tot = 0
        for j, h in enumerate(HOURS):
            n = int(len(zon_d[zon_d['Hour']==h]))
            day_tot += n
            gc = ws.cell(gtr, 3+j, n if n > 0 else '')
            gc.font = fn(9, bold=True, color=NAV); gc.fill = F(TAB)
            gc.alignment = al('center')
            gc.border = Border(right=Side(style='thin', color=BDR),
                               bottom=Side(style='thin', color=BDR))
        gc_t = ws.cell(gtr, 3+len(HOURS), day_tot if day_tot > 0 else '')
        gc_t.font = fn(9, bold=True, color=NAV); gc_t.fill = F(TAB)
        gc_t.alignment = al('center')
        gc_t.border = Border(right=Side(style='thin', color=BDR),
                             bottom=Side(style='thin', color=BDR))

        return gtr + 1   # next row after this block

    hr_start = row_cursor + 2
    sec_hdr(ws, hr_start, '🕐  Hourly Zone Traffic Heatmap  ·  แยกตามวัน')

    hm_r = hr_start + 1
    for di, (ds, dcolor) in enumerate(zip(EVENT_DATES, DAY_CSS_XL)):
        _, z_d = day_slices[ds]
        dlbl = f'Day {di+1}  ·  {EVENT_LABELS[di]}'
        hm_r = write_heatmap_block(ws, hm_r, z_d, dlbl, dcolor, zones_in_data)
        # small gap between day blocks
        if di < 2:
            gap(ws, hm_r, 4); hm_r += 1

    # Heatmap legend
    rh(ws, hm_r, 14)
    ws.merge_cells(start_row=hm_r, start_column=2, end_row=hm_r, end_column=11)
    lc = ws.cell(hm_r, 2,
        '  Heatmap intensity:  ░ Very Low   ▒ Low   ▓ Medium   █ High   ██ Peak'
        '     (intensity relative to each day\'s own peak)')
    lc.font = fn(8, italic=True, color=MUT)
    lc.fill = F(GRY); lc.alignment = al('left')
    gap(ws, hm_r+1)

    # ══════════════════════════════════════════════════════════════
    # 🗺️  ACTIVITY × ZONE TRAFFIC MATRIX — per day, rows grouped
    # ══════════════════════════════════════════════════════════════

    # Pre-compute all activity rows with zone breakdowns
    all_act_rows = []
    for ds in EVENT_DATES:
        e_d, z_d = day_slices[ds]
        for t_s, t_e, name in ACTIVITIES.get(ds, []):
            sh, sm = int(t_s[:2]), int(t_s[3:])
            eh, em = int(t_e[:2]), int(t_e[3:])
            s_min, e_min = sh*60+sm, eh*60+em
            def in_w(sub, _ds=ds, _sm=s_min, _em=e_min):
                mins = sub['Time'].dt.hour*60 + sub['Time'].dt.minute
                return sub[(sub['Date']==_ds) & (mins >= _sm) & (mins <= _em)]
            v_w = in_w(e_d)
            z_w = in_w(z_d)
            zone_counts = {}
            for z in zones_in_data:
                sub  = z_w[z_w['Location']==z]
                zone_counts[z] = len(sub)
            all_act_rows.append({
                'date': ds, 'time': f'{t_s}–{t_e}', 'name': name,
                'visitors': len(v_w), 'unique': int(uv_count(v_w)),
                'zone_counts': zone_counts,
            })

    mtr = hm_r + 2
    sec_hdr(ws, mtr, '🗺️  Activity × Zone Traffic Matrix  ·  แยกตามวัน')

    short_zones = [z.replace('TV Touchscreen ','TV-').replace(' Total','')[:11]
                   for z in zones_in_data]

    # Header
    rh(ws, mtr+1, 30)
    for ci, lbl in enumerate(['Activity'] + short_zones + ['Visitors']):
        c = ws.cell(mtr+1, 2+ci, lbl)
        c.font = fn(8, bold=True, color=WHT)
        c.fill = F(NAV)
        c.alignment = al('center', wrap=True)
        c.border = Border(right=Side(style='thin', color='336699'),
                          bottom=Side(style='medium', color=BLU))

    # Global max for heatmap scale
    max_val = max((n for row in all_act_rows for n in row['zone_counts'].values()), default=1)
    max_val = max(max_val, 1)

    rr = mtr + 2
    prev_date = None
    for row in all_act_rows:
        # Day separator header when date changes
        if row['date'] != prev_date:
            di   = EVENT_DATES.index(row['date'])
            dlbl = f'  Day {di+1}  ·  {EVENT_LABELS[di]}'
            dclr = DAY_CSS_XL[di]
            rh(ws, rr, 16)
            ws.merge_cells(start_row=rr, start_column=2,
                           end_row=rr,   end_column=2+len(zones_in_data)+1)
            dc = ws.cell(rr, 2, dlbl)
            dc.font = fn(9, bold=True, color=WHT)
            dc.fill = F(dclr); dc.alignment = al('left')
            dc.border = Border(bottom=Side(style='thin', color=BDR))
            rr += 1
            prev_date = row['date']

        rh(ws, rr, 17)
        di   = EVENT_DATES.index(row['date'])
        dclr = DAY_CSS_XL[di]
        alt  = (rr % 2 == 0)

        # Activity name
        nc = ws.cell(rr, 2, f'{row["time"]}  {row["name"]}')
        nc.font = fn(8, bold=True, color=dclr)
        nc.fill = F(YLW if alt else WHT)
        nc.alignment = al('left', indent=1)
        nc.border = Border(right=Side(style='thin', color=BDR),
                           bottom=Side(style='thin', color=BDR))

        # Zone count cells
        for j, z in enumerate(zones_in_data):
            n = row['zone_counts'].get(z, 0)
            mc = ws.cell(rr, 3+j, n if n > 0 else '')
            mc.alignment = al('center')
            mc.border = Border(right=Side(style='thin', color=BDR),
                               bottom=Side(style='thin', color=BDR))
            intensity = n / max_val
            if   intensity >= 0.70: mc.fill = F('1A5C38'); mc.font = fn(9, bold=True, color=WHT)
            elif intensity >= 0.45: mc.fill = F('2E8B57'); mc.font = fn(9, bold=True, color=WHT)
            elif intensity >= 0.20: mc.fill = F('A8D8A8'); mc.font = fn(9,            color=BLK)
            elif intensity >  0.01: mc.fill = F('DFF0D8'); mc.font = fn(9,            color=MUT)
            else:                   mc.fill = F(GRY if alt else WHT); mc.font = fn(9, color=BDR)

        # Visitors summary col
        vc = ws.cell(rr, 3+len(zones_in_data), row['visitors'] if row['visitors'] > 0 else '—')
        vc.font = fn(9, bold=True, color=BLK); vc.fill = F(GRY if alt else WHT)
        vc.alignment = al('center')
        vc.border = Border(right=Side(style='thin', color=BDR),
                           bottom=Side(style='thin', color=BDR))
        rr += 1

    gap(ws, rr)

    # ══════════════════════════════════════════════════════════════
    # 📊  ACTIVITY VISITS COMPARISON CHART — per day
    # ══════════════════════════════════════════════════════════════
    acr = rr + 1
    sec_hdr(ws, acr, '📊  Activity Visits Comparison  ·  Visitors per Activity (แยกตามวัน)')
    for ri in range(acr+1, acr+16): rh(ws, ri, 14)

    act_chart_rows = [(f'D{EVENT_DATES.index(r["date"])+1} {r["name"][:18]}',
                       r['visitors'], sum(r['zone_counts'].values()))
                      for r in all_act_rows]
    write_hidden_data(ws, acr+1, 15,
                      ['Activity', 'Visitors', 'Zone Visits'],
                      act_chart_rows)
    make_bar_chart(ws, f'B{acr+1}', acr+1, 15, len(act_chart_rows), 2,
                   [CH_V, CH_U], bar_type='bar', height=14)
    gap(ws, acr+16)

    footer(ws, acr+17)



# ══════════════════════════════════════════════════════════════════
# PER-DAY DASHBOARD
# ══════════════════════════════════════════════════════════════════
def build_day_sheet(wb, df, date_str, day_num, date_label):
    _day_palette_tab = ['1E6B9A','196B45','8B4500','5B217F','8B2020','006B6B','8B5E00']
    _day_palette_acc = [BLU, GRN, AMB, '7B2D8B', 'B02020', '006B6B', '8B5E00']
    tab_c = _day_palette_tab[(day_num-1) % len(_day_palette_tab)]
    ws = setup_ws(wb, f'Day {day_num}  ·  {date_label}', tab_color=tab_c)

    accent = _day_palette_acc[(day_num-1) % len(_day_palette_acc)]
    d   = df[df['Date']==date_str]
    ent = d[(d['Type']=='Entrance') & (d['Event']=='in')]
    psb = d[d['Type']=='Passerby']
    zon = d[(d['Type']=='Zone') & (d['Event']=='in')]
    has = len(d) > 0

    page_header(ws, 1, f'{EVENT_NAME}  —  Day {day_num} Traffic Report',
        f'{ORGANIZER} {VENUE_TYPE}  ·  {VENUE}  ·  {SYSTEM_CREDIT}',
        date_label)

    if not has:
        gap(ws, 4)
        ws.merge_cells(start_row=5, start_column=2, end_row=5, end_column=11)
        c = ws.cell(5, 2, f'⏳  No data available for Day {day_num} ({date_label}) — pending upload')
        c.font = fn(11, italic=True, color=MUT); c.alignment = al('center')
        footer(ws, 8); return

    # KPIs
    gap(ws, 4, 8); rh(ws, 5, 16); rh(ws, 6, 28); rh(ws, 7, 14); gap(ws, 8, 8)
    tv  = len(ent); uv = uv_count(ent); tp = len(psb) + len(ent)
    ph  = ent.groupby('Hour').size().idxmax() if tv>0 else 0
    pv  = ent.groupby('Hour').size().max()    if tv>0 else 0
    mn  = int((d['Gender']=='Male').sum()); fn2 = int((d['Gender']=='Female').sum())
    kpi(ws, 5,  2, 'BOOTH VISITORS',    tv,  'Via entrance gates',             accent)
    kpi(ws, 5,  4, 'UNIQUE VISITORS',   uv,  'Distinct individuals',           GRN)
    kpi(ws, 5,  6, 'PASSERSBY',         tp,  'Walk-past traffic',              AMB)
    kpi(ws, 5,  8, 'PEAK HOUR',         pv,  f'at {ph:02d}:00 — busiest hour', RED)
    kpi(ws, 5, 10, 'GENDER SPLIT',
        f'{mn/(mn+fn2)*100:.0f}% M' if (mn+fn2)>0 else '—',
        f'{mn:,} Male  /  {fn2:,} Female', MUT)

    acts = ACTIVITIES.get(date_str, [])
    r = hourly_section(ws, 9, ent, psb, acts, dc=15)
    _p = PROFILE_CONFIG.get(EVENT_PROFILE, PROFILE_CONFIG['full'])
    if _p['has_zones']:
        r = zone_section(ws, r, zon, dc=15)
    r = demo_section(ws, r, d)

    # Activity table for this day
    if acts:
        sec_hdr(ws, r, '│ BOOTH ACTIVITY ANALYTICS  ·  Traffic During Activities')
        tbl_hdr(ws, r+1, ['Time','Activity','Visitors','Unique','Passersby','Zone Visits'])
        for i,(ts,te,name) in enumerate(acts):
            sh2,sm2=int(ts[:2]),int(ts[3:]);  eh2,em2=int(te[:2]),int(te[3:])
            sm_=sh2*60+sm2; em_=eh2*60+em2
            def iw(sub):
                mins=sub['Time'].dt.hour*60+sub['Time'].dt.minute
                return sub[(mins>=sm_)&(mins<=em_)]
            vw=iw(ent); pw=iw(psb); zw=iw(zon)  # ent already filtered to Event='in'
            tbl_row(ws, r+2+i,
                    [f'{ts}–{te}', name, len(vw), int(uv_count(vw)),
                     len(pw), len(zw)],
                    alt=(i%2==0), hi=True)
        r = r+3+len(acts)
        gap(ws, r)
        r += 1

    footer(ws, r)


# ══════════════════════════════════════════════════════════════════
# OVERALL SUMMARY SHEET
# ══════════════════════════════════════════════════════════════════
def build_overall_sheet(wb, df):
    ws = setup_ws(wb, 'Overall Summary', tab_color=NAV)

    ent = df[(df['Type']=='Entrance') & (df['Event']=='in')]; psb = df[df['Type']=='Passerby']
    zon = df[(df['Type']=='Zone') & (df['Event']=='in')]

    page_header(ws, 1, f'{EVENT_NAME}  —  {len(EVENT_DATES)}-Day Booth Traffic Summary',
        f'{ORGANIZER} {VENUE_TYPE}  ·  {VENUE}  ·  {EVENT_LABELS[0]} – {EVENT_LABELS[-1]}  ·  {SYSTEM_CREDIT}',
        f'{EVENT_LABELS[0][:6]} – {EVENT_LABELS[-1][:6]} {EVENT_DATES[-1][-4:]}')

    gap(ws, 4, 8); rh(ws, 5, 16); rh(ws, 6, 28); rh(ws, 7, 14); gap(ws, 8, 8)
    tv  = len(ent); uv = uv_count(ent); tp = len(psb) + len(ent)
    ph  = ent.groupby('Hour').size().idxmax() if tv>0 else 0
    pv  = ent.groupby('Hour').size().max()    if tv>0 else 0
    mn  = int((df['Gender']=='Male').sum()); fn2 = int((df['Gender']=='Female').sum())

    kpi(ws, 5,  2, 'TOTAL VISITORS',   tv,  f'{df["Date"].nunique()} day(s) with data', BLU)
    kpi(ws, 5,  4, 'UNIQUE VISITORS',  uv,  'Distinct individuals',                     GRN)
    kpi(ws, 5,  6, 'TOTAL PASSERSBY',  tp,  'Walk-past traffic',                        AMB)
    kpi(ws, 5,  8, 'PEAK HOUR',        pv,  f'at {ph:02d}:00',                          RED)
    kpi(ws, 5, 10, 'GENDER SPLIT',
        f'{mn/(mn+fn2)*100:.0f}% M' if (mn+fn2)>0 else '—',
        f'{mn:,} M  /  {fn2:,} F', MUT)

    # Day comparison table
    r = 9
    sec_hdr(ws, r, '1 │ DAY-BY-DAY COMPARISON')
    tbl_hdr(ws, r+1, ['Date','Day','Visitors','Unique','Passersby','Zone Visits','Peak Hour','Peak Traffic'])
    for i,(ds,dl,dn) in enumerate(zip(EVENT_DATES, EVENT_LABELS, [1,2,3])):
        e=df[(df['Date']==ds)&(df['Type']=='Entrance')&(df['Event']=='in')]
        p=df[(df['Date']==ds)&(df['Type']=='Passerby')]
        z=df[(df['Date']==ds)&(df['Type']=='Zone')&(df['Event']=='in')]
        psb_total = len(p) + len(e)
        has = len(e)>0
        ph2,pv2 = (e.groupby('Hour').size().idxmax(),
                   e.groupby('Hour').size().max()) if has else (0,0)
        tbl_row(ws, r+2+i,
                [dl, f'Day {dn}',
                 len(e) if has else 'No Data',
                 uv_count(e) if has else '—',
                 psb_total if has else '—',
                 len(z) if has else '—',
                 f'{ph2:02d}:00' if has else '—',
                 pv2 if has else '—'],
                alt=(i%2==0))
        if not has:
            for ci in range(2,10): ws.cell(r+2+i,ci).font=fn(9,color=MUT,italic=True)

    tbl_row(ws, r+5, [f'{len(EVENT_DATES)}-Day Total','—', tv, uv, tp, len(zon),'—','—'], total=True)  # tp = psb+visitors
    gap(ws, r+6)

    # Build per-day dicts for multi-day hourly table
    all_days_data = {}
    for ds in EVENT_DATES:
        e_d = df[(df['Date']==ds) & (df['Type']=='Entrance') & (df['Event']=='in')]
        p_d = df[(df['Date']==ds) & (df['Type']=='Passerby')]
        all_days_data[ds] = (e_d, p_d)
    r2 = hourly_section(ws, r+7, ent, psb, [], dc=15, all_days_data=all_days_data)
    _p = PROFILE_CONFIG.get(EVENT_PROFILE, PROFILE_CONFIG['full'])
    if _p['has_zones']:
        r2 = zone_section(ws, r2, zon, dc=15)
    r2 = demo_section(ws, r2, df)
    footer(ws, r2)


# ══════════════════════════════════════════════════════════════════
# RAW DATA SHEET
# ══════════════════════════════════════════════════════════════════
def build_raw_sheet(wb, df, date_str, date_label, day_num):
    ws = setup_ws(wb, f'Raw Data · Day {day_num}', tab_color='888888')
    raw_widths = [1.2, 7, 20, 16, 28, 16, 24, 8, 14]
    for i,w in enumerate(raw_widths,1):
        ws.column_dimensions[get_column_letter(i)].width = w

    page_header(ws, 1, f'{EVENT_NAME}  —  Raw Data  ·  Day {day_num}',
        f'Source Records  ·  {date_label}  ·  All sensor events', date_label)

    d = df[df['Date']==date_str].sort_values('Time').reset_index(drop=True)

    gap(ws, 4, 6)
    hdrs = ['#','Time','Location','Type','Age Group','Gender','Body ID (abbrev.)']
    tbl_hdr(ws, 5, hdrs, cs=2)

    for i,(_,row) in enumerate(d.iterrows()):
        r = 6+i; rh(ws, r, 16)
        bg = GRY if i%2==0 else WHT
        bid = str(row['BodyID']); bid_short = bid[:12]+'…' if len(bid)>12 else bid
        vals = [i+1,
                row['Time'].strftime('%H:%M:%S') if hasattr(row['Time'],'strftime') else str(row['Time']),
                row['Location'], row['Type'], row['AgeGroup'],
                row['Gender'], bid_short]
        for j,v in enumerate(vals):
            c = ws.cell(r, j+2, v)
            c.font = fn(8, color=BLK); c.fill = F(bg)
            c.alignment = al('left' if j in [2,3,4] else 'center', indent=1 if j in [2,3,4] else 0)
            c.border = Border(right=Side(style='thin',color=BDR),
                              bottom=Side(style='thin',color=BDR))

    sr = 6+len(d)+1; rh(ws, sr, 20)
    ws.merge_cells(start_row=sr, start_column=2, end_row=sr, end_column=8)
    sc = ws.cell(sr, 2,
        f'Total: {len(d):,}  │  Entrance: {len(d[d["Type"]=="Entrance"]):,}  '
        f'│  Passerby: {len(d[d["Type"]=="Passerby"]):,}  '
        f'│  Zone: {len(d[d["Type"]=="Zone"]):,}')
    sc.font = fn(9, bold=True, color=NAV); sc.fill = F(TAB)
    sc.alignment = al('left', indent=2)
    sc.border = Border(top=Side(style='medium', color=NAV))

    ws.freeze_panes = 'B6'
    ws.auto_filter.ref = f'B5:{get_column_letter(len(hdrs)+1)}5'
    footer(ws, sr+2)


# ══════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════


"""
Generates the complete Event Analytics Dashboard
replacing generate_heatmap_html() in otc_dashboard_v4.py
"""

def generate_full_html(df, output_path):
    import json, re as _re, os

    # ── Data prep ──────────────────────────────────────────────────
    ent = df[(df['Type']=='Entrance') & (df['Event']=='in')]
    psb = df[df['Type']=='Passerby']
    zon = df[(df['Type']=='Zone')     & (df['Event']=='in')]

    HOURS     = list(range(DISPLAY_HOURS_START, DISPLAY_HOURS_END))
    HOUR_LBLS = [f'{h:02d}:00' for h in HOURS]

    EVENT_DATES_  = EVENT_DATES   # from config
    EVENT_LABELS_ = EVENT_LABELS  # from config
    DAY_SHORT     = [l.replace(' 2026','').replace(' 20','')[:6].strip() for l in EVENT_LABELS]  # auto-shorten
    DAY_CSS       = DAY_COLORS[:len(EVENT_DATES)]  # from config

    ACTIVITIES_   = ACTIVITIES   # from config

    ZONE_ORDER_ = ZONE_ORDER  # from config
    ZONE_ABBREV_  = ZONE_ABBREV   # from config
    zones_avail = zon['Location'].unique().tolist()
    ZONES = [z for z in ZONE_ORDER_ if z in zones_avail] + \
            [z for z in zones_avail if z not in ZONE_ORDER_]

    # ── Helper functions ───────────────────────────────────────────
    def h_vals(sub, mode='count'):
        return [int(len(sub[sub['Hour']==h])) if mode=='count'
                else int(uv_count(sub[sub['Hour']==h]))
                for h in HOURS]

    def fmt(n):
        return f'{n:,}' if isinstance(n, int) and n > 0 else ('—' if n == 0 else str(n))

    def pct(a, b):
        return f'{a/b*100:.1f}%' if b > 0 else '—'

    def peak_info(vals):
        mx = max(vals)
        if mx == 0: return '—'
        return f'{HOUR_LBLS[vals.index(mx)]} ({mx:,})'

    def heat_cls(v, mx):
        if v == 0 or mx == 0: return 'e'
        r = v/mx
        return 'h4' if r>=.8 else 'h3' if r>=.55 else 'h2' if r>=.3 else 'h1' if r>=.1 else 'h0'

    def trim_n(all_vals_list):
        mx = 9
        for vals in all_vals_list:
            for i,v in enumerate(vals):
                if v > 0: mx = max(mx, i)
        return mx + 1

    def in_window(sub, ds, sm, em):
        mins = sub['Time'].dt.hour*60 + sub['Time'].dt.minute
        return sub[(sub['Date']==ds) & (mins>=sm) & (mins<=em)]

    # ── Compute dwell ──────────────────────────────────────────────
    zone_io = df[(df['Type']=='Zone') & df['Event'].isin(['in','out'])].sort_values(['BodyID','Location','Time'])
    dwell_sessions = []
    for (bid,loc),grp in zone_io.groupby(['BodyID','Location'], sort=False):
        ins  = grp[grp['Event']=='in']['Time'].tolist()
        outs = grp[grp['Event']=='out']['Time'].tolist()
        oi = 0
        for t_in in ins:
            while oi < len(outs) and outs[oi] <= t_in: oi += 1
            if oi < len(outs):
                sec = (outs[oi] - t_in).total_seconds()
                if DWELL_MIN_SEC <= sec <= DWELL_MAX_SEC:  # 10 sec – 60 min (exhibition context)
                    dwell_sessions.append({'loc': loc, 'bid': bid, 'min': sec/60,
                                           'date': t_in.strftime('%Y-%m-%d')})
                oi += 1
    import pandas as pd
    ddf = pd.DataFrame(dwell_sessions)

    # ── Generate insights ──────────────────────────────────────────
    _insights = {}
    if _HAS_INSIGHTS:
        try:
            _insights = _gen_insights(
                df=df, profile=EVENT_PROFILE,
                event_dates=EVENT_DATES_,
                event_labels=EVENT_LABELS_,
                activities=ACTIVITIES_,
                zone_order=ZONE_ORDER_,
                ddf=ddf,
                dwell_threshold=ENGAGEMENT_THRESHOLD_SEC,
            )
        except Exception as _e:
            print(f'  ⚠ Insight engine error: {_e}')

    def _get_ins(category):
        return _insights.get(category, [])


    # ── HTML component builders ────────────────────────────────────
    def kpi_cards(*items):
        # items = (label, value, sub, accent_css)
        cards = ''
        for label, value, sub, acc in items:
            cards += (f'<div class="kpi" style="border-left-color:{acc}">'
                      f'<div class="kv">{value}</div>'
                      f'<div class="kl">{label}</div>'
                      f'<div class="ks">{sub}</div></div>')
        return f'<div class="krow">{cards}</div>'

    def sec_hdr(icon, title, sub=''):
        if sub:
            return (f'<div class="shdr"><span class="sico">{icon}</span>'
                    f'<div><span class="stit">{title}</span>'
                    f'<span class="ssub">{sub}</span></div></div>')
        return (f'<div class="shdr"><span class="sico">{icon}</span>'
                f'<div><span class="stit">{title}</span></div></div>')

    def insight_block(insights_list: list, max_show: int = 4) -> str:
        """
        Renders a compact insight panel above a section.
        Shows headline insights first, then key, then detail.
        max_show = max number of insight cards to display.
        """
        if not insights_list:
            return ''
        ordered = (
            [i for i in insights_list if i['level'] == 'headline'] +
            [i for i in insights_list if i['level'] == 'key'] +
            [i for i in insights_list if i['level'] == 'detail']
        )[:max_show]

        cards = []
        for ins in ordered:
            pos = ins.get('positive')
            border = '#1A7A45' if pos is True else ('#B02020' if pos is False else '#005B9A')
            bg     = '#F0FFF4' if pos is True else ('#FFF5F5' if pos is False else '#EFF6FB')
            cards.append(
                f'<div class="ins-card" style="border-left:3px solid {border};background:{bg}">'
                f'<span class="ins-icon">{ins["icon"]}</span>'
                f'<div class="ins-body">'
                f'<span class="ins-val" style="color:{border}">{ins["value"]}</span>'
                f'<span class="ins-en">{ins["en"]}</span>'
                f'<span class="ins-th">{ins["th"]}</span>'
                f'</div></div>'
            )
        return (
            '<div class="ins-panel">'
            + ''.join(cards)
            + '</div>'
        )

    def funnel_block(total_psb: int, total_v: int,
                     total_uv: int, engaged_uv: int,
                     insights_list: list = None) -> str:
        """Renders a visual funnel: Passersby → Visitors → Unique → Engaged."""
        steps = [
            ('Passersby',          total_psb,  '#6B8299', '👟'),
            ('Visitors (Entries)', total_v,    '#005B9A', '🚶'),
            ('Unique Visitors',    total_uv,   '#1A7A45', '👤'),
            ('Engaged (>60s)',     engaged_uv, '#B06000', '🔥'),
        ]
        html = ['<div class="funnel-wrap">']
        max_v = max(s[1] for s in steps) or 1
        for i, (label, val, color, icon) in enumerate(steps):
            w = max(20, round(val / max_v * 100))
            conv = ''
            if i > 0:
                prev = steps[i-1][1]
                pct = round(val / prev * 100, 1) if prev > 0 else 0
                conv = f'<span class="funnel-conv">{pct}% conversion</span>'
            html.append(
                f'<div class="funnel-step">'
                f'<div class="funnel-label">{icon} {label}</div>'
                f'<div class="funnel-bar-wrap">'
                f'<div class="funnel-bar" style="width:{w}%;background:{color}">'
                f'<span class="funnel-n">{val:,}</span></div>'
                f'{conv}</div></div>'
            )
        html.append('</div>')
        if insights_list:
            html.append(insight_block(insights_list, max_show=2))
        return '\n'.join(html)


    # ══════════════════════════════════════════════════════════════
    # ACTIVITY CORRELATION BLOCK
    # ══════════════════════════════════════════════════════════════
    def activity_correlation_block(act_rows: list, insights_list: list = None) -> str:
        if not act_rows:
            return ''
        html = ['<div class="act-corr-wrap">']
        max_v = max([max(a['before'], a['during'], a['after']) for a in act_rows] + [1])
        for a in act_rows:
            lift = a.get('lift_pct', 0)
            if lift >= 30:
                lift_color, lift_icon = '#1A7A45', '📈'
            elif lift >= 0:
                lift_color, lift_icon = '#B06000', '➡️'
            else:
                lift_color, lift_icon = '#B02020', '📉'
            wb = round(a['before'] / max_v * 100)
            wd = round(a['during'] / max_v * 100)
            wa = round(a['after']  / max_v * 100)
            html.append(
                f'<div class="act-row">'
                f'<div class="act-info">'
                f'<div class="act-name">{a["name"]}</div>'
                f'<div class="act-time">{a["date"]} &middot; {a["time"]}</div>'
                f'</div>'
                f'<div class="act-bars">'
                f'<div class="act-bar-group">'
                f'<span class="act-lbl">Before</span>'
                f'<div class="act-bar-wrap"><div class="act-bar" style="width:{wb}%;background:#B8D4F0">'
                f'<span>{a["before"]:,}</span></div></div></div>'
                f'<div class="act-bar-group">'
                f'<span class="act-lbl act-lbl-active">During</span>'
                f'<div class="act-bar-wrap"><div class="act-bar act-bar-during" style="width:{wd}%">'
                f'<span>{a["during"]:,}</span></div></div></div>'
                f'<div class="act-bar-group">'
                f'<span class="act-lbl">After</span>'
                f'<div class="act-bar-wrap"><div class="act-bar" style="width:{wa}%;background:#B8D4F0">'
                f'<span>{a["after"]:,}</span></div></div></div>'
                f'</div>'
                f'<div class="act-lift" style="color:{lift_color}">'
                f'<span class="act-lift-icon">{lift_icon}</span>'
                f'<span class="act-lift-val">{lift:+.0f}%</span>'
                f'</div>'
                f'</div>'
            )
        html.append('</div>')
        if insights_list:
            html.append(insight_block(insights_list, max_show=2))
        return '\n'.join(html)


    # ══════════════════════════════════════════════════════════════
    # EXECUTIVE SUMMARY BLOCK (Sprint 4)
    # ══════════════════════════════════════════════════════════════
    def exec_summary_block(kpi_cards_list, headline_insights, prog_score=None,
                            best_activity=None, top_zone=None, top_time=None):
        """
        High-level 1-page overview for C-level / client leadership.
        - kpi_cards_list: list of 4 tuples (title, value, sub, color)
        - headline_insights: list of insight dicts (top 3-5)
        - prog_score: {'pct': 22, 'winners': 2, 'total': 9} or None
        - best_activity: dict with name, during, lift_pct or None
        - top_zone: dict with name, visits, dwell_min or None
        - top_time: dict with hour, visitors, label or None
        """
        html = []
        html.append('<div class="exec-wrap">')

        # ── Top KPI row (big numbers) ──
        html.append('<div class="exec-kpi-row">')
        for title, value, sub, color in kpi_cards_list:
            html.append(
                f'<div class="exec-kpi" style="border-top:4px solid {color}">'
                f'<div class="exec-kpi-title">{title}</div>'
                f'<div class="exec-kpi-val" style="color:{color}">{value}</div>'
                f'<div class="exec-kpi-sub">{sub}</div>'
                f'</div>'
            )
        html.append('</div>')

        # ── Programme Effectiveness (if available) ──
        if prog_score:
            pct = prog_score['pct']
            score_color = '#1A7A45' if pct >= 50 else ('#B06000' if pct >= 30 else '#B02020')
            score_label = 'Excellent' if pct >= 70 else ('Good' if pct >= 50 else ('Fair' if pct >= 30 else 'Needs improvement'))
            html.append(
                f'<div class="exec-score" style="border-left:4px solid {score_color}">'
                f'<div class="exec-score-l">'
                f'<span class="exec-score-title">📊 Programme Effectiveness</span>'
                f'<span class="exec-score-sub">{prog_score["winners"]} of {prog_score["total"]} activities drove ≥+20% traffic lift</span>'
                f'</div>'
                f'<div class="exec-score-r">'
                f'<span class="exec-score-num" style="color:{score_color}">{pct}%</span>'
                f'<span class="exec-score-lbl" style="color:{score_color}">{score_label}</span>'
                f'</div>'
                f'</div>'
            )

        # ── 3-column highlights: Best Activity / Top Zone / Peak Time ──
        html.append('<div class="exec-highlights">')
        if best_activity:
            html.append(
                f'<div class="exec-hl"><div class="exec-hl-icon">🏆</div>'
                f'<div class="exec-hl-lbl">Best Activity</div>'
                f'<div class="exec-hl-val">{best_activity["name"]}</div>'
                f'<div class="exec-hl-sub">+{best_activity["lift_pct"]:.0f}% lift · {best_activity["during"]:,} visitors</div>'
                f'</div>'
            )
        if top_zone:
            html.append(
                f'<div class="exec-hl"><div class="exec-hl-icon">🗺️</div>'
                f'<div class="exec-hl-lbl">Most Visited Zone</div>'
                f'<div class="exec-hl-val">{top_zone["name"]}</div>'
                f'<div class="exec-hl-sub">{top_zone["visits"]:,} visits · {top_zone["dwell_min"]:.1f} min avg</div>'
                f'</div>'
            )
        if top_time:
            html.append(
                f'<div class="exec-hl"><div class="exec-hl-icon">⏰</div>'
                f'<div class="exec-hl-lbl">Peak Hour</div>'
                f'<div class="exec-hl-val">{top_time["hour"]:02d}:00</div>'
                f'<div class="exec-hl-sub">{top_time["visitors"]:,} visitors · {top_time["label"]}</div>'
                f'</div>'
            )
        html.append('</div>')

        # ── Headline insights (top 3-5) ──
        if headline_insights:
            html.append('<div class="exec-insights-title">💡 Key Insights</div>')
            html.append(insight_block(headline_insights, max_show=5))

        html.append('</div>')
        return '\n'.join(html)

    # ══════════════════════════════════════════════════════════════
    # TOP ZONE RANKING
    # ══════════════════════════════════════════════════════════════
    def top_zone_ranking(zone_stats: list, max_show: int = 5) -> str:
        if not zone_stats:
            return ''
        sorted_z = sorted(zone_stats, key=lambda x: x.get('visits', 0), reverse=True)[:max_show]
        if not sorted_z:
            return ''
        max_visits = max(z['visits'] for z in sorted_z) or 1
        max_dwell  = max([z.get('dwell_min', 0) for z in sorted_z] + [1])
        html = ['<div class="rank-wrap">']
        medals = ['🥇','🥈','🥉']
        for i, z in enumerate(sorted_z):
            medal = medals[i] if i < 3 else f'#{i+1}'
            v_pct = round(z['visits'] / max_visits * 100)
            d_pct = round(z.get('dwell_min', 0) / max_dwell * 100) if max_dwell > 0 else 0
            eng   = z.get('engagement_pct', 0)
            eng_color = '#1A7A45' if eng >= 30 else '#B06000'
            html.append(
                f'<div class="rank-row">'
                f'<div class="rank-medal">{medal}</div>'
                f'<div class="rank-name">{z["name"]}</div>'
                f'<div class="rank-metrics">'
                f'<div class="rank-metric"><span class="rank-ml">Visits</span>'
                f'<div class="rank-bar-wrap"><div class="rank-bar" style="width:{v_pct}%;background:#005B9A"></div>'
                f'<span class="rank-val">{z["visits"]:,}</span></div></div>'
                f'<div class="rank-metric"><span class="rank-ml">Dwell</span>'
                f'<div class="rank-bar-wrap"><div class="rank-bar" style="width:{d_pct}%;background:#2D7DD2"></div>'
                f'<span class="rank-val">{z.get("dwell_min",0):.1f}m</span></div></div>'
                f'<div class="rank-metric"><span class="rank-ml">Engmt</span>'
                f'<span class="rank-eng" style="color:{eng_color}">{eng:.0f}%</span></div>'
                f'</div></div>'
            )
        html.append('</div>')
        return '\n'.join(html)


    # ══════════════════════════════════════════════════════════════
    # SPONSOR VALUE BLOCK
    # ══════════════════════════════════════════════════════════════
    def sponsor_value_block(stage_zone: str, exposure_v: int, exposure_uv: int,
                            during_activity_v: int, avg_dwell_min: float,
                            engaged_pct: float,
                            insights_list: list = None) -> str:
        cards = [
            ('Total Exposure',  f'{exposure_v:,}',
             f'{exposure_uv:,} unique individuals · all visits to {stage_zone}',
             '#005B9A', '👁️'),
            ('Activity Audience', f'{during_activity_v:,}',
             'Visitors present during scheduled activities on stage',
             '#1A7A45', '🎪'),
            ('Avg Dwell',       f'{avg_dwell_min:.1f} min',
             f'Mean time visitors stayed in {stage_zone}',
             '#B06000', '⏱️'),
            ('Engaged Share',   f'{engaged_pct:.0f}%',
             f'Unique visitors with dwell > 60 seconds',
             '#7B2D8B', '🔥'),
        ]
        html = ['<div class="sponsor-grid">']
        for title, val, sub, color, icon in cards:
            html.append(
                f'<div class="sponsor-card" style="border-top:3px solid {color}">'
                f'<div class="sp-icon">{icon}</div>'
                f'<div class="sp-title">{title}</div>'
                f'<div class="sp-val" style="color:{color}">{val}</div>'
                f'<div class="sp-sub">{sub}</div>'
                f'</div>'
            )
        html.append('</div>')
        if insights_list:
            html.append(insight_block(insights_list, max_show=2))
        return '\n'.join(html)

        s = f'<span class="ssub">{sub}</span>' if sub else ''
        return (f'<div class="shdr"><span class="sico">{icon}</span>'
                f'<div><span class="stit">{title}</span>{s}</div></div>')

    def day_hdr(label, color):
        return f'<div class="dhdr" style="background:{color}">{label}</div>'

    def tbl_hdr(*cols):
        ths = ''.join(f'<th>{c}</th>' for c in cols)
        return f'<thead><tr>{ths}</tr></thead>'

    def tbl_row(*vals, alt=False, hi=False, total=False):
        cls = ' total' if total else (' hi' if hi else (' alt' if alt else ''))
        tds = ''.join(f'<td>{v}</td>' for v in vals)
        return f'<tr class="{cls.strip()}">{tds}</tr>'

    def heatmap_tbl(rows, gmax, sn):
        sh = HOUR_LBLS[:sn]
        t = ['<table class="hm"><thead><tr><th class="zth"></th>']
        t += [f'<th>{h}</th>' for h in sh]
        t += ['<th class="tot-th">Total</th></tr></thead><tbody>']
        for row in rows:
            vals = row['vals'][:sn]
            has  = any(v>0 for v in vals)
            dot_c = row.get("color", "#888")
            t.append(f'<tr><td class="zl"><span class="dot" style="background:{dot_c}"></span>{row["label"]}</td>')
            for v in vals:
                cls = heat_cls(v, gmax) if has else 'e'
                t.append(f'<td class="cell {cls}">{v:,}' + '</td>' if v>0 else f'<td class="cell {cls}"></td>')
            rtot = sum(vals)
            t.append(f'<td class="tot-cell">{rtot:,}</td>' if rtot>0 else '<td class="tot-cell">—</td>')
            t.append('</tr>')
        t.append('</tbody></table>')
        return ''.join(t)

    def heatmap_combined(day_data, sn):
        """
        3 metrics (V/UV/PSB) side-by-side per hour, per day.
        day_data = [{label, color, v:[], uv:[], p:[], v_global:int}, ...]
        sn = number of hours to show
        """
        sh = HOUR_LBLS[:sn]

        # Per-day, per-metric max → independent colour scale per day
        for d in day_data:
            d['gm_v']  = max(d['v'][:sn])  if any(x>0 for x in d['v'][:sn])  else 1
            d['gm_uv'] = max(d['uv'][:sn]) if any(x>0 for x in d['uv'][:sn]) else 1
            d['gm_p']  = max(d['p'][:sn])  if any(x>0 for x in d['p'][:sn])  else 1

        t = ['<table class="hm-comb"><thead>']
        # Row 1: Day headers (span 3 cols each) + scale annotation
        t.append('<tr><th class="hc-hour"></th>')
        for d in day_data:
            has = any(x>0 for x in d['v'])
            bg  = d['color'] if has else '#AAAAAA'
            _ncols = 2 + (1 if PROFILE_CONFIG.get(EVENT_PROFILE,PROFILE_CONFIG['full'])['has_psb'] else 0)
            t.append(f'<th colspan="{_ncols}" class="hc-day" style="background:{bg}">{d["label"]}</th>')
        t.append('</tr>')
        # Row 2: Sub-col labels + peak value per metric per day
        t.append('<tr><th class="hc-hour">Hour</th>')
        for d in day_data:
            has = any(x>0 for x in d['v'])
            c   = d['color'] if has else '#AAAAAA'
            _has_psb_hm = (PROFILE_CONFIG.get(EVENT_PROFILE, PROFILE_CONFIG['full'])['has_psb'] and SHOW_PASSERBY)
            _metrics = [('V','gm_v'),('UV','gm_uv')] + ([('PSB','gm_p')] if _has_psb_hm else [])
            for lbl, gm_key in _metrics:
                pk = d[gm_key] if has else 0
                tip = f'peak {pk:,}' if has else 'no data'
                t.append(f'<th class="hc-sub" style="color:{c}" title="{tip}">{lbl}</th>')
        t.append('</tr></thead><tbody>')

        # Data rows per hour — each day uses its OWN max for colour
        for hi, hour in enumerate(sh):
            t.append(f'<tr><td class="hc-hour-lbl">{hour}</td>')
            for d in day_data:
                v  = d['v'][hi]  if hi < len(d['v'])  else 0
                uv = d['uv'][hi] if hi < len(d['uv']) else 0
                p  = d['p'][hi]  if hi < len(d['p'])  else 0
                has = any(x>0 for x in d['v'])
                _metrics_d = [(v,'gm_v'),(uv,'gm_uv')] + ([(p,'gm_p')] if PROFILE_CONFIG.get(EVENT_PROFILE,PROFILE_CONFIG['full'])['has_psb'] else [])
                for val, gm_key in _metrics_d:
                    c = heat_cls(val, d[gm_key]) if has else 'e'
                    t.append(f'<td class="hcc {c}">{val:,}</td>' if val>0 else f'<td class="hcc {c}"></td>')
            t.append('</tr>')

        # Total row
        t.append('<tr class="hc-tot"><td class="hc-hour-lbl"><b>Total</b></td>')
        for d in day_data:
            has = sum(d['v'][:sn]) > 0
            tv  = sum(d['v'][:sn])
            tuv = d.get('v_global', uv_count(d['ent_d']) if 'ent_d' in d else sum(d['uv'][:sn]))
            tp  = sum(d['p'][:sn])
            _tot_vals = [tv, tuv] + ([tp] if PROFILE_CONFIG.get(EVENT_PROFILE,PROFILE_CONFIG['full'])['has_psb'] else [])
            for val in _tot_vals:
                t.append(f'<td class="hcc-tot">{val:,}</td>' if has else '<td class="hcc-tot">—</td>')
        t.append('</tr>')

        # Legend row
        t.append(f'<tr><td colspan="{1+len(day_data)*3}" style="padding:4px 0 0">')
        t.append('<span style="font-size:7.5px;color:var(--mut)">V = Visitors &nbsp;&middot;&nbsp; UV = Unique Visitors &nbsp;&middot;&nbsp; PSB = Passersby (PSB gates + Visitors) &nbsp;&middot;&nbsp; Colour scale independent per metric</span>')
        t.append('</td></tr>')

        t.append('</tbody></table>')
        return ''.join(t)


    def legend(mx):
        return (f'<div class="leg"><span class="ll">Low</span>'
                f'<span class="ls h0"></span><span class="ls h1"></span>'
                f'<span class="ls h2"></span><span class="ls h3"></span>'
                f'<span class="ls h4"></span>'
                f'<span class="ll">High &nbsp;(max {mx:,})</span></div>')

    def no_data_block():
        return '<div class="nodata">&#9203; No data for this day — pending upload</div>'

    def page_break():
        return '<div class="pb"></div>'

    def page_footer(page_num, total_pages):
        return (f'<div class="pfooter">DITECH — Digital Intelligence Technology Co., Ltd. '
                f'· AI Vision People Counting · OTC Asia 2026, Malaysia · Confidential'
                f' · Page {page_num}</div>')

    # ══════════════════════════════════════════════════════════════
    # BUILD SECTIONS
    # ══════════════════════════════════════════════════════════════
    sections = []

    # ─────────────────────────────────────────────────────────────
    # SECTION 0: EXECUTIVE SUMMARY (1-page overview for leadership)
    # ─────────────────────────────────────────────────────────────
    _es = []
    _es.append(sec_hdr('⭐', 'Executive Summary',
                       f'{EVENT_NAME} · {EVENT_LABELS_[0]} – {EVENT_LABELS_[-1]} · '
                       f'Key metrics, insights, and impact analysis'))

    # Master KPIs
    _ex_v   = len(ent)
    _ex_uv  = int(uv_count(ent))
    _ex_psb = len(psb) + _ex_v
    _ex_conv = (_ex_v / _ex_psb * 100) if _ex_psb > 0 else 0

    # Engagement (staff-aware: filter at source so numerator matches uv_count denominator)
    _zio = _non_staff(df)
    _zio = _zio[(_zio['Type']=='Zone') & _zio['Event'].isin(['in','out'])].sort_values(['BodyID','Location','Time'])
    _ex_eng = set()
    for (_b,_l),_g in _zio.groupby(['BodyID','Location'],sort=False):
        _is=_g[_g['Event']=='in']['Time'].tolist()
        _os=_g[_g['Event']=='out']['Time'].tolist(); _oi=0
        for _ti in _is:
            while _oi<len(_os) and _os[_oi]<=_ti: _oi+=1
            if _oi<len(_os):
                if (_os[_oi]-_ti).total_seconds() > ENGAGEMENT_THRESHOLD_SEC: _ex_eng.add(_b)
                _oi+=1
    _ex_eng_pct = (len(_ex_eng) / max(_ex_uv,1) * 100)

    _ex_kpis = [
        (f'{VENUE_TYPE} Visitors', f'{_ex_v:,}', f'{len(EVENT_DATES_)}-day total', '#005B9A'),
        ('Unique Visitors',        f'{_ex_uv:,}', 'Distinct individuals', '#1A7A45'),
        ('Entry Conversion',       f'{_ex_conv:.1f}%', 'Passersby → Visitors', '#B06000'),
        ('Engagement Rate',        f'{_ex_eng_pct:.0f}%', 'UV dwell > 60 sec', '#7B2D8B'),
    ]

    # Best activity + Programme score
    _ex_best_act = None
    _ex_prog    = None
    _p_ex = PROFILE_CONFIG.get(EVENT_PROFILE, PROFILE_CONFIG['full'])
    if _p_ex['has_activity']:
        _best_lift = -999; _best = None; _all = []
        for ds in EVENT_DATES_:
            e_d = ent[ent['Date']==ds]
            for ts, te, name in ACTIVITIES_.get(ds, []):
                sh,sm=int(ts[:2]),int(ts[3:]); eh,em=int(te[:2]),int(te[3:])
                sm_=sh*60+sm; em_=eh*60+em; dur=max(em_-sm_,15)
                vw=in_window(e_d,ds,sm_,em_); bw=in_window(e_d,ds,max(0,sm_-dur),sm_)
                b,d=len(bw),len(vw)
                if b >= 20: base = b
                else:
                    op_s = DISPLAY_HOURS_START*60; op_e = DISPLAY_HOURS_END*60
                    op_ent = e_d[(e_d['Time'].dt.hour*60 + e_d['Time'].dt.minute >= op_s) &
                                 (e_d['Time'].dt.hour*60 + e_d['Time'].dt.minute <= op_e)]
                    tw = max((op_e - op_s) // dur, 1)
                    base = max(len(op_ent) // tw, 1)
                lift = max(-100, min(300, (d - base) / base * 100 if base > 0 else 0))
                _all.append({'name':name, 'during':d, 'lift_pct':lift})
                if lift > _best_lift:
                    _best_lift = lift; _best = {'name':name, 'during':d, 'lift_pct':lift}
        _ex_best_act = _best
        win_n = sum(1 for a in _all if a['lift_pct'] >= 20)
        _ex_prog = {'pct': round(win_n/max(len(_all),1)*100), 'winners': win_n, 'total': len(_all)}

    _ex_top_zone = None
    if _p_ex['has_zones'] and len(zon) > 0:
        _zv = zon.groupby('Location').size().sort_values(ascending=False)
        if len(_zv) > 0:
            _tz = _zv.index[0]
            _zd = ddf[ddf['loc']==_tz] if 'loc' in ddf.columns else pd.DataFrame()
            _dw = float(_zd['min'].mean()) if len(_zd) > 0 else 0.0
            _ex_top_zone = {'name': _tz, 'visits': int(_zv.iloc[0]), 'dwell_min': _dw}

    _ex_top_time = None
    if len(ent) > 0:
        _hc = ent.groupby('Hour').size()
        _ex_top_time = {'hour': int(_hc.idxmax()), 'visitors': int(_hc.max()),
                        'label': f'{(_hc.max()/len(ent)*100):.1f}% of total'}

    # Top headline insights across categories
    _ex_headlines = []
    for cat in ['funnel','activity','traffic','zone','demographic']:
        for ins in _get_ins(cat):
            if ins['level'] == 'headline':
                _ex_headlines.append(ins)
    _ex_headlines = _ex_headlines[:5]

    _es.append(exec_summary_block(
        kpi_cards_list=_ex_kpis,
        headline_insights=_ex_headlines,
        prog_score=_ex_prog,
        best_activity=_ex_best_act,
        top_zone=_ex_top_zone,
        top_time=_ex_top_time,
    ))
    sections.append(('\n'.join(_es), 0))

    # ─────────────────────────────────────────────────────────────
    # SECTION 1: OVERALL SUMMARY KPIs + Hourly Traffic
    # ─────────────────────────────────────────────────────────────
    total_v   = len(ent)
    unique_v  = uv_count(ent)
    total_p   = len(psb) + len(ent)
    total_z   = len(zon)
    peak_h_v  = ent.groupby('Hour').size().idxmax() if total_v > 0 else 0
    peak_v_v  = int(ent.groupby('Hour').size().max()) if total_v > 0 else 0
    male_n    = int((ent['Gender']=='Male').sum())
    female_n  = int((ent['Gender']=='Female').sum())
    tg        = male_n + female_n

    s = []
    s.append(sec_hdr('\U0001f4ca', 'Overall Booth Traffic Summary',
                     f'{EVENT_LABELS[0]} \u00b7 Day 1 of {len(EVENT_DATES)}'))
    # Compute zone dwell avg for KPI
    _zone_io = df[(df['Type']=='Zone') & df['Event'].isin(['in','out'])].sort_values(['BodyID','Location','Time'])
    _dwell_all = []
    for (_bid,_loc),_grp in _zone_io.groupby(['BodyID','Location'],sort=False):
        _ins=_grp[_grp['Event']=='in']['Time'].tolist(); _outs=_grp[_grp['Event']=='out']['Time'].tolist(); _oi=0
        for _t in _ins:
            while _oi<len(_outs) and _outs[_oi]<=_t: _oi+=1
            if _oi<len(_outs):
                _sec=(_outs[_oi]-_t).total_seconds()
                if DWELL_MIN_SEC<=_sec<=DWELL_MAX_SEC: _dwell_all.append(_sec/60)  # 10 sec – 60 min
                _oi+=1
    _davg = round(sum(_dwell_all)/len(_dwell_all),1) if _dwell_all else 0
    _dmed = round(sorted(_dwell_all)[len(_dwell_all)//2],1) if _dwell_all else 0

    # Per-day stats for sub-labels (show all days)
    def _day_v(ds):
        e=ent[ent['Date']==ds]; return len(e) if len(e)>0 else None
    def _day_uv(ds):
        e=ent[ent['Date']==ds]; return uv_count(e) if len(e)>0 else None
    def _day_p(ds):
        e=ent[ent['Date']==ds]; pb=psb[psb['Date']==ds]
        return len(pb)+len(e) if len(e)>0 else None
    def _day_peak(ds):
        e=ent[ent['Date']==ds]
        if len(e)==0: return None,None
        ph=e.groupby('Hour').size().idxmax(); pv=int(e.groupby('Hour').size().max())
        return ph,pv
    _day_labels = []
    for _di,_ds in enumerate(EVENT_DATES_):
        _dv=_day_v(_ds); _ph,_pv=_day_peak(_ds)
        if _dv: _day_labels.append(f'D{_di+1}: {_dv:,} V · Peak {_ph:02d}:00')
        else:   _day_labels.append(f'D{_di+1}: No data')
    _visitors_sub = ' / '.join(_day_labels)

    s.append(kpi_cards(
        (f'{VENUE_TYPE} Visitors', fmt(total_v),
            ' · '.join([(f'D{i+1}: {len(ent[ent["Date"]==ds]):,}' if len(ent[ent['Date']==ds])>0 else f'D{i+1}: —')
                         for i,ds in enumerate(EVENT_DATES_)]), '#005B9A'),
        ('Unique Visitors',   fmt(unique_v), 'Distinct individuals (global)',               '#1A7A45'),
        *([('Passersby', fmt(total_p), 'PSB gates + Visitors', '#B06000')]
          if (PROFILE_CONFIG.get(EVENT_PROFILE, PROFILE_CONFIG['full'])['has_psb'] and SHOW_PASSERBY) else []),
        *([('Zone Visits', fmt(total_z), 'All zones combined', '#003865')]
          if PROFILE_CONFIG.get(EVENT_PROFILE, PROFILE_CONFIG['full'])['has_zones'] else []),
        ('Peak Hour',         f'{peak_h_v:02d}:00', f'{peak_v_v:,} visitors · Day 1',      '#8B0000'),
        *([('Zone Dwell (Avg)', f'{_davg} min', f'Median: {_dmed} min · {len(_dwell_all):,} sessions', '#2D7DD2')]
          if PROFILE_CONFIG.get(EVENT_PROFILE, PROFILE_CONFIG['full'])['has_dwell'] else []),
    ))

    # Traffic + demographic insights
    _ins1 = _get_ins('traffic') + _get_ins('demographic')
    if _ins1: s.append(insight_block(_ins1, max_show=3))

    sections.append(('\n'.join(s), 1))

    # ─────────────────────────────────────────────────────────────
    # SECTION 2: ZONE TRAFFIC SUMMARY (per day) + HEATMAP
    # ─────────────────────────────────────────────────────────────
    s = []

    # ── Visitor Funnel (standard/full only) ──────────────────
    _p2 = PROFILE_CONFIG.get(EVENT_PROFILE, PROFILE_CONFIG['full'])
    if _p2['has_psb'] and SHOW_PASSERBY:
        # Staff-aware: filter so engaged_uv numerator matches uv_count denominator
        _zon_io = _non_staff(df)
        _zon_io = _zon_io[(_zon_io['Type']=='Zone') & _zon_io['Event'].isin(['in','out'])].sort_values(['BodyID','Location','Time'])
        _eng = set()
        for (_bid,_loc),_grp in _zon_io.groupby(['BodyID','Location'],sort=False):
            _ins_t=_grp[_grp['Event']=='in']['Time'].tolist()
            _outs_t=_grp[_grp['Event']=='out']['Time'].tolist(); _oi=0
            for _ti in _ins_t:
                while _oi<len(_outs_t) and _outs_t[_oi]<=_ti: _oi+=1
                if _oi<len(_outs_t):
                    if (_outs_t[_oi]-_ti).total_seconds()>ENGAGEMENT_THRESHOLD_SEC: _eng.add(_bid)
                    _oi+=1
        s.append(sec_hdr('🔻', 'Visitor Funnel',
                         'Passersby → Visitors → Unique → Engaged'))
        s.append(funnel_block(
            total_psb=len(psb)+len(ent),
            total_v=len(ent), total_uv=uv_count(ent),
            engaged_uv=len(_eng),
            insights_list=_get_ins('funnel'),
        ))

    s.append(sec_hdr('🗺️', 'Zone Traffic Summary — All Zones',
                     'Total Visits / Unique / Engmt % / Dwell Time · แยกตามวัน'))

    # ── Pre-compute engaged UV (dwell > 60s) per zone PER DAY ────
    # Staff-aware: filter so engagement % aligns with zone unique
    zone_io_eng = _non_staff(df)
    zone_io_eng = zone_io_eng[(zone_io_eng['Type']=='Zone') & zone_io_eng['Event'].isin(['in','out'])].sort_values(['BodyID','Location','Time'])
    engaged_day = {}   # (loc, date) → set of BodyID
    for (bid, loc), grp in zone_io_eng.groupby(['BodyID','Location'], sort=False):
        ins  = grp[grp['Event']=='in']['Time'].tolist()
        outs = grp[grp['Event']=='out']['Time'].tolist()
        oi   = 0
        for t_in in ins:
            while oi < len(outs) and outs[oi] <= t_in: oi += 1
            if oi < len(outs):
                if (outs[oi] - t_in).total_seconds() > ENGAGEMENT_THRESHOLD_SEC:
                    ds_key = t_in.strftime('%Y-%m-%d')
                    engaged_day.setdefault((loc, ds_key), set()).add(bid)
                oi += 1
    # Also global (all days)
    engaged_all = {}
    for (loc, ds_key), bids in engaged_day.items():
        engaged_all.setdefault(loc, set()).update(bids)

    # ── Per-day zone tables ───────────────────────────────────────
    for di, ds in enumerate(EVENT_DATES_):
        z_d = zon[zon['Date']==ds]
        has = len(z_d) > 0

        s.append(day_hdr(f'Day {di+1}  ·  {EVENT_LABELS_[di]}', DAY_CSS[di]))

        if not has:
            s.append(no_data_block())
            continue

        day_stats = (z_d.groupby('Location')
                     .agg(V=('No','count'),
                          UV=('BodyID', lambda s: (_non_staff(z_d.loc[s.index])['BodyID'].nunique()
                                                   if 'CustomerType' in z_d.columns else s.nunique())))
                     .reset_index().sort_values('V', ascending=False))
        tv_day = int(day_stats['V'].sum())

        s.append('<table class="dt"><thead><tr>'
                 '<th>Zone</th><th>Visits</th><th>Unique</th><th>Engmt %</th>'
                 '<th>Male</th><th>Female</th><th>Avg Dwell</th><th>Half stayed ></th><th>Ratio %</th>'
                 '</tr></thead><tbody>')

        for i, (_, zr) in enumerate(day_stats.iterrows()):
            loc   = zr['Location']
            zm    = int((z_d[z_d['Location']==loc]['Gender']=='Male').sum())
            zf    = int((z_d[z_d['Location']==loc]['Gender']=='Female').sum())
            uv_t  = int(zr['UV'])
            uv_e  = len(engaged_day.get((loc, ds), set()))
            eng_n = uv_e/uv_t*100 if uv_t > 0 else 0
            eng   = f'{eng_n:.1f}%' if uv_t > 0 else '—'
            dw    = ddf[(ddf['loc']==loc) & (ddf['date']==ds)] if len(ddf)>0 else None
            avg_d = f"{dw['min'].mean():.1f}" if dw is not None and len(dw)>0 else '—'
            med_d = f"{dw['min'].median():.1f}" if dw is not None and len(dw)>0 else '—'
            ratio = f"{int(zr['V'])/tv_day*100:.1f}%"
            eng_s = ('color:#1A7A45;font-weight:700' if eng_n>=25 else
                     'color:#B06000;font-weight:700' if eng_n>=15 else
                     'color:#B02020;font-weight:700')
            alt   = 'alt' if i%2==0 else ''
            s.append(f'<tr class="{alt}"><td>{loc}</td>'
                     f'<td class="mono">{int(zr["V"]):,}</td>'
                     f'<td class="mono">{uv_t:,}</td>'
                     f'<td class="mono" style="{eng_s}">{eng}</td>'
                     f'<td class="mono">{zm:,}</td>'
                     f'<td class="mono">{zf:,}</td>'
                     f'<td class="mono">{avg_d}</td>'
                     f'<td class="mono">{med_d}</td>'
                     f'<td class="mono">{ratio}</td></tr>')

        # Totals for every column
        tot_zm  = int((z_d['Gender']=='Male').sum())
        tot_zf  = int((z_d['Gender']=='Female').sum())
        tot_uv  = int(uv_count(z_d))
        # Engmt% total = UV engaged (any zone) / total UV
        all_eng_day = set(bid for (loc_k, ds_k), bids in engaged_day.items()
                          if ds_k == ds for bid in bids)
        tot_eng_n   = len(all_eng_day) / tot_uv * 100 if tot_uv > 0 else 0
        tot_eng     = f'{tot_eng_n:.1f}%' if tot_uv > 0 else '—'
        tot_eng_s   = ('color:#1A7A45;font-weight:700' if tot_eng_n>=25 else
                       'color:#B06000;font-weight:700' if tot_eng_n>=15 else
                       'color:#B02020;font-weight:700')
        # Overall dwell (all zones, all days — ddf already filtered 10s-3hr)
        ddf_day = ddf[ddf['date']==ds] if len(ddf)>0 else ddf
        tot_avg = f"{ddf_day['min'].mean():.1f}" if len(ddf_day)>0 else '—'
        tot_med = f"{ddf_day['min'].median():.1f}" if len(ddf_day)>0 else '—'
        s.append(f'<tr class="total">'
                 f'<td><b>TOTAL — Day {di+1}</b></td>'
                 f'<td class="mono">{tv_day:,}</td>'
                 f'<td class="mono">{tot_uv:,}</td>'
                 f'<td class="mono" style="{tot_eng_s}">{tot_eng}</td>'
                 f'<td class="mono">{tot_zm:,}</td>'
                 f'<td class="mono">{tot_zf:,}</td>'
                 f'<td class="mono">{tot_avg}</td>'
                 f'<td class="mono">{tot_med}</td>'
                 f'<td class="mono">100%</td></tr>')
        s.append('</tbody></table>')

    s.append('<p class="tbl-note">&#9432;&nbsp; <b>Engmt %</b> = Unique visitors who stayed &gt;60 sec ÷ Zone unique visitors that day &nbsp;&middot;&nbsp; <b>Avg / Half stayed &gt;</b> = all days combined (dwell sessions)</p>')

    # ── Zone Heatmap per day ──────────────────────────────────────
    s.append(sec_hdr('🔥', 'Hourly Zone Traffic Heatmap', 'แยกตามวัน · intensity relative to each day\'s peak'))
    for di, ds in enumerate(EVENT_DATES_):
        z_d = zon[zon['Date']==ds]
        has = len(z_d) > 0
        # Insert page-break before days 2+ if many days (prevents overflow)
        if di > 0 and len(EVENT_DATES_) > 2:
            pass  # page-break disabled
        s.append(day_hdr(f'Day {di+1}  ·  {EVENT_LABELS_[di]}', DAY_CSS[di]))
        if not has:
            s.append(no_data_block()); continue

        rows_z = [{'label': ZONE_ABBREV_.get(z, z), 'color': '#7A8FA6',
                   'vals': [int(len(z_d[(z_d['Location']==z)&(z_d['Hour']==h)])) for h in HOURS]}
                  for z in ZONES]
        av = [v for r in rows_z for v in r['vals']]
        gm = max(av) if av else 1
        sn = trim_n([r['vals'] for r in rows_z])
        s.append(heatmap_tbl(rows_z, gm, sn))
        s.append(legend(gm))

    # ── Top Zone Ranking block ────────────────────────────
    zone_stats_all = []
    for z in ZONE_ORDER_:
        zv = zon[zon['Location']==z]
        if len(zv) == 0: continue
        vcount = len(zv)
        uv = int(uv_count(zv))
        # Dwell avg for this zone
        zd = ddf[ddf['loc']==z] if 'loc' in ddf.columns else pd.DataFrame()
        d_avg = float(zd['min'].mean()) if len(zd) > 0 else 0.0
        # Engagement %
        total_zone_uv = int(uv_count(zv))
        eng_uv = int(zd[zd['min']*60 > ENGAGEMENT_THRESHOLD_SEC]['bid'].nunique()) if len(zd) > 0 else 0
        eng_pct = (eng_uv / total_zone_uv * 100) if total_zone_uv > 0 else 0
        zone_stats_all.append({
            'name': z, 'visits': vcount, 'unique': uv,
            'dwell_min': d_avg, 'engagement_pct': eng_pct,
        })

    if zone_stats_all:
        s.append(sec_hdr('🏆', 'Top Zones Ranking', 'Highest traffic + engagement · ranked by visits'))
        s.append(top_zone_ranking(zone_stats_all, max_show=5))

    # ── Sponsor Value blocks (one per sponsored zone) ──────
    _sponsored = [z for z in SPONSOR_ZONES if z in zon['Location'].unique()]
    if _sponsored:
        pass  # page-break disabled
        s.append(sec_hdr('💎', f'Sponsor Value Analysis',
                         f'{len(_sponsored)} sponsored zone(s) · Exposure + Engagement metrics'))
        for si, zone_name in enumerate(_sponsored):
            stage_zon = zon[zon['Location']==zone_name]
            if len(stage_zon) == 0:
                continue
            exp_v  = len(stage_zon)
            exp_uv = int(uv_count(stage_zon))
            during_v = 0
            for date_str, acts in ACTIVITIES_.items():
                for ts, te, _ in acts:
                    sm = int(ts[:2])*60+int(ts[3:])
                    em = int(te[:2])*60+int(te[3:])
                    mod = stage_zon['Time'].dt.hour*60 + stage_zon['Time'].dt.minute
                    during_v += len(stage_zon[(stage_zon['Date']==date_str) & (mod>=sm) & (mod<=em)])
            st_d = ddf[ddf['loc']==zone_name] if 'loc' in ddf.columns else pd.DataFrame()
            avg_dw = float(st_d['min'].mean()) if len(st_d) > 0 else 0.0
            eng_uv = int(st_d[st_d['min']*60 > ENGAGEMENT_THRESHOLD_SEC]['bid'].nunique()) if len(st_d) > 0 else 0
            eng_pct = (eng_uv / exp_uv * 100) if exp_uv > 0 else 0

            # Sub-header per sponsored zone
            s.append(f'<div class="sponsor-zone-hdr" '
                     f'style="background:linear-gradient(90deg,#7B2D8B,#B06000);color:#fff;'
                     f'padding:6px 10px;border-radius:4px;margin:8px 0 6px 0;'
                     f'font-size:10px;font-weight:700">'
                     f'💎 {zone_name}</div>')
            s.append(sponsor_value_block(
                stage_zone=zone_name,
                exposure_v=exp_v, exposure_uv=exp_uv,
                during_activity_v=during_v,
                avg_dwell_min=avg_dw, engaged_pct=eng_pct,
                insights_list=None,
            ))
            # Break between multiple sponsors (if >1)
            if si < len(_sponsored)-1 and len(_sponsored) > 1:
                s.append('<div style="height:6px"></div>')

    sections.append(('\n'.join(s), 2))
    sections.append(('\n'.join(s), 2))

    # ─────────────────────────────────────────────────────────────
    # SECTION 3: ACTIVITY ANALYTICS
    # ─────────────────────────────────────────────────────────────
    s = []
    _act_ins = _get_ins('activity')
    s.append(sec_hdr('🎪', f'{VENUE_TYPE} Activity Analytics', 'Traffic correlation · Before / During / After'))

    # ── Build activity rows with before/during/after windows ──
    all_acts = []
    act_corr_rows = []
    for ds in EVENT_DATES_:
        e_d = ent[ent['Date']==ds]; p_d = psb[psb['Date']==ds]; z_d = zon[zon['Date']==ds]
        di = EVENT_DATES_.index(ds)
        for ts, te, name in ACTIVITIES_.get(ds, []):
            sh,sm=int(ts[:2]),int(ts[3:]); eh,em=int(te[:2]),int(te[3:])
            sm_=sh*60+sm; em_=eh*60+em
            dur_min = max(em_ - sm_, 15)
            vw=in_window(e_d,ds,sm_,em_); pw=in_window(p_d,ds,sm_,em_); zw=in_window(z_d,ds,sm_,em_)
            all_acts.append({'ds':ds,'di':di,'time':f'{ts}–{te}','name':name,
                             'v':len(vw),'uv':int(uv_count(vw)),'p':len(pw),'z':len(zw)})
            # Before/after windows of equal duration
            bw = in_window(e_d, ds, max(0, sm_-dur_min), sm_)
            aw = in_window(e_d, ds, em_, em_+dur_min)
            b, d, a = len(bw), len(vw), len(aw)
            # Calibrate lift: if before-window has too few visitors (pre-opening),
            # use after-window as baseline instead. If both too low, use day median
            MIN_BASELINE = 20
            if b >= MIN_BASELINE:
                baseline = b
            elif a >= MIN_BASELINE:
                baseline = a
            else:
                # Day median visitors per equivalent-duration window during operating hours
                op_start = max(DISPLAY_HOURS_START*60, 9*60)
                op_end   = min(DISPLAY_HOURS_END*60, 22*60)
                op_ent   = e_d[(e_d['Time'].dt.hour*60 + e_d['Time'].dt.minute >= op_start) &
                               (e_d['Time'].dt.hour*60 + e_d['Time'].dt.minute <= op_end)]
                total_op = len(op_ent)
                total_windows = max((op_end - op_start) // dur_min, 1)
                baseline = max(total_op // total_windows, 1)
            lift_pct = (d - baseline) / baseline * 100 if baseline > 0 else 0
            # Cap lift at ±300% for sanity
            lift_pct = max(-100, min(300, lift_pct))
            act_corr_rows.append({
                'date': EVENT_LABELS_[di], 'time': f'{ts}–{te}', 'name': name,
                'before': b, 'during': d, 'after': a, 'lift_pct': lift_pct,
                'baseline': baseline,
            })

    # ── Render Activity Correlation visual block ──
    if act_corr_rows:
        s.append(activity_correlation_block(act_corr_rows, insights_list=_act_ins))

    # ── Event Impact Insights (Sprint 3) ──
    try:
        from insight_engine import event_impact_insights
        _impact_ins = event_impact_insights(act_corr_rows)
        if _impact_ins:
            # Break before impact analysis if many activities
            if len(act_corr_rows) > 6:
                pass  # page-break disabled
            s.append(sec_hdr('📊', 'Event Impact Analysis', 'Programme effectiveness summary'))
            s.append(insight_block(_impact_ins, max_show=3))
    except Exception as _e:
        print(f'  ⚠ impact insight err: {_e}')

    # ── Before/During/After Numeric Summary Table ──
    if act_corr_rows:
        s.append(sec_hdr('📐', 'Before / During / After — Numeric Summary',
                         'Exact visitor counts per activity window'))
        s.append('<table class="dt bda-tbl"><thead><tr>'
                 '<th>Activity</th><th>Time</th>'
                 '<th class="bda-before">Before</th>'
                 '<th class="bda-during">During</th>'
                 '<th class="bda-after">After</th>'
                 '<th>Incremental</th><th>Lift %</th>'
                 '</tr></thead><tbody>')
        for i, r in enumerate(act_corr_rows):
            incr = r['during'] - r['before']
            incr_sign = f'+{incr:,}' if incr >= 0 else f'{incr:,}'
            incr_color = '#1A7A45' if incr > 0 else ('#B02020' if incr < 0 else '#6B8299')
            lift = r['lift_pct']
            lift_sign = f'+{lift:.0f}%' if lift >= 0 else f'{lift:.0f}%'
            lift_color = '#1A7A45' if lift >= 20 else ('#B06000' if lift >= 0 else '#B02020')
            cls = 'alt' if i % 2 == 0 else ''
            s.append(
                f'<tr class="{cls}">'
                f'<td class="act-name">{r["name"]}</td>'
                f'<td class="mono">{r["date"]} · {r["time"]}</td>'
                f'<td class="mono">{r["before"]:,}</td>'
                f'<td class="mono" style="font-weight:700">{r["during"]:,}</td>'
                f'<td class="mono">{r["after"]:,}</td>'
                f'<td class="mono" style="color:{incr_color};font-weight:700">{incr_sign}</td>'
                f'<td class="mono" style="color:{lift_color};font-weight:700">{lift_sign}</td>'
                f'</tr>'
            )
        # Totals
        tot_b = sum(r['before'] for r in act_corr_rows)
        tot_d = sum(r['during'] for r in act_corr_rows)
        tot_a = sum(r['after']  for r in act_corr_rows)
        tot_incr = tot_d - tot_b
        tot_lift = (tot_incr / max(tot_b, 1)) * 100
        s.append(
            f'<tr class="total"><td>TOTAL</td><td></td>'
            f'<td class="mono">{tot_b:,}</td>'
            f'<td class="mono">{tot_d:,}</td>'
            f'<td class="mono">{tot_a:,}</td>'
            f'<td class="mono">{tot_incr:+,}</td>'
            f'<td class="mono">{tot_lift:+.0f}%</td></tr>'
        )
        s.append('</tbody></table>')

    # ── Original table for detail ──
    s.append(sec_hdr('📋', 'Activity Traffic Detail Table', 'Per-activity visitors, unique, passersby, zone visits'))

    prev_ds = None
    s.append('<table class="dt"><thead><tr>'
             '<th>Time</th><th>Activity</th><th>Visitors</th><th>Unique</th><th>Passersby</th><th>Zone Visits</th>'
             '</tr></thead><tbody>')
    for i,row in enumerate(all_acts):
        if row['ds'] != prev_ds:
            di = row['di']
            s.append(f'<tr class="day-sep"><td colspan="6" style="background:{DAY_CSS[di]};'
                     f'color:#fff;font-weight:700;padding:4px 8px">'
                     f'Day {di+1}  ·  {EVENT_LABELS_[di]}</td></tr>')
            prev_ds = row['ds']
        s.append(f'<tr class="hi"><td class="mono">{row["time"]}</td>'
                 f'<td class="act-name">{row["name"]}</td>'
                 f'<td class="mono">{row["v"]:,}</td><td class="mono">{row["uv"]:,}</td>'
                 f'<td class="mono">{row["p"]:,}</td><td class="mono">{row["z"]:,}</td></tr>')
    s.append('</tbody></table>')

    # Activity × Zone matrix — always break before (large)
    pass  # page-break disabled
    s.append(sec_hdr('🗺️', 'Activity × Zone Traffic Matrix', 'Heatmap · visits per zone during each activity'))
    short_z = [ZONE_ABBREV_.get(z,z) for z in ZONES]
    s.append('<div style="overflow-x:auto"><table class="hm" style="font-size:8px;table-layout:auto"><thead><tr>'
             '<th class="zth" style="width:150px;white-space:normal;word-break:break-word">Activity</th>')
    s += [f'<th style="font-size:7.5px">{z}</th>' for z in short_z]
    s.append('<th class="tot-th">V</th></tr></thead><tbody>')

    max_mval = 1
    matrix = []
    for row in all_acts:
        ds=row['ds']; sh,sm=int(row['time'][:2]),int(row['time'][3:5])
        eh,em=int(row['time'][6:8]),int(row['time'][9:11])
        sm_=sh*60+sm; em_=eh*60+em
        z_d=zon[zon['Date']==ds]
        zcounts=[]
        for z in ZONES:
            sub=z_d[z_d['Location']==z]; mins=sub['Time'].dt.hour*60+sub['Time'].dt.minute
            n=int(len(sub[(mins>=sm_)&(mins<=em_)])); zcounts.append(n); max_mval=max(max_mval,n)
        matrix.append(zcounts)

    prev_ds = None
    for i,(row,zcounts) in enumerate(zip(all_acts,matrix)):
        if row['ds'] != prev_ds:
            di=row['di']
            s.append(f'<tr><td colspan="{len(ZONES)+2}" style="background:{DAY_CSS[di]};'
                     f'color:#fff;font-size:8.5px;font-weight:700;padding:3px 6px">'
                     f'Day {di+1}  ·  {EVENT_LABELS_[di]}</td></tr>')
            prev_ds = row['ds']
        clr = DAY_CSS[row['di']]
        s.append(f'<tr><td class="zl" style="color:{clr};font-weight:600;font-size:8px;white-space:normal;word-break:break-word;max-width:150px">'
                 f'{row["time"]}<br>{row["name"]}</td>')
        for n in zcounts:
            c = heat_cls(n, max_mval)
            s.append(f'<td class="cell {c}" style="height:22px;font-size:8px">{n if n>0 else ""}</td>')
        s.append(f'<td class="tot-cell" style="font-size:8px">{row["v"]:,}</td></tr>')
    s.append('</tbody></table></div>')

    sections.append(('\n'.join(s), 3))

    # ─────────────────────────────────────────────────────────────
    # SECTION 4: DWELL TIME ANALYSIS
    # ─────────────────────────────────────────────────────────────
    s = []
    _zone_ins = _get_ins('zone')
    if _zone_ins: s.append(insight_block(_zone_ins, max_show=2))
    s.append(sec_hdr('⏱️', 'Dwell Time Analysis', 'Time spent per zone · in→out pairing · 10 sec – 60 min filter'))

    if len(ddf) > 0:
        # Booth dwell
        booth_io = df[df['Type'].isin(['Entrance','Zone']) & df['Event'].isin(['in','out'])]
        bspans = booth_io.groupby('BodyID')['Time'].agg(['min','max']).reset_index()
        bspans['dmin'] = (bspans['max'] - bspans['min']).dt.total_seconds()/60
        bspans = bspans[bspans['dmin']>0]
        zone_avg = round(ddf['min'].mean(), 1)
        zone_med = round(ddf['min'].median(), 1)
        s.append(kpi_cards(
            ('Avg Zone Dwell Time',     f'{zone_avg} min', f'{len(ddf):,} valid zone sessions (in→out pairs)', '#003865'),
            ('Median Zone Dwell',       f'{zone_med} min', 'Half of visitors stayed longer than this',           '#005B9A'),
            ('Visitors with Dwell Data',f'{ddf["bid"].nunique():,}', f'of {int(uv_count(zon)):,} zone unique visitors',  '#1A7A45'),
            ('Zones with Dwell Data',   f'{ddf["loc"].nunique()}',   'Zones with valid in→out pairs',                         '#B06000'),
        ))

        # Dwell distribution
        s.append('<p class="tbl-note" style="margin-bottom:6px">'
                 '<b>Avg</b> = Average dwell time &nbsp;&middot;&nbsp; '
                 '<b>Half stayed &gt;</b> = Median (50% stayed longer than this) &nbsp;&middot;&nbsp; '
                 '<b>Top 25% stayed &gt;</b> = 25% of visitors stayed longer &nbsp;&middot;&nbsp; '
                 '<b>Top 10% stayed &gt;</b> = only 10% stayed this long</p>')
        s.append('<table class="dt"><thead><tr>'
                 '<th>Zone</th><th>Unique Visitors with Dwell</th>'
                 '<th>Avg (min)</th>'
                 '<th title="Median: half stayed longer than this">Half stayed &gt; (min)</th>'
                 '<th title="75th percentile: 25% stayed longer than this">Top 25% &gt; (min)</th>'
                 '<th title="90th percentile: only 10% stayed longer than this">Top 10% &gt; (min)</th>'
                 '</tr></thead><tbody>')
        for i,z in enumerate(ZONES):
            g = ddf[ddf['loc']==z]
            if len(g)==0: continue
            avg_d=g['min'].mean(); med_d=g['min'].median()
            p75=g['min'].quantile(.75); p90=g['min'].quantile(.90)
            col = '#1A7A45' if avg_d<5 else '#B06000' if avg_d<15 else '#B02020'
            s.append(f'<tr class="{"alt" if i%2==0 else ""}"><td>{z}</td>'
                     f'<td class="mono">{g["bid"].nunique():,}</td>'
                     f'<td class="mono" style="color:{col};font-weight:700">{avg_d:.1f}</td>'
                     f'<td class="mono">{med_d:.1f}</td>'
                     f'<td class="mono">{p75:.1f}</td><td class="mono">{p90:.1f}</td></tr>')
        s.append('</tbody></table>')
        s.append('<p class="tbl-note">&#9432;&nbsp; '
                 '<b>Unique Visitors</b> here = visitors with a valid in&#8594;out pair (10 sec – 60 min). '
                 'This is lower than Zone Summary UV because some visitors had no recorded exit '
                 '(still inside at data export, or sensor did not detect exit).</p>')
    else:
        s.append(no_data_block())

    # ── Dwell Time Benchmark by Zone (behavior-based) ─────────────
    # Shown only when SHOW_DWELL_BENCHMARK is on. Compares actual avg
    # zone dwell against per-zone targets set in event config.
    if SHOW_DWELL_BENCHMARK and len(ddf) > 0 and ZONE_BENCHMARK:
        s.append(sec_hdr('\U0001F3AF', 'Dwell Time Benchmark by Zone',
                         'Behavior-based targets per zone \u00b7 actual vs benchmark'))
        s.append('<p class="tbl-note" style="margin-bottom:6px">'
                 'Benchmarks reflect expected visitor behavior in each zone '
                 '(interaction complexity), not a single flat target.</p>')
        s.append('<table class="dt"><thead><tr>'
                 '<th>Zone</th><th>Description</th>'
                 '<th>Actual Avg Dwell</th><th>Benchmark</th><th>Status</th>'
                 '</tr></thead><tbody>')
        _row_i = 0
        for z in ZONES:
            _bench_sec = ZONE_BENCHMARK.get(z)
            if _bench_sec is None:
                continue   # zone has no benchmark set — skip
            _g = ddf[ddf['loc'] == z]
            _actual_min = _g['min'].mean() if len(_g) > 0 else None
            _bench_min = _bench_sec / 60.0
            _desc = ZONE_DESC.get(z, '')
            _mode = ZONE_BENCHMARK_MODE.get(z, 'higher_better')
            _dir = '\u2264' if _mode == 'lower_better' else '\u2265'  # ≤ or ≥
            if _actual_min is None:
                _actual_str = '\u2014'
                _status = '<span style="color:#9CA3AF">no data</span>'
            else:
                _actual_str = f'{_actual_min:.1f} min'
                if _mode == 'lower_better':
                    _met = _actual_min <= _bench_min
                    _fail_label = '\u2717 Over'   # stayed too long
                else:
                    _met = _actual_min >= _bench_min
                    _fail_label = '\u2717 Below'  # stayed too short
                if _met:
                    _status = '<span style="color:#1A7A45;font-weight:700">\u2713 Met</span>'
                else:
                    _status = f'<span style="color:#B02020;font-weight:700">{_fail_label}</span>'
            _alt = 'alt' if _row_i % 2 == 0 else ''
            _row_i += 1
            s.append(f'<tr class="{_alt}"><td>{z}</td>'
                     f'<td>{_desc}</td>'
                     f'<td class="mono">{_actual_str}</td>'
                     f'<td class="mono">{_dir} {_bench_min:.0f} min</td>'
                     f'<td>{_status}</td></tr>')
        s.append('</tbody></table>')
        s.append('<p class="tbl-note">&#9432;&nbsp; '
                 '<b>Status</b> = \u2713 Met when a zone hits its target. '
                 '\u2265 means "higher is better" (longer dwell = good, e.g. consultation); '
                 '\u2264 means "lower is better" (shorter dwell = good, e.g. photo booth). '
                 'Zones without a configured benchmark are omitted.</p>')

    sections.append(('\n'.join(s), 4))

    # ─────────────────────────────────────────────────────────────
    # SECTION 5: VISITOR DEMOGRAPHICS
    # ─────────────────────────────────────────────────────────────
    s = []
    s.append(sec_hdr('👥', 'Visitor Demographics', f'Gender & Age Group · {VENUE_TYPE} Entrance only'))

    AGE_ORDER = ['Young Adults (19-35 yrs)','Middle-Aged (36-55 yrs)','Juvenile (0-18 yrs)','Seniors (55+ yrs)']
    AGE_LBL   = ['Young Adults (19-35)','Middle-Aged (36-55)','Juvenile (0-18)','Seniors (55+)']

    def demo_block(uv_df, label=None, color=None):
        """Render gender + age tables for a subset of unique visitors."""
        _m = int((uv_df['Gender']=='Male').sum())
        _f = int((uv_df['Gender']=='Female').sum())
        _t = _m + _f
        _age = uv_df['AgeGroup'].value_counts()
        _ta  = len(uv_df)
        rows = []
        if label:
            bg = color or '#005B9A'
            rows.append(f'<div class="dhdr" style="background:{bg};color:#fff;'
                        f'padding:5px 10px;font-size:9px;font-weight:700;margin-bottom:4px">'
                        f'{label}</div>')
        rows.append('<div class="two-col">')
        # Gender
        rows.append('<div><table class="dt"><thead><tr><th>Gender</th><th>Count</th><th>Share %</th></tr></thead><tbody>')
        for i,(lbl,n) in enumerate([('Male',_m),('Female',_f)]):
            rows.append(f'<tr class="{"alt" if i%2==0 else ""}"><td>{lbl}</td>'
                        f'<td class="mono">{n:,}</td><td class="mono">{pct(n,_t)}</td></tr>')
        rows.append(f'<tr class="total"><td>TOTAL</td><td class="mono">{_t:,}</td><td class="mono">100%</td></tr>')
        rows.append('</tbody></table></div>')
        # Age
        rows.append('<div><table class="dt"><thead><tr><th>Age Group</th><th>Count</th><th>Share %</th></tr></thead><tbody>')
        for i,(full,lbl) in enumerate(zip(AGE_ORDER,AGE_LBL)):
            n = int(_age.get(full,0))
            rows.append(f'<tr class="{"alt" if i%2==0 else ""}"><td>{lbl}</td>'
                        f'<td class="mono">{n:,}</td><td class="mono">{pct(n,_ta)}</td></tr>')
        rows.append(f'<tr class="total"><td>TOTAL</td><td class="mono">{_ta:,}</td><td class="mono">100%</td></tr>')
        rows.append('</tbody></table></div>')
        rows.append('</div>')  # two-col
        return rows

    _demo_prof = PROFILE_CONFIG.get(EVENT_PROFILE, PROFILE_CONFIG['full'])

    if _demo_prof['has_zones']:  # standard/full: overall only
        uniq_v = _non_staff(ent).drop_duplicates('BodyID')
        male_e = int((uniq_v['Gender']=='Male').sum())
        fem_e  = int((uniq_v['Gender']=='Female').sum())
        tge    = male_e + fem_e
        s.append(kpi_cards(
            ('Male Unique Visitors',   f'{pct(male_e,tge)}', f'{male_e:,} unique visitors', '#005B9A'),
            ('Female Unique Visitors', f'{pct(fem_e,tge)}',  f'{fem_e:,} unique visitors', '#B06000'),
            ('Total Unique',           fmt(tge), 'Distinct individuals only', '#003865'),
        ))
        s.extend(demo_block(uniq_v))

    else:  # simple profile: overall KPIs + per-day side-by-side
        uniq_v = _non_staff(ent).drop_duplicates('BodyID')
        male_e = int((uniq_v['Gender']=='Male').sum())
        fem_e  = int((uniq_v['Gender']=='Female').sum())
        tge    = male_e + fem_e
        s.append(kpi_cards(
            ('Male Unique Visitors',   f'{pct(male_e,tge)}', f'{male_e:,} unique visitors', '#005B9A'),
            ('Female Unique Visitors', f'{pct(fem_e,tge)}',  f'{fem_e:,} unique visitors', '#B06000'),
            ('Total Unique',           fmt(tge), 'Distinct individuals only', '#003865'),
        ))
        # Per-day side-by-side (compact) — all days in a multi-col grid
        # Build each day's table as a column in a flex row
        AGE_ORDER2 = ['Young Adults (19-35 yrs)','Middle-Aged (36-55 yrs)',
                      'Juvenile (0-18 yrs)','Seniors (55+ yrs)']
        AGE_LBL2   = ['19–35','36–55','0–18','55+']

        def compact_day_col(uv_df, di, dl):
            """Render one day as a compact single column with gender+age stacked."""
            _m = int((uv_df['Gender']=='Male').sum())
            _f = int((uv_df['Gender']=='Female').sum())
            _t = _m + _f
            _age = uv_df['AgeGroup'].value_counts()
            _ta  = len(uv_df)
            col_color = DAY_CSS[di]
            html = []
            html.append(f'<div style="flex:1;min-width:0">')
            html.append(f'<div style="background:{col_color};color:#fff;font-size:8.5px;'
                        f'font-weight:700;padding:4px 8px;border-radius:4px 4px 0 0;'
                        f'margin-bottom:2px">Day {di+1} · {dl} · {_t:,} UV</div>')
            # Gender
            html.append('<table class="dt" style="margin-bottom:4px"><thead>'
                        '<tr><th>Gender</th><th>Unique Visitors</th><th>%</th></tr></thead><tbody>')
            for _i,(_lbl,_n) in enumerate([('Male',_m),('Female',_f)]):
                html.append(f'<tr class="{"alt" if _i%2==0 else ""}">'
                            f'<td>{_lbl}</td><td class="mono">{_n:,}</td>'
                            f'<td class="mono">{pct(_n,_t)}</td></tr>')
            html.append(f'<tr class="total"><td>Total</td><td class="mono">{_t:,}</td>'
                        f'<td class="mono">100%</td></tr>')
            html.append('</tbody></table>')
            # Age
            html.append('<table class="dt"><thead>'
                        '<tr><th>Age</th><th>Unique Visitors</th><th>%</th></tr></thead><tbody>')
            for _i,(_full,_lbl) in enumerate(zip(AGE_ORDER2,AGE_LBL2)):
                _n = int(_age.get(_full,0))
                html.append(f'<tr class="{"alt" if _i%2==0 else ""}">'
                            f'<td>{_lbl}</td><td class="mono">{_n:,}</td>'
                            f'<td class="mono">{pct(_n,_ta)}</td></tr>')
            html.append(f'<tr class="total"><td>Total</td><td class="mono">{_ta:,}</td>'
                        f'<td class="mono">100%</td></tr>')
            html.append('</tbody></table>')
            html.append('</div>')
            return html

        # Wrap all days in a flex row
        n_days_avail = len(EVENT_DATES_)
        col_width = max(160, 182 // max(n_days_avail, 1))  # mm per col
        s.append(f'<div style="display:flex;gap:8px;align-items:flex-start">')
        for di, (ds, dl) in enumerate(zip(EVENT_DATES_, EVENT_LABELS_)):
            ent_d = ent[ent['Date']==ds]
            uv_d  = ent_d.drop_duplicates('BodyID')
            if len(uv_d) == 0:
                s.append(f'<div style="flex:1;min-width:0">'
                         f'<div style="background:{DAY_CSS[di]};color:#fff;font-size:8.5px;'
                         f'font-weight:700;padding:4px 8px;border-radius:4px 4px 0 0">'
                         f'Day {di+1} · {dl}</div>'
                         f'<p style="color:#999;font-size:9px;padding:8px">No data</p></div>')
            else:
                s.extend(compact_day_col(uv_d, di, dl))
        s.append('</div>')  # end flex row

    sections.append(('\n'.join(s), 5))

    # ─────────────────────────────────────────────────────────────
    # ─────────────────────────────────────────────────────────────
    # SECTION 6: ENTRANCE & PASSERBY GATE BREAKDOWN
    # ─────────────────────────────────────────────────────────────
    s = []
    s.append(sec_hdr('🚪', f'{VENUE_TYPE} Entrance Gate & Passerby Gate Breakdown', 'Per gate traffic summary'))

    _p6 = PROFILE_CONFIG.get(EVENT_PROFILE, PROFILE_CONFIG['full'])
    _show_psb = _p6['has_psb'] and SHOW_PASSERBY

    if _show_psb:
        s.append('<div class="two-col">')

    # ENT gates
    s.append('<div>')
    s.append(f'<div class="col-title">{VENUE_TYPE} Entrance Gates (ENT)</div>')
    s.append('<table class="dt"><thead><tr><th>Gate</th><th>Visitors</th><th>Unique</th><th>Share %</th></tr></thead><tbody>')
    ent_tot = len(ent)
    for i,(loc,g) in enumerate(ent.groupby('Location')):
        n=len(g); uv=uv_count(g)
        s.append(f'<tr class="{"alt" if i%2==0 else ""}"><td>{loc}</td>'
                 f'<td class="mono">{n:,}</td><td class="mono">{uv:,}</td>'
                 f'<td class="mono">{pct(n,ent_tot)}</td></tr>')
    s.append(f'<tr class="total"><td>TOTAL</td><td class="mono">{ent_tot:,}</td>'
             f'<td class="mono">{uv_count(ent):,}</td><td class="mono">100%</td></tr>')
    s.append('</tbody></table></div>')

    # PSB gates — standard/full only
    if _show_psb:
        s.append('<div>')
        s.append('<div class="col-title">Passerby Gates (PSB)</div>')
        s.append('<table class="dt"><thead><tr><th>Gate</th><th>Passersby</th><th>Share %</th></tr></thead><tbody>')
        psb_tot = len(psb)
        for i,(loc,g) in enumerate(psb.groupby('Location')):
            n=len(g)
            s.append(f'<tr class="{"alt" if i%2==0 else ""}"><td>{loc}</td>'
                     f'<td class="mono">{n:,}</td><td class="mono">{pct(n,psb_tot)}</td></tr>')
        s.append(f'<tr class="total"><td>TOTAL</td><td class="mono">{psb_tot:,}</td><td class="mono">100%</td></tr>')
        s.append('</tbody></table></div>')
        s.append('</div>')  # close PSB inner div
        s.append('</div>')  # close two-col

    sections.append(('\n'.join(s), 6))

    # ─────────────────────────────────────────────────────────────
    # SECTION 7: OVERALL HEATMAP — combined V/UV/PSB side-by-side
    # ─────────────────────────────────────────────────────────────
    s = []
    s.append(sec_hdr('🌡️', 'Overall Hourly Heatmap — Booth Traffic',
                     'Visitors (V) / Unique Visitors (UV) / Passersby (PSB) · แสดงรายวัน · สีแต่ละ metric ใช้ scale ของตัวเอง'))

    # Build day_data: one entry per day with all 3 metrics
    all_vals_v = []; all_vals_uv = []; all_vals_p = []
    day_data_hm = []
    for di, ds in enumerate(EVENT_DATES_):
        e_d = ent[ent['Date']==ds]; p_d = psb[psb['Date']==ds]
        v_h  = h_vals(e_d)
        uv_h = h_vals(_non_staff(e_d), 'nunique')
        p_h  = [h_vals(p_d)[i] + v_h[i] for i in range(len(HOURS))]
        day_data_hm.append({
            'label':    DAY_SHORT[di],
            'color':    DAY_CSS[di],
            'v':        v_h,
            'uv':       uv_h,
            'p':        p_h,
            'v_global': uv_count(e_d),  # true global unique for total row (staff-aware)
        })
        all_vals_v  += v_h; all_vals_uv += uv_h; all_vals_p += p_h

    sn = trim_n([d['v'] for d in day_data_hm] +
                [d['uv'] for d in day_data_hm] +
                [d['p']  for d in day_data_hm])

    # Overall KPI strip
    gv  = len(ent); guv = uv_count(ent); gp = len(psb)+len(ent)
    gpv = max(all_vals_v);  gpvh = HOUR_LBLS[all_vals_v.index(gpv)  % len(HOUR_LBLS)] if gpv>0  else '--'
    gpuv= max(all_vals_uv); gpuvh= HOUR_LBLS[all_vals_uv.index(gpuv)% len(HOUR_LBLS)] if gpuv>0 else '--'
    gpp = max(all_vals_p);  gpph = HOUR_LBLS[all_vals_p.index(gpp)  % len(HOUR_LBLS)] if gpp>0  else '--'
    # Staff-exclusion footnote: only set when filter is active and staff exist
    if EXCLUDE_STAFF and 'CustomerType' in ent.columns:
        # staff entrance 'in' events removed from visitor count.
        # NOTE: do NOT report unique BodyID as a "staff count" — ReID groups
        # staff by uniform template (3 templates here), not by person.
        _staff_entries = len(ent[ent['CustomerType']=='Staff'])
        _staff_note = f' · Excluded {_staff_entries:,} staff entries' if _staff_entries > 0 else ''
    else:
        _staff_note = ''
    s.append(kpi_cards(
        (f'Total Visitors ({len(EVENT_DATES)} Days)',       fmt(gv),  f'Peak: {gpvh} ({gpv:,})',  '#005B9A'),
        (f'Total Unique Visitors ({len(EVENT_DATES)} Days)',fmt(guv), f'Peak: {gpuvh} ({gpuv:,}){_staff_note}','#1A7A45'),
        *([(f'Total Passersby ({len(EVENT_DATES)} Days)', fmt(gp), f'Peak: {gpph} ({gpp:,})', '#B06000')]
          if (PROFILE_CONFIG.get(EVENT_PROFILE, PROFILE_CONFIG['full'])['has_psb'] and SHOW_PASSERBY) else []),
    ))

    s.append(heatmap_combined(day_data_hm, sn))

    sections.append(('\n'.join(s), 7))

    # ─────────────────────────────────────────────────────────────
    # SECTION 8: HOURLY TRAFFIC TABLE (moved after Heatmap)
    # ─────────────────────────────────────────────────────────────
    s = []
    # Hourly table for each day
    _s8_show_psb = PROFILE_CONFIG.get(EVENT_PROFILE, PROFILE_CONFIG['full'])['has_psb'] and SHOW_PASSERBY
    _s8_title = ('Hourly Traffic — Visitors / Unique / Passersby' if _s8_show_psb
                 else 'Hourly Traffic — Visitors / Unique')
    s.append(sec_hdr('🕐', _s8_title, 'แยกตามวัน'))
    for di, ds in enumerate(EVENT_DATES_):
        e_d = ent[ent['Date']==ds]
        p_d = psb[psb['Date']==ds]
        has = len(e_d) > 0
        acts = ACTIVITIES_.get(ds, [])

        # Insert page-break before days 2+ if many days (prevents overflow)
        if di > 0 and len(EVENT_DATES_) > 2:
            pass  # page-break disabled
        s.append(day_hdr(f'Day {di+1}  ·  {EVENT_LABELS_[di]}', DAY_CSS[di]))
        if not has:
            s.append(no_data_block()); continue

        rows_h = []; peak_vv = 0
        for h in HOURS:
            eh = e_d[e_d['Hour']==h]; ph = p_d[p_d['Hour']==h]
            v=len(eh); uv=int(uv_count(eh)); p=len(ph)+v
            note_hits = [a[2] for a in acts
                         if int(a[0][:2])*60+int(a[0][3:]) < h*60+60
                         and int(a[1][:2])*60+int(a[1][3:]) > h*60]
            note = ' / '.join(note_hits) if note_hits else ''
            rows_h.append((h, v, uv, p, note))
            peak_vv = max(peak_vv, v)

        show = max((i for i,(h,v,*_) in enumerate(rows_h) if v>0), default=9) + 1
        # Build table header conditionally
        _psb_th = '<th>Passersby</th>' if _s8_show_psb else ''
        s.append(f'<table class="dt"><thead><tr>'
                 f'<th>Hour</th><th>Visitors</th><th>Unique</th>{_psb_th}<th>Activity</th>'
                 f'</tr></thead><tbody>')
        for i,(h,v,uv,p,note) in enumerate(rows_h[:show]):
            row_cls = 'peak' if v==peak_vv and peak_vv>0 else ('hi' if note else ('alt' if i%2==0 else ''))
            _psb_td = f'<td class="mono">{p:,}</td>' if _s8_show_psb else ''
            s.append(f'<tr class="{row_cls}"><td class="mono">{h:02d}:00</td>'
                     f'<td class="mono">{v:,}</td><td class="mono">{uv:,}</td>'
                     f'{_psb_td}<td class="act">{note}</td></tr>')
        tot_v=sum(r[1] for r in rows_h[:show])
        tot_p=sum(r[3] for r in rows_h[:show])
        tot_uv = uv_count(e_d)
        _psb_tot_td = f'<td class="mono">{tot_p:,}</td>' if _s8_show_psb else ''
        s.append(f'<tr class="total"><td>TOTAL</td><td class="mono">{tot_v:,}</td>'
                 f'<td class="mono">{tot_uv:,}</td>{_psb_tot_td}<td></td></tr>')
        s.append('</tbody></table>')
        s.append('<p class="tbl-note">&#9432;&nbsp; Unique per hour = distinct visitors in that hour only. Total Unique = global distinct BodyID count. Sum of hourly unique will be higher because the same visitor may enter across multiple hours.</p>')

    sections.append(('\n'.join(s), 8))

    # ══════════════════════════════════════════════════════════════
    # ASSEMBLE HTML
    # ══════════════════════════════════════════════════════════════
    CSS = """
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Sans:ital,wght@0,300;0,400;0,500;0,600;0,700;1,400&family=IBM+Plex+Mono:wght@400;500&display=swap');
*,*::before,*::after{box-sizing:border-box;margin:0;padding:0}
:root{
  --nav:#003865;--blu:#005B9A;--bg:#F2F6FA;--card:#fff;--bdr:#D0DCE8;
  --txt:#1A2A3A;--mut:#6B8299;--alt:#F7FAFD;--hi:#FFF3CD;--tot:#D6E4F0;
  --h0:#EFF6FB;--h1:#C6DBEF;--h2:#6BAED6;--h3:#2171B5;--h4:#084594;
}
body{font-family:'IBM Plex Sans',sans-serif;background:var(--bg);color:var(--txt);
     font-size:11px;line-height:1.45;padding:14px}

/* ── Continuous document — one flat flow ───────────────── */
body{max-width:1100px;margin:0 auto}
.doc-section{display:block}
.ph{margin-bottom:11px}

/* Footer at bottom of doc on screen */
.pfooter{text-align:center;font-size:8px;color:var(--mut);
         margin-top:14px;padding-top:8px;
         border-top:1px solid var(--bdr);
         word-break:break-word;line-height:1.5}

/* No-op helpers */
.page-break-before,.page-break-after,.page,.pb{display:block;height:0;margin:0;padding:0}

/* ── Print: force A4 pages with proper breaks ──────────── */
@media print{
  @page{size:A4 portrait;margin:10mm 10mm 12mm 10mm}
  body{background:#fff;padding:0;font-size:9.5px;max-width:100%;margin:0}
  .no-print{display:none!important}

  /* Disable any page-break helpers */
  .page-break-before,.page-break-after,.page,.pb{
    page-break-before:auto!important;page-break-after:auto!important;
    break-before:auto!important;break-after:auto!important;
    display:block!important;height:0!important;margin:0!important;padding:0!important
  }

  /* Footer as a normal flow element at the very end of the document.
     It is rendered only once (is_last), so static positioning places it
     after the final content instead of floating on top of it. */
  .pfooter{position:static;
           text-align:center;font-size:7px;color:#666;
           margin-top:8mm;padding-top:2mm;
           border-top:1px solid #ccc;
           background:#fff;
           page-break-inside:avoid;break-inside:avoid;
           word-break:break-word;line-height:1.4}

  /* Prevent SMALL atomic blocks from splitting — let containers flow */
  /* Small/atomic items: keep together */
  .kpi,.ins-card,.funnel-step,.act-row,.rank-row,
  .sponsor-card,.exec-kpi,.exec-hl,.exec-score,
  .dt thead,.hm thead{
    page-break-inside:avoid;break-inside:avoid
  }
  /* Repeat table headers at the top of every printed page */
  .dt thead,.hm thead{
    display:table-header-group
  }
  /* Table rows: keep together (all report tables use .dt or .hm) */
  .dt tbody tr,.hm tbody tr,.dt tr,.hm tr{
    page-break-inside:avoid;break-inside:avoid
  }
  /* Section headers should stay with at least the next block */
  .shdr{page-break-after:avoid;break-after:avoid}
  /* Large containers — let them break naturally to fill pages */
  .exec-wrap,.act-corr-wrap,.rank-wrap,.sponsor-grid,
  .funnel-wrap,.ins-panel,.krow{
    page-break-inside:auto;break-inside:auto
  }
}

/* ── Page header ────────────────────────────────────────── */
.ph{background:var(--nav);border-radius:6px;padding:9px 13px;
    margin-bottom:11px;display:flex;align-items:center;justify-content:space-between;gap:8px}
.ph h1{font-size:12.5px;font-weight:700;color:#fff;letter-spacing:-.2px}
.ph p{font-size:8px;color:#B8D4F0;margin-top:1px}
.phb{background:rgba(255,255,255,.12);border:1px solid rgba(255,255,255,.2);
     color:#fff;font-size:8px;font-weight:600;padding:3px 9px;
     border-radius:14px;white-space:nowrap;flex-shrink:0}

/* ── Section header ─────────────────────────────────────── */
.shdr{display:flex;align-items:flex-start;gap:7px;background:#E8F0F8;page-break-inside:avoid;break-inside:avoid;
      border-left:3px solid var(--nav);border-radius:5px;
      padding:6px 9px;margin-bottom:8px}
.sico{font-size:13px;line-height:1;margin-top:1px;flex-shrink:0}
.stit{font-size:11px;font-weight:700;color:var(--nav);display:block}
.ssub{font-size:8px;color:var(--mut);display:block;margin-top:1px}

/* ── Section divider ────────────────────────────────────── */
.sdiv{border:none;border-top:1px solid var(--bdr);margin:10px 0}

/* ── KPI row ────────────────────────────────────────────── */
.krow{display:flex;gap:6px;margin-bottom:7px;flex-wrap:wrap;page-break-inside:avoid;break-inside:avoid}
.kpi{background:#F4F7FB;border:1px solid var(--bdr);border-radius:6px;
     padding:5px 9px;flex:1;min-width:80px;border-left:3px solid var(--blu)}
.kv{font-size:15px;font-weight:700;color:var(--nav);
    font-family:'IBM Plex Mono',monospace;letter-spacing:-0.5px;line-height:1.1}
.kl{font-size:7.5px;color:var(--mut);font-weight:600;
    text-transform:uppercase;letter-spacing:.3px;margin-top:2px}
.ks{font-size:7.5px;color:var(--blu);font-weight:500;margin-top:1px}

/* ── Data table ─────────────────────────────────────────── */
.dt{width:100%;border-collapse:collapse;font-size:9.5px;margin-bottom:8px}
.dt thead tr{background:var(--nav)}
.dt thead th{color:#fff;font-weight:600;padding:5px 6px;text-align:center;
             border-right:1px solid #336699;font-size:9px}
.dt thead th:first-child{text-align:left}
.dt tbody tr td{padding:3px 6px;border-bottom:1px solid var(--bdr);
                border-right:1px solid var(--bdr);vertical-align:middle}
.dt tbody tr td:first-child{text-align:left}
.dt tbody tr td:not(:first-child){text-align:center}
.dt tbody tr.alt td{background:var(--alt)}
.dt tbody tr.hi td{background:var(--hi)}
.dt tbody tr.peak td:nth-child(2){background:var(--blu);color:#fff;font-weight:700}
.dt tbody tr.total td{background:var(--tot);color:var(--nav);font-weight:700;
                      border-top:2px solid var(--bdr)}
.dt tbody tr.day-sep td{padding:3px 8px}
.mono{font-family:'IBM Plex Mono',monospace;font-size:9px}
.act{font-size:8px;color:#8B4500;font-weight:500}
.act-name{font-weight:600;color:var(--nav)}
.col-title{font-size:9.5px;font-weight:700;color:var(--nav);
           padding:4px 0 5px;margin-bottom:4px;
           border-bottom:2px solid var(--nav)}

/* ── Two column layout ──────────────────────────────────── */
.two-col{display:grid;grid-template-columns:1fr 1fr;gap:12px;margin-bottom:6px}

/* ── Heatmap table ──────────────────────────────────────── */
.hm-comb{width:100%;border-collapse:separate;border-spacing:2px;font-family:'IBM Plex Mono',monospace;margin-bottom:4px}
.hc-hour{width:40px;font-size:7.5px;font-weight:600;color:var(--mut);text-align:center;padding:2px 1px}
.hc-hour-lbl{font-size:8px;font-weight:600;color:var(--mut);text-align:center;vertical-align:middle;padding:0 2px;font-family:'IBM Plex Mono',monospace}
.hc-day{font-size:8px;font-weight:700;color:#fff;text-align:center;padding:3px 2px;border-radius:3px 3px 0 0}
.hc-sub{font-size:7.5px;font-weight:700;text-align:center;padding:2px 1px;min-width:28px}
.hcc{height:22px;text-align:center;vertical-align:middle;font-size:8px;font-weight:600;border-radius:2px;min-width:28px}
.hcc-tot{background:var(--tot);color:var(--nav);font-weight:700;font-size:8px;text-align:center;border-radius:2px;padding:0 2px}
.hc-tot td{padding:2px 1px}
.hm{width:100%;border-collapse:separate;border-spacing:2px;
    table-layout:fixed;font-family:'IBM Plex Mono',monospace;margin-bottom:4px}
.zth{width:80px}
.hm thead th{font-size:7.5px;font-weight:600;color:var(--mut);
             text-align:center;padding:2px 1px}
.tot-th{font-size:7.5px;font-weight:700;color:var(--nav);text-align:center;width:38px}
.zl{font-size:8px;font-weight:600;color:var(--txt);
    padding:2px 4px 2px 0;white-space:normal;word-break:break-word;
    vertical-align:middle;line-height:1.3}
.dot{display:inline-block;width:6px;height:6px;border-radius:50%;
     margin-right:3px;vertical-align:middle}
.cell{height:auto;min-height:24px;text-align:center;vertical-align:middle;
      font-size:8.5px;font-weight:600;border-radius:3px;padding:3px 1px}
.e {background:#F0F4F8;color:transparent}
.h0{background:var(--h0);color:#9BB5CC}
.h1{background:var(--h1);color:#4A7FA5}
.h2{background:var(--h2);color:#fff}
.h3{background:var(--h3);color:#fff}
.h4{background:var(--h4);color:#fff;font-weight:700}
.tot-cell{background:var(--tot);color:var(--nav);font-weight:700;
          font-size:8.5px;text-align:center;border-radius:3px;
          padding:0 3px;white-space:nowrap}
.leg{display:flex;align-items:center;gap:3px;margin-bottom:6px}
.ll{font-size:7.5px;color:var(--mut)}
.ls{display:inline-block;width:16px;height:9px;border-radius:2px}
.ls.h0{background:var(--h0)}.ls.h1{background:var(--h1)}
.ls.h2{background:var(--h2)}.ls.h3{background:var(--h3)}.ls.h4{background:var(--h4)}

/* ── Day header badge ───────────────────────────────────── */
.dhdr{color:#fff;font-size:9.5px;font-weight:700;padding:3px 10px;
      border-radius:14px;display:inline-block;margin-bottom:6px}

/* ── Metric block (heatmap) ─────────────────────────────── */
.metric-block{margin-bottom:10px}
.metric-block+.metric-block{border-top:1px dashed var(--bdr);padding-top:8px}
.mhd{padding:3px 8px;margin-bottom:6px;background:#F4F7FB;
     border-radius:3px;font-size:9.5px;font-weight:700}

/* ── No data ────────────────────────────────────────────── */
.tbl-note{font-size:7.5px;color:var(--mut);font-style:italic;margin:3px 0 6px;padding:2px 6px;border-left:2px solid var(--bdr)}
.nodata{padding:10px;text-align:center;color:var(--mut);
        font-size:10px;background:#F4F7FB;border-radius:5px;margin-bottom:6px}

/* ── Page footer ────────────────────────────────────────── */
.pfooter{text-align:center;font-size:6.5px;color:var(--mut);
         margin-top:8px;padding-top:5px;border-top:1px solid var(--bdr);
         word-break:break-word;line-height:1.4;display:block}

/* ── Insight Cards ─────────────────────────────────────── */
.ins-panel{display:flex;flex-direction:column;gap:5px;margin-bottom:10px}
.ins-card{display:flex;gap:8px;padding:6px 9px;border-radius:4px;
          page-break-inside:avoid;break-inside:avoid}
.ins-icon{font-size:14px;flex-shrink:0;line-height:1.3}
.ins-body{display:flex;flex-direction:column;gap:1px;min-width:0}
.ins-val{font-size:13px;font-weight:700;color:var(--nav)}
.ins-en{font-size:8.5px;color:var(--txt);line-height:1.4}
.ins-th{font-size:8px;color:var(--mut);line-height:1.4;font-style:italic}

/* ── Funnel ─────────────────────────────────────────────── */
.funnel-wrap{display:flex;flex-direction:column;gap:6px;margin-bottom:10px}
.funnel-step{display:flex;align-items:center;gap:8px}
.funnel-label{width:140px;flex-shrink:0;font-size:8.5px;font-weight:600;color:var(--nav)}
.funnel-bar-wrap{flex:1;display:flex;align-items:center;gap:6px}
.funnel-bar{height:22px;border-radius:3px;display:flex;align-items:center;
            padding:0 8px;min-width:40px;transition:width .3s}
.funnel-n{font-size:8.5px;font-weight:700;color:#fff;white-space:nowrap}
.funnel-conv{font-size:7.5px;color:var(--mut);white-space:nowrap}

/* ── Activity Correlation ─────────────────────────────── */
.act-corr-wrap{display:flex;flex-direction:column;gap:6px;margin-bottom:10px}
.act-row{display:flex;gap:10px;align-items:center;padding:7px 10px;
         background:var(--alt);border-radius:4px;page-break-inside:avoid}
.act-info{width:130px;flex-shrink:0}
.act-name{font-size:9px;font-weight:700;color:var(--nav);line-height:1.2}
.act-time{font-size:7.5px;color:var(--mut);margin-top:1px}
.act-bars{flex:1;display:flex;flex-direction:column;gap:2px;min-width:0}
.act-bar-group{display:flex;align-items:center;gap:6px}
.act-lbl{font-size:7.5px;color:var(--mut);width:42px;flex-shrink:0;text-align:right}
.act-lbl-active{color:var(--nav);font-weight:700}
.act-bar-wrap{flex:1;min-width:0}
.act-bar{height:14px;border-radius:2px;display:flex;align-items:center;padding:0 5px;min-width:28px}
.act-bar-during{background:linear-gradient(90deg,#005B9A 0%,#1A7A45 100%)}
.act-bar span{font-size:7.5px;font-weight:700;color:#fff;white-space:nowrap}
.act-bar[style*="#B8D4F0"] span{color:var(--nav)}
.act-lift{display:flex;flex-direction:column;align-items:center;width:60px;flex-shrink:0}
.act-lift-icon{font-size:13px;line-height:1}
.act-lift-val{font-size:10px;font-weight:700;margin-top:2px}

/* ── Top Zone Ranking ─────────────────────────────────── */
.rank-wrap{display:flex;flex-direction:column;gap:5px;margin-bottom:10px}
.rank-row{display:flex;gap:8px;align-items:center;padding:6px 9px;
          background:var(--alt);border-radius:4px;border-left:3px solid var(--blu);
          page-break-inside:avoid}
.rank-medal{font-size:16px;width:26px;flex-shrink:0;text-align:center;
            font-weight:700;color:var(--nav)}
.rank-name{width:130px;flex-shrink:0;font-size:9px;font-weight:700;color:var(--nav);
           white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.rank-metrics{flex:1;display:flex;gap:10px;min-width:0;align-items:center}
.rank-metric{flex:1;display:flex;flex-direction:column;gap:1px;min-width:0}
.rank-ml{font-size:7px;color:var(--mut);font-weight:600;text-transform:uppercase}
.rank-bar-wrap{position:relative;height:14px;background:#F0F4F8;border-radius:2px;
               display:flex;align-items:center}
.rank-bar{height:100%;border-radius:2px;transition:width .3s}
.rank-val{position:absolute;right:4px;font-size:7.5px;font-weight:700;color:var(--nav)}
.rank-eng{font-size:12px;font-weight:700;padding-top:2px}

/* ── Sponsor Value Grid ────────────────────────────────── */
.sponsor-grid{display:grid;grid-template-columns:repeat(4,1fr);gap:6px;margin-bottom:10px}
.sponsor-card{background:#fff;border:1px solid var(--bdr);border-radius:4px;
              padding:8px 9px;page-break-inside:avoid;display:flex;flex-direction:column;gap:2px}
.sp-icon{font-size:14px;line-height:1}
.sp-title{font-size:7.5px;color:var(--mut);font-weight:600;text-transform:uppercase;margin-top:3px}
.sp-val{font-size:15px;font-weight:700;line-height:1.1;margin-top:1px}
.sp-sub{font-size:7.5px;color:var(--mut);line-height:1.3;margin-top:2px}

/* ── Executive Summary ──────────────────────────────────── */
.exec-wrap{display:flex;flex-direction:column;gap:10px;margin-bottom:10px}
.exec-kpi-row{display:grid;grid-template-columns:repeat(4,1fr);gap:7px}
.exec-kpi{background:#fff;border:1px solid var(--bdr);border-radius:5px;
          padding:10px 12px;text-align:center;page-break-inside:avoid}
.exec-kpi-title{font-size:8px;color:var(--mut);text-transform:uppercase;
                font-weight:600;letter-spacing:.3px}
.exec-kpi-val{font-size:22px;font-weight:700;line-height:1.1;margin:4px 0}
.exec-kpi-sub{font-size:7.5px;color:var(--mut);line-height:1.3}

.exec-score{display:flex;justify-content:space-between;align-items:center;
            background:var(--alt);border-radius:5px;padding:10px 14px;
            page-break-inside:avoid}
.exec-score-l{display:flex;flex-direction:column}
.exec-score-title{font-size:11px;font-weight:700;color:var(--nav)}
.exec-score-sub{font-size:8.5px;color:var(--mut);margin-top:2px}
.exec-score-r{display:flex;flex-direction:column;align-items:flex-end}
.exec-score-num{font-size:28px;font-weight:700;line-height:1}
.exec-score-lbl{font-size:8.5px;font-weight:600;margin-top:2px;text-transform:uppercase}

.exec-highlights{display:grid;grid-template-columns:repeat(3,1fr);gap:7px}
.exec-hl{background:#fff;border:1px solid var(--bdr);border-left:3px solid var(--blu);
         border-radius:4px;padding:8px 10px;page-break-inside:avoid}
.exec-hl-icon{font-size:16px;line-height:1}
.exec-hl-lbl{font-size:7.5px;color:var(--mut);text-transform:uppercase;
             font-weight:600;margin-top:4px}
.exec-hl-val{font-size:12px;font-weight:700;color:var(--nav);line-height:1.2;margin-top:2px}
.exec-hl-sub{font-size:7.5px;color:var(--mut);line-height:1.3;margin-top:2px}

.exec-insights-title{font-size:10px;font-weight:700;color:var(--nav);
                     margin:4px 0 -2px 0;padding-left:2px}

/* ── Before/During/After Table ────────────────────────── */
.bda-tbl th.bda-before{background:#D6E4F0}
.bda-tbl th.bda-during{background:#005B9A;color:#fff}
.bda-tbl th.bda-after{background:#D6E4F0}

/* ── Print button ───────────────────────────────────────── */

.print-btn{display:block;margin:0 auto 14px;padding:8px 22px;
           background:var(--nav);color:#fff;border:none;border-radius:7px;
           font-size:12px;font-weight:600;cursor:pointer;
           font-family:'IBM Plex Sans',sans-serif}
.print-btn:hover{background:var(--blu)}
"""

    # Group sections into pages (2-3 sections per page to fit A4 portrait)
    # Page assignment:
    # Page 1: Overall KPIs + Hourly Traffic  (sec 1)
    # Page 2: Zone Summary + Zone Heatmap    (sec 2)
    # Page 3: Activity Analytics             (sec 3)
    # Page 4: Dwell Time + Demographics      (sec 4 + 5)
    # Page 5: Gate Breakdown + Overall Heatmap (sec 6 + 7)
    _PG = PAGE_GROUPS  # page groups from config

    page_html = []
    total_pages = len(PAGE_GROUPS)
    for pg_num, group in enumerate(_PG, 1):
        sec_map = {sec_num: sec_content for sec_content, sec_num in sections}
        body = '\n'.join(sec_map[n] for n in group if n in sec_map)
        # Build page titles dynamically from _PG section order
        _sec_label = {
            0: (f'{EVENT_NAME} — Executive Summary',
                f'C-level overview &middot; Top insights &middot; Programme effectiveness'),
            1: (f'{VENUE_TYPE} Level — Overall Dashboard',
                f'Overall Traffic KPIs &middot; Hourly Heatmap'),
            7: (f'{VENUE_TYPE} Level — Overall Dashboard',
                f'Overall Traffic KPIs &middot; Hourly Heatmap'),
            5: (f'{VENUE_TYPE} Level — Visitor Profile',
                'Demographics &middot; Gate Breakdown'),
            6: (f'{VENUE_TYPE} Level — Visitor Profile',
                'Demographics &middot; Gate Breakdown'),
            8: (f'{VENUE_TYPE} Level — Hourly Traffic Breakdown',
                ('Visitors / Unique Visitors / Passersby &middot; แยกตามวัน'
                 if (PROFILE_CONFIG.get(EVENT_PROFILE, PROFILE_CONFIG['full'])['has_psb'] and SHOW_PASSERBY)
                 else 'Visitors / Unique Visitors &middot; แยกตามวัน')),
            3: (f'{VENUE_TYPE} Level — Activity Analytics',
                'Traffic during activities &middot; Activity &times; Zone Matrix'),
            2: ('Zone Level — Zone Traffic', 'Zone Summary &middot; Hourly Zone Heatmap'),
            4: ('Zone Level — Dwell Time Analysis', 'Time spent per zone &middot; In &rarr; Out pairing'),
        }
        page_titles = {pn: _sec_label.get(_grp[0],
                       (f'{EVENT_NAME} — {ORGANIZER} {VENUE_TYPE} Analytics Dashboard', ''))
                       for pn, _grp in enumerate(_PG, 1)}
        pg_title, pg_sub = page_titles.get(pg_num, (f'{EVENT_NAME} \u2014 {ORGANIZER} {VENUE_TYPE} Analytics Dashboard', ''))
        # Flat continuous document — each section just contributes body content
        # Header rendered once (first), footer rendered once (last)
        is_first = (pg_num == 1)
        is_last  = (pg_num == total_pages)
        ph_html = f"""<div class="ph">
    <div>
      <h1>{pg_title}</h1>
      <p>{EVENT_NAME} &nbsp;&middot;&nbsp; {ORGANIZER} {VENUE_TYPE}, {VENUE} &nbsp;&middot;&nbsp; AI People Counting by DITECH</p>
    </div>
  </div>""" if is_first else ""
        pf_html = f"""<div class="pfooter">{SYSTEM_CREDIT} &nbsp;&middot;&nbsp; {EVENT_NAME}{', ' + VENUE if VENUE else ''}</div>""" if is_last else ""
        page_html.append(f"""{ph_html}
{body}
{pf_html}""")

    final = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>{EVENT_NAME} &mdash; Full Analytics Dashboard</title>
<style>{CSS}</style>
</head>
<body>
<button class="print-btn no-print" onclick="window.print()">&#128424; Print / Save as PDF &nbsp; (A4 Portrait)</button>
{''.join(page_html)}
</body>
</html>"""

    import re as _re
    final = _re.sub(r'\\n', '', final)
    final = _re.sub(r'\n{3,}', '\n', final)

    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(final)
    print(f'  Full HTML dashboard saved: {output_path}')



def main():
    global ZONE_TYPES, OUTPUT, DAY_CSS_XL
    global EVENT_NAME, ORGANIZER, VENUE, SYSTEM_CREDIT, CONFIDENTIAL
    global EVENT_DATES, EVENT_LABELS, DAY_COLORS
    global ENTRANCE_GATES, PASSERBY_GATES, ZONE_ORDER, ZONE_ABBREV
    global ACTIVITIES, DWELL_MIN_SEC, DWELL_MAX_SEC
    global ENGAGEMENT_THRESHOLD_SEC, DISPLAY_HOURS_START, DISPLAY_HOURS_END
    global INPUT_FILES, EVENT_PROFILE, PAGE_GROUPS, VENUE_TYPE, SHOW_PASSERBY, SPONSOR_ZONES
    global EXCLUDE_STAFF
    global SHOW_DWELL_BENCHMARK, ZONE_BENCHMARK, ZONE_DESC, ZONE_BENCHMARK_MODE

    # ── Load config (Excel _config sheet or event_config.py) ───
    if _USE_EXCEL_CONFIG:
        _cfg = load_config_from_excel(_RAWDATA_FILE)
        EVENT_NAME     = _cfg.get("event_name", "Event Dashboard")
        ORGANIZER      = _cfg.get("organizer", "")
        VENUE          = _cfg.get("venue", "")
        SYSTEM_CREDIT  = _cfg.get("system_credit", "AI People Counting")
        CONFIDENTIAL   = _cfg.get("confidential", True)
        OUTPUT_XLSX_   = os.path.join(_DIR, _cfg.get("output_xlsx", "Dashboard.xlsx"))
        OUTPUT_HTML_   = os.path.join(_DIR, _cfg.get("output_html", "Dashboard.html"))
        EVENT_DATES    = _cfg["EVENT_DATES"]
        EVENT_LABELS   = _cfg["EVENT_LABELS"]
        DAY_COLORS     = _cfg["DAY_COLORS"]
        ENTRANCE_GATES = _cfg["ENTRANCE_GATES"]
        PASSERBY_GATES = _cfg["PASSERBY_GATES"]
        ZONE_ORDER     = _cfg["ZONE_ORDER"]
        ZONE_ABBREV    = _cfg["ZONE_ABBREV"]
        ACTIVITIES     = _cfg["ACTIVITIES"]
        DWELL_MIN_SEC  = _cfg["DWELL_MIN_SEC"]
        DWELL_MAX_SEC  = _cfg["DWELL_MAX_SEC"]
        ENGAGEMENT_THRESHOLD_SEC = _cfg["ENGAGEMENT_THRESHOLD_SEC"]
        DISPLAY_HOURS_START      = _cfg["DISPLAY_HOURS_START"]
        DISPLAY_HOURS_END        = _cfg["DISPLAY_HOURS_END"]
        EXCLUDE_STAFF            = _cfg.get("EXCLUDE_STAFF", True)
        print(f'  Staff exclusion: {"enabled" if EXCLUDE_STAFF else "disabled"}')
        SHOW_DWELL_BENCHMARK     = _cfg.get("SHOW_DWELL_BENCHMARK", False)
        ZONE_BENCHMARK           = _cfg.get("ZONE_BENCHMARK", {})
        ZONE_DESC                = _cfg.get("ZONE_DESC", {})
        ZONE_BENCHMARK_MODE      = _cfg.get("ZONE_BENCHMARK_MODE", {})
        if SHOW_DWELL_BENCHMARK:
            print(f'  Dwell benchmark table: enabled ({len(ZONE_BENCHMARK)} zone target(s))')
        INPUT_FILES    = [_RAWDATA_FILE]
        # event_type / profile from Section A of _config sheet
        EVENT_PROFILE  = _cfg.get('event_type', 'full').lower().strip()
        if EVENT_PROFILE not in PROFILE_CONFIG:
            print(f'  ⚠  Unknown event_type "{EVENT_PROFILE}" — defaulting to "full"')
            EVENT_PROFILE = 'full'
        VENUE_TYPE     = _cfg.get('venue_type', 'Booth').strip()
        SHOW_PASSERBY  = str(_cfg.get('show_passerby', 'True')).strip().lower() != 'false'
        # sponsor_zones — comma-separated list in config, e.g. 'Main Stage, VIP Lounge'
        _sz = _cfg.get('sponsor_zones', 'Main Stage')
        if _sz and isinstance(_sz, str):
            SPONSOR_ZONES = [z.strip() for z in _sz.split(',') if z.strip()]

    else:
        OUTPUT_XLSX_ = OUTPUT_XLSX
        OUTPUT_HTML_ = OUTPUT_HTML
        # Load optional fields from event_config.py
        try:
            from event_config import VENUE_TYPE as _vt
            VENUE_TYPE = _vt
        except (ImportError, AttributeError): pass
        try:
            from event_config import SHOW_PASSERBY as _sp
            SHOW_PASSERBY = _sp
        except (ImportError, AttributeError): pass
        try:
            from event_config import SPONSOR_ZONES as _sz
            SPONSOR_ZONES = _sz
        except (ImportError, AttributeError): pass

    # Build derived globals from config
    ZONE_TYPES.clear()
    for _z in ZONE_ORDER:     ZONE_TYPES[_z] = "Zone"
    for _g in ENTRANCE_GATES: ZONE_TYPES[_g] = "Entrance"
    for _g in PASSERBY_GATES: ZONE_TYPES[_g] = "Passerby"
    DAY_CSS_XL[:] = [c.lstrip("#") for c in DAY_COLORS]
    OUTPUT = OUTPUT_XLSX_
    # Apply profile — set PAGE_GROUPS from profile
    _prof_now = PROFILE_CONFIG.get(EVENT_PROFILE, PROFILE_CONFIG['full'])
    if not SHOW_PASSERBY and _prof_now['has_psb']:
        print('  ℹ  Passerby hidden (show_passerby=False)')
    _base_groups = _prof_now['page_groups']

    # ── Auto-paginate based on actual data volume ──
    # Count activities across all dates
    _n_activities = sum(len(v) for v in ACTIVITIES.values()) if ACTIVITIES else 0
    _effective_has_psb = _prof_now['has_psb'] and SHOW_PASSERBY
    PAGE_GROUPS = auto_paginate(
        base_groups=_base_groups,
        n_days=len(EVENT_DATES),
        n_gates=len(ENTRANCE_GATES) + (len(PASSERBY_GATES) if _effective_has_psb else 0),
        n_zones=len(ZONE_ORDER),
        n_activities=_n_activities,
        profile=EVENT_PROFILE,
        has_psb=_effective_has_psb,
    )
    print(f'  ℹ  Auto-paginate: {len(_base_groups)} → {len(PAGE_GROUPS)} pages '
          f'(days={len(EVENT_DATES)}, zones={len(ZONE_ORDER)}, activities={_n_activities})')

    print("\n" + "="*46)
    print(f" {EVENT_NAME} — Dashboard Generator v4")
    print("="*46)
    print("\nLoading data...")
    if _USE_EXCEL_CONFIG:
        df = load_rawdata_from_excel(_RAWDATA_FILE, COL_NAMES, ZONE_TYPES, EVENT_DATES)
    else:
        df = load_data()
    print(f'  Total: {len(df):,} records  |  Dates: {sorted(df["Date"].unique())}')

    wb = Workbook()
    wb.remove(wb.active)

    _prof = PROFILE_CONFIG.get(EVENT_PROFILE, PROFILE_CONFIG['full'])
    print(f'\nProfile : {_prof["desc"]}')
    print('Building sheets…')
    build_overall_sheet(wb, df)
    print('  ✓  Overall Summary')

    if os.environ.get('SKIP_XLSX_DAY', '0') != '1':
        for i,(ds,dl) in enumerate(zip(EVENT_DATES, EVENT_LABELS), 1):
            build_day_sheet(wb, df, ds, i, dl)
        print(f'  ✓  Day {i} — {dl}')

    if _prof['has_dwell']:
        build_dwell_sheet(wb, df)
        print('  ✓  Dwell Time Analysis')
    else:
        print('  —  Dwell Time Analysis (skipped)')

    if _prof['has_zones']:
        build_zone_sheet(wb, df)
        print('  ✓  Zone Dashboard')
    else:
        print('  —  Zone Dashboard (skipped)')

    if _prof['has_activity']:
        build_activity_sheet(wb, df)
        print('  ✓  Activity Analytics')
    else:
        print('  —  Activity Analytics (skipped)')

    if os.environ.get('SKIP_RAW', '0') != '1':
        for i,(ds,dl) in enumerate(zip(EVENT_DATES, EVENT_LABELS), 1):
            try:
                build_raw_sheet(wb, df, ds, dl, i)
                print(f'  ✓  Raw Data Day {i}')
            except Exception as e:
                print(f'  ⚠ Raw Data Day {i} skipped: {e}')
    else:
        print('  —  Raw Data sheets skipped (SKIP_RAW=1)')

    # ── Heatmap HTML ──────────────────────────────────────────────
    heatmap_path = OUTPUT_HTML_
    generate_full_html(df, heatmap_path)

    # ── Add Heatmap link sheet to Excel ───────────────────────────
    ws_link = setup_ws(wb, '🌡 Heatmap Report', tab_color='2D7DD2')
    page_header(ws_link, 1,
        f'{EVENT_NAME}  —  Hourly Traffic Heatmap Report',
        f'Interactive HTML report  ·  Visitors / Unique / Passersby per Hour  ·  All {len(EVENT_DATES)} Days',
        f'{EVENT_LABELS[0][:6]} – {EVENT_LABELS[-1][:6]} {EVENT_DATES[-1][-4:]}')

    rh(ws_link, 4, 8); rh(ws_link, 5, 40); rh(ws_link, 6, 20); rh(ws_link, 7, 8)
    ws_link.merge_cells(start_row=5, start_column=2, end_row=5, end_column=11)
    ic = ws_link.cell(5, 2, '🌡  Hourly Traffic Heatmap — Interactive Report')
    ic.font = Font(name='Calibri', size=16, bold=True, color='003865')
    ic.fill = PatternFill('solid', start_color='E5EEF6', fgColor='E5EEF6')
    ic.alignment = Alignment(horizontal='center', vertical='center')

    ws_link.merge_cells(start_row=6, start_column=2, end_row=6, end_column=11)
    sc = ws_link.cell(6, 2,
        f'Open the file: {os.path.basename(heatmap_path)} (same folder as this Excel)')
    sc.font = Font(name='Calibri', size=11, color='6B8299', italic=True)
    sc.fill = PatternFill('solid', start_color='F4F7FB', fgColor='F4F7FB')
    sc.alignment = Alignment(horizontal='center', vertical='center')

    import os as _os
    hm_fname = _os.path.basename(heatmap_path)

    # Instructions table
    sec_hdr(ws_link, 8, 'ℹ️  How to use the Heatmap Report')
    instructions = [
        ('1', 'เปิดไฟล์', f'{hm_fname}  (ใน folder เดียวกัน)'),
        ('2', 'เลือก Tab', f'Overall ({len(EVENT_DATES)} Days) / ' + ' / '.join(f'Day {i+1}' for i in range(len(EVENT_DATES)))),
        ('3', 'เลือก Metric', 'Visitors (ENT in) / Unique Visitors / Passersby'),
        ('4', 'Hover บน Cell', 'ดูตัวเลขละเอียด + ชื่อ Hour'),
        ('5', 'อัปเดตข้อมูล', 'รัน dashboard_engine.py ใหม่เมื่อมีข้อมูลวันใหม่'),
    ]
    tbl_hdr(ws_link, 9, ['Step', 'Action', 'Details'])
    for i, (step, action, detail) in enumerate(instructions):
        tbl_row(ws_link, 10+i, [step, action, detail], alt=(i%2==0))

    footer(ws_link, 16)
    print('  ✓  Heatmap link sheet added')

    wb.save(OUTPUT)
    print(f'\n  Saved → {OUTPUT_XLSX_}')
    print(f'  Saved → {OUTPUT_HTML_}')
    print('  Done ✅\n')

if __name__ == '__main__':
    main()
